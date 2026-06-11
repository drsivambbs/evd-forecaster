import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from scipy.stats import gamma as gamma_dist
from datetime import date, datetime
from io import BytesIO


def _slug(text: str) -> str:
    """Filename-safe slug derived from a free-text scenario label."""
    import re
    s = re.sub(r"[^A-Za-z0-9]+", "_", (text or "").strip()).strip("_")
    return s[:40].lower() or "scenario"


def _try_fig_to_png(fig, width: int = 1100, height: int = 480):
    """Convert a Plotly figure to PNG bytes. Returns None if static export is
    unavailable — the Excel report still generates without the image.

    Guard: plotly >= 6 paired with kaleido < 1.0 is a broken combination whose
    to_image() hangs indefinitely (notably on Windows), so detect it and skip
    fast rather than freezing the whole app."""
    def _major_minor(mod):
        parts = (getattr(mod, "__version__", "0").split(".") + ["0", "0"])[:2]
        try:
            return tuple(int(p) for p in parts)
        except ValueError:
            return (0, 0)
    try:
        import plotly, plotly.io as pio
        try:
            import kaleido
        except Exception:
            return None
        if _major_minor(plotly)[0] >= 6 and _major_minor(kaleido) < (1, 0):
            return None  # known-hanging combo — skip image embedding
        return pio.to_image(fig, format="png", width=width, height=height,
                             scale=2)
    except Exception:
        return None


def _apply_header_style(ws, row: int, cols: int, fill: str, font_color: str = "FFFFFF"):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    fill_obj = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    font_obj = Font(name="Calibri", size=11, bold=True, color=font_color)
    align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill_obj
        cell.font = font_obj
        cell.alignment = align
        cell.border = border
    ws.row_dimensions[row].height = 22


def _write_title_block(ws, title: str, subtitle: str = ""):
    from openpyxl.styles import Font, Alignment
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(row=1, column=1, value=title).font = Font(
        name="Calibri", size=18, bold=True, color="1F4E79")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left",
                                                    vertical="center")
    ws.row_dimensions[1].height = 30
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
        ws.cell(row=2, column=1, value=subtitle).font = Font(
            name="Calibri", size=11, italic=True, color="5B6573")
    ws.row_dimensions[2].height = 18


def _auto_size(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_excel_report() -> bytes:
    """Bundle every input + output currently in session_state into a polished
    multi-sheet .xlsx with Cover, Inputs, per-step sheets (chart + data +
    interpretation), and a final Interpretation sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage

    ss = st.session_state
    scenario_name = ss.get("scenario_name", "EVD Forecaster run")
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    brand = "1F4E79"
    brand_light = "DDE7F0"
    grey_band = "F1F4F8"

    H2 = Font(name="Calibri", size=14, bold=True, color=brand)
    BOLD = Font(name="Calibri", size=11, bold=True)
    fill_brand = PatternFill(start_color=brand, end_color=brand, fill_type="solid")
    fill_brand_light = PatternFill(start_color=brand_light,
                                    end_color=brand_light, fill_type="solid")
    fill_grey = PatternFill(start_color=grey_band, end_color=grey_band,
                             fill_type="solid")
    thin = Side(border_style="thin", color="CCCCCC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

    wb = Workbook()

    # =====================================================================
    # SHEET 1 — Cover / Summary
    # =====================================================================
    ws = wb.active
    ws.title = "Cover"
    _write_title_block(ws, f"EVD Forecaster — {scenario_name}",
                        f"Generated {now}")

    row = 4
    ws.cell(row=row, column=1, value="Section").fill = fill_brand
    ws.cell(row=row, column=1).font = Font(name="Calibri", size=11, bold=True,
                                            color="FFFFFF")
    ws.cell(row=row, column=2, value="Status").fill = fill_brand
    ws.cell(row=row, column=2).font = Font(name="Calibri", size=11, bold=True,
                                            color="FFFFFF")
    ws.cell(row=row, column=3, value="Headline").fill = fill_brand
    ws.cell(row=row, column=3).font = Font(name="Calibri", size=11, bold=True,
                                            color="FFFFFF")
    for c in range(1, 4):
        ws.cell(row=row, column=c).alignment = LEFT
        ws.cell(row=row, column=c).border = BORDER
    ws.row_dimensions[row].height = 22

    series = ss.get("result_series")
    rt_df = ss.get("rt_df")
    scenarios_out = ss.get("fc_scenarios")
    horizon_dates = ss.get("fc_dates")
    eoo_probs = ss.get("eoo_probs")
    eoo_days = ss.get("eoo_days")
    valid_plc = ss.get("eoo_valid_plc")
    # Stable SI key from Step 2 (widget keys are dropped on navigation, which
    # would make this filter miss the rt_df rows in the exported report).
    si_used = float(ss.get("rt_si_mean", 15.3))

    sections = []
    sections.append((
        "Step 1 — Daily incidence",
        "Built" if series is not None and len(series) > 0 else "Pending",
        (f"{len(series)} days of incidence ({pd.to_datetime(series['date']).min():%d %b %Y} → "
         f"{pd.to_datetime(series['date']).max():%d %b %Y})"
         if series is not None and len(series) > 0 else "—")
    ))
    if rt_df is not None and not rt_df.empty:
        prim = rt_df[rt_df["si_mean_used"] == si_used].dropna(subset=["rt_mean"])
        if not prim.empty:
            latest = prim.iloc[-1]
            mean_rt = float(latest["rt_mean"])
            direction = "growing" if mean_rt > 1 else "declining"
            sections.append((
                "Step 2 — R_t estimation",
                "Built",
                f"Latest R_t = {mean_rt:.2f} (95% CrI {float(latest['rt_lower']):.2f}–{float(latest['rt_upper']):.2f}) → outbreak {direction}"
            ))
        else:
            sections.append(("Step 2 — R_t estimation", "Built",
                              "Estimates computed"))
    else:
        sections.append(("Step 2 — R_t estimation", "Pending", "—"))

    if scenarios_out and horizon_dates is not None:
        scen_lbl = {"S1": "Delayed", "S2": "Moderate", "S3": "Strong"}
        lines = []
        for name, data in scenarios_out.items():
            v = data["cum_confirmed"]
            med = v["median"] if isinstance(v, dict) else v
            lines.append(f"{scen_lbl.get(name, name)}: {float(med[-1]):,.0f} cumulative confirmed")
        sections.append(("Step 3 — Forecast", "Built", " · ".join(lines)))
    else:
        sections.append(("Step 3 — Forecast", "Pending", "—"))

    # Step 4 status — "Built" whenever any EOO output exists in session_state,
    # even if only one scenario had a valid projected last case.
    if eoo_probs is not None and eoo_days is not None and len(eoo_probs) > 0:
        thr = ss.get("eoo_threshold", 0.95)
        cross = np.where(eoo_probs >= thr)[0]
        if valid_plc and len(cross) > 0:
            cross_day = int(eoo_days[cross[0]])
            lines = []
            for s_name, plc in valid_plc.items():
                lbl = {"S1": "Delayed", "S2": "Moderate", "S3": "Strong"}.get(s_name, s_name)
                d = (plc + pd.Timedelta(days=cross_day)).strftime("%d %b %Y")
                lines.append(f"{lbl}: P≥{thr:.0%} on {d}")
            sections.append(("Step 4 — End of outbreak", "Built",
                              " · ".join(lines)))
        elif valid_plc:
            sections.append(("Step 4 — End of outbreak", "Built",
                              f"P(extinct) does not reach {thr:.0%} within "
                              f"{int(eoo_days[-1])} days"))
        else:
            sections.append(("Step 4 — End of outbreak", "Built",
                              "Simulation ran; no scenario with valid "
                              "projected last case"))
    else:
        sections.append(("Step 4 — End of outbreak", "Pending",
                          "Run Step 4 (EOO simulation) to populate"))

    row += 1
    for sect, status, headline in sections:
        ws.cell(row=row, column=1, value=sect).font = BOLD
        ws.cell(row=row, column=2, value=status)
        if status == "Built":
            ws.cell(row=row, column=2).font = Font(
                name="Calibri", size=11, bold=True, color="1E8449")
        else:
            ws.cell(row=row, column=2).font = Font(
                name="Calibri", size=11, italic=True, color="B22222")
        ws.cell(row=row, column=3, value=headline)
        for c in range(1, 4):
            ws.cell(row=row, column=c).alignment = WRAP
            ws.cell(row=row, column=c).border = BORDER
            ws.cell(row=row, column=c).fill = (
                fill_brand_light if row % 2 == 0 else fill_grey)
        ws.row_dimensions[row].height = 32
        row += 1

    # Sources / methods reference
    row += 2
    ws.cell(row=row, column=1, value="Methods & references").font = H2
    row += 1
    refs = [
        ("Cori et al. 2013", "Am J Epidemiol 178:1505",
         "https://doi.org/10.1093/aje/kwt133"),
        ("Nouvellet et al. 2018", "Epidemics 22:3",
         "https://doi.org/10.1016/j.epidem.2017.02.012"),
        ("WHO Ebola Response Team 2014", "N Engl J Med 371:1481",
         "https://doi.org/10.1056/NEJMoa1411100"),
        ("Lloyd-Smith et al. 2005", "Nature 438:355",
         "https://doi.org/10.1038/nature04153"),
        ("Wamala et al. 2010", "Emerg Infect Dis 16:1087",
         "https://doi.org/10.3201/eid1607.090536"),
        ("Legrand et al. 2007", "Epidemiol Infect 135:610",
         "https://doi.org/10.1017/S0950268806007217"),
    ]
    for citation, journal, doi in refs:
        ws.cell(row=row, column=1, value=citation).font = BOLD
        ws.cell(row=row, column=2, value=journal)
        ws.cell(row=row, column=3, value=doi).hyperlink = doi
        ws.cell(row=row, column=3).font = Font(name="Calibri", size=11,
                                                color=brand, underline="single")
        row += 1

    _auto_size(ws, {"A": 28, "B": 14, "C": 78})

    # =====================================================================
    # SHEET 2 — Inputs
    # =====================================================================
    ws = wb.create_sheet("Inputs")
    _write_title_block(ws, "Inputs", "Every parameter value with its source")

    row = 4
    headers = ["Parameter", "Value", "Source / DOI"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.fill = fill_brand
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.alignment = LEFT
        cell.border = BORDER
    ws.row_dimensions[row].height = 22

    INPUT_ROWS = [
        ("Step 1 — Data", None, None, None, None),
        ("Scenario name", ss.get("scenario_name", ""), None, "string", None),
        ("Step 2 — R_t estimation", None, None, None, None),
        ("SI mean primary (days)", 15.3, "rt_si_mean", "float",
         "si_mean_primary_src"),
        ("SI SD primary (days)", 9.3, "rt_si_sd", "float",
         "si_sd_primary_src"),
        ("Sensitivity SI mean (days)", 12.0, "si_mean_sens_val", "float",
         "si_mean_sens_src"),
        ("Sensitivity SI SD (days)", 5.0, "si_sd_sens_val", "float",
         "si_sd_sens_src"),
        ("Sliding window (days)", 7, "window_val", "int", "window_src"),
        ("Prior shape", 1.0, "shape_prior_val", "float", "shape_prior_src"),
        ("Prior rate", 0.2, "rate_prior_val", "float", "rate_prior_src"),
        ("R_t selected for forecast", None, "selected_rt", "float", None),
        ("R_t basis", "", "selected_rt_basis", "string", None),
        ("Step 3 — Forecast", None, None, None, None),
        ("CFR (fraction)", 0.34, "fc_cfr", "float", "cfr_src"),
        ("Onset-to-death lag (days)", 10, "lag_val", "int", "lag_src"),
        ("Forecast horizon (days)", 365, "forecast_horizon", "int", None),
        ("Forecast start date", "", "forecast_start_date", "string", None),
        ("S1 target R_t", 0.9, "S1_target", "float", None),
        ("S1 days to target", 180, "S1_days", "int", None),
        ("S2 target R_t", 0.9, "S2_target", "float", None),
        ("S2 days to target", 90, "S2_days", "int", None),
        ("S3 target R_t", 0.6, "S3_target", "float", None),
        ("S3 days to target", 30, "S3_days", "int", None),
        ("Step 4 — End of outbreak", None, None, None, None),
        ("EOO simulations", 1000, "eoo_n_sim", "int", None),
        ("EOO days range", 180, "eoo_max_days", "int", None),
        ("EOO declaration threshold", 0.95, "eoo_threshold", "float", None),
        ("EOO offspring dispersion k", 0.18, "eoo_k_disp", "float", None),
    ]
    row += 1
    for label, default, key, kind, src_key in INPUT_ROWS:
        if kind is None:
            ws.cell(row=row, column=1, value=label).font = Font(
                name="Calibri", size=11, bold=True, color=brand)
            for c in range(1, 4):
                ws.cell(row=row, column=c).fill = fill_brand_light
                ws.cell(row=row, column=c).border = BORDER
            ws.row_dimensions[row].height = 20
            row += 1
            continue
        val = ss.get(key, default) if key else default
        src = ss.get(src_key, "") if src_key else ""
        if val is None:
            val = ""
        ws.cell(row=row, column=1, value=label).font = Font(name="Calibri",
                                                              size=11)
        ws.cell(row=row, column=2, value=val).alignment = LEFT
        ws.cell(row=row, column=3, value=src).alignment = LEFT
        if src:
            ws.cell(row=row, column=3).hyperlink = src
            ws.cell(row=row, column=3).font = Font(name="Calibri", size=10,
                                                    color=brand,
                                                    underline="single")
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = BORDER
        row += 1
    _auto_size(ws, {"A": 36, "B": 22, "C": 60})

    # =====================================================================
    # Helper: append a data table from a pandas DataFrame onto a sheet
    # =====================================================================
    def _append_table(ws, df, start_row: int, header_fill: str = brand):
        for j, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=start_row, column=j, value=str(col))
            cell.fill = PatternFill(start_color=header_fill,
                                     end_color=header_fill, fill_type="solid")
            cell.font = Font(name="Calibri", size=10, bold=True,
                              color="FFFFFF")
            cell.alignment = LEFT
            cell.border = BORDER
        ws.row_dimensions[start_row].height = 20
        r = start_row + 1
        for _, row_data in df.iterrows():
            for j, col in enumerate(df.columns, start=1):
                v = row_data[col]
                if isinstance(v, (np.floating, float)):
                    cell = ws.cell(row=r, column=j, value=float(v))
                    cell.number_format = "#,##0.00"
                elif isinstance(v, (np.integer, int)):
                    cell = ws.cell(row=r, column=j, value=int(v))
                    cell.number_format = "#,##0"
                else:
                    cell = ws.cell(row=r, column=j, value=str(v))
                cell.border = BORDER
                cell.alignment = LEFT
                if (r - start_row) % 2 == 0:
                    cell.fill = fill_grey
            r += 1
        return r

    def _embed_png(ws, png_bytes, anchor: str, w_px: int = 880):
        if png_bytes is None:
            return
        try:
            img = XLImage(BytesIO(png_bytes))
            img.width = w_px
            img.height = int(w_px * 0.44)
            ws.add_image(img, anchor)
        except Exception:
            pass

    def _interpretation_block(ws, row: int, text: str):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        c = ws.cell(row=row, column=1, value="Interpretation")
        c.font = Font(name="Calibri", size=11, bold=True, color=brand)
        c.fill = fill_brand_light
        c.alignment = LEFT
        ws.row_dimensions[row].height = 20
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row,
                       end_column=10)
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(name="Calibri", size=11)
        c.alignment = WRAP
        c.fill = fill_grey
        ws.row_dimensions[row].height = 60
        return row + 2

    # =====================================================================
    # SHEET 3 — Step 1: Daily incidence
    # =====================================================================
    if series is not None and len(series) > 0:
        ws = wb.create_sheet("Step 1 — Daily incidence")
        _write_title_block(ws, "Step 1 — Daily incidence builder",
                            "From cumulative DON snapshots → daily series")

        # Embed daily-new chart (read directly from session_state)
        _fig_daily = ss.get("chart_daily")
        if _fig_daily is not None:
            _embed_png(ws, _try_fig_to_png(_fig_daily), "A4")

        row_data_start = 30
        ws.cell(row=row_data_start - 1, column=1,
                 value="Daily incidence — model-ready table").font = H2
        di = series.copy()
        di["date"] = pd.to_datetime(di["date"]).dt.strftime("%Y-%m-%d")
        end_row = _append_table(ws, di, row_data_start)

        interp_text = (
            f"The daily incidence series spans {len(series)} days. "
            f"Total new confirmed cases: {float(series['new_confirmed'].sum()):,.0f}. "
            f"Total suspected: {float(series['new_suspected'].sum()):,.0f}. "
            f"Total deaths: {float(series['new_deaths'].sum()):,.0f}. "
            "Daily counts between cumulative snapshots are linearly interpolated and "
            "first-differenced; small non-integer values are an artefact of "
            "interpolation, not observed data."
        )
        if (ss.get("cfr_active")
                and "new_cfr_estimated" in series.columns):
            interp_text += (
                f" CFR backcalculation applied (Imperial / Garske et al.): "
                f"CFR = {ss.get('cfr_pct_active', '?')}%, "
                f"role = {ss.get('cfr_role_active', 'Total cases')}. "
                f"Estimated true total cases "
                f"= {float(series['new_cfr_estimated'].sum()):,.0f} "
                f"(phase-bias multiplier (1+r/β)^α applied to deaths/CFR)."
            )
        _interpretation_block(ws, end_row + 1, interp_text)

        # Column widths
        for col_letter, w in {"A": 14, "B": 18, "C": 18, "D": 18, "E": 26,
                                "F": 22, "G": 22, "H": 22, "I": 40}.items():
            ws.column_dimensions[col_letter].width = w

    # =====================================================================
    # SHEET 4 — Step 2: R_t estimation
    # =====================================================================
    if rt_df is not None and not rt_df.empty:
        ws = wb.create_sheet("Step 2 — R_t")
        _write_title_block(ws, "Step 2 — R_t estimation",
                            f"Cori 2013 sliding-window, SI Gamma({ss.get('rt_si_mean', 15.3)}, {ss.get('rt_si_sd', 9.3)})")

        # Embed R_t chart (from session_state — set when Step 2 rendered)
        _fig_rt = ss.get("chart_rt")
        if _fig_rt is not None:
            _embed_png(ws, _try_fig_to_png(_fig_rt, width=1100, height=520),
                        "A4")

        # Headline R_t banner
        prim = rt_df[rt_df["si_mean_used"] == si_used].dropna(subset=["rt_mean"])
        if not prim.empty:
            latest = prim.iloc[-1]
            mean_rt = float(latest["rt_mean"])
            direction = "growing" if mean_rt > 1 else "declining"
            colour = "B22222" if mean_rt > 1 else "1E8449"
            ws.cell(row=30, column=1,
                     value=f"Latest R_t = {mean_rt:.2f}  (95% CrI {float(latest['rt_lower']):.2f}–{float(latest['rt_upper']):.2f}) — {direction}")
            ws.cell(row=30, column=1).font = Font(name="Calibri", size=13,
                                                    bold=True, color=colour)

        row_data_start = 33
        ws.cell(row=row_data_start - 1, column=1,
                 value="R_t estimates — full table").font = H2
        rt_out = rt_df.copy()
        rt_out["date"] = pd.to_datetime(rt_out["date"]).dt.strftime("%Y-%m-%d")
        for c in ["rt_mean", "rt_lower", "rt_upper", "shape_post", "rate_post"]:
            if c in rt_out.columns:
                rt_out[c] = rt_out[c].round(3)
        end_row = _append_table(ws, rt_out, row_data_start)

        interp_lines = []
        if not prim.empty:
            interp_lines.append(
                f"The most recent R_t estimate is {mean_rt:.2f} (95% CrI {float(latest['rt_lower']):.2f}–{float(latest['rt_upper']):.2f})."
            )
            if mean_rt > 1:
                interp_lines.append(
                    "R_t > 1 indicates the outbreak is in a growth phase — each case "
                    "is, on average, generating more than one secondary case."
                )
            else:
                interp_lines.append(
                    "R_t < 1 indicates the outbreak is declining — each case is generating "
                    "fewer than one secondary case on average."
                )
        interp_lines.append(
            f"R_t selected for forecast: {float(ss.get('selected_rt', float('nan'))):.2f} "
            f"(basis: {ss.get('selected_rt_basis', 'Latest')})."
        )
        _interpretation_block(ws, end_row + 1, " ".join(interp_lines))

        for col_letter, w in {"A": 14, "B": 14, "C": 14, "D": 14, "E": 18,
                                "F": 18, "G": 18, "H": 14, "I": 16}.items():
            ws.column_dimensions[col_letter].width = w

    # =====================================================================
    # SHEET 5 — Step 3: Forecast
    # =====================================================================
    if scenarios_out and horizon_dates is not None:
        ws = wb.create_sheet("Step 3 — Forecast")
        _write_title_block(ws, "Step 3 — Renewal-equation forecast",
                            "Nouvellet 2018; 3 response scenarios with 90% PI bands")

        # Embed forecast chart (from session_state)
        scen_defaults = {"S1": {"label": "Delayed response"},
                          "S2": {"label": "Moderate response"},
                          "S3": {"label": "Strong response"}}
        _fig_forecast = ss.get("chart_forecast")
        if _fig_forecast is not None:
            _embed_png(ws, _try_fig_to_png(_fig_forecast, width=1200,
                                            height=520), "A4")

        # Per-scenario headline metrics
        row = 32
        ws.cell(row=row, column=1, value="Per-scenario outcomes").font = H2
        row += 1
        for col_i, h in enumerate(
            ["Scenario", "Final confirmed (median)",
             "Final confirmed (90% PI)", "Final deaths (median)",
             "Peak new confirmed", "Peak day"], start=1):
            cell = ws.cell(row=row, column=col_i, value=h)
            cell.fill = fill_brand
            cell.font = Font(name="Calibri", size=10, bold=True,
                              color="FFFFFF")
            cell.alignment = LEFT
            cell.border = BORDER
        row += 1
        for name, data in scenarios_out.items():
            def _m(metric):
                v = data[metric]
                return v["median"] if isinstance(v, dict) else v
            def _lo(metric):
                v = data[metric]
                return v["lower"] if isinstance(v, dict) else v
            def _hi(metric):
                v = data[metric]
                return v["upper"] if isinstance(v, dict) else v
            med_conf = _m("cum_confirmed")
            med_death = _m("cum_deaths")
            med_new = _m("new_confirmed")
            peak_idx = int(np.argmax(med_new))
            cells = [
                scen_defaults[name]["label"],
                float(med_conf[-1]),
                f"{float(_lo('cum_confirmed')[-1]):,.0f} – {float(_hi('cum_confirmed')[-1]):,.0f}",
                float(med_death[-1]),
                float(med_new[peak_idx]),
                horizon_dates[peak_idx].strftime("%d %b %Y"),
            ]
            for col_i, v in enumerate(cells, start=1):
                cell = ws.cell(row=row, column=col_i, value=v)
                if isinstance(v, float):
                    cell.number_format = "#,##0"
                cell.border = BORDER
                cell.alignment = LEFT
                cell.fill = fill_grey if row % 2 == 0 else fill_brand_light
            ws.row_dimensions[row].height = 20
            row += 1

        # Forecast table
        row += 2
        ws.cell(row=row, column=1,
                 value="Forecast table — median + 90% PI").font = H2
        row += 1
        rows = []
        for name, data in scenarios_out.items():
            for i, d in enumerate(horizon_dates):
                def _v(metric, key):
                    v = data[metric]
                    return (float(v[key][i]) if isinstance(v, dict)
                            else float(v[i]))
                rows.append({
                    "date": d.strftime("%Y-%m-%d"),
                    "scenario": scen_defaults[name]["label"],
                    "new_confirmed_med": round(_v("new_confirmed", "median"), 1),
                    "cum_confirmed_med": round(_v("cum_confirmed", "median"), 0),
                    "cum_confirmed_lo": round(_v("cum_confirmed", "lower"), 0),
                    "cum_confirmed_hi": round(_v("cum_confirmed", "upper"), 0),
                    "cum_deaths_med": round(_v("cum_deaths", "median"), 0),
                    "cum_deaths_lo": round(_v("cum_deaths", "lower"), 0),
                    "cum_deaths_hi": round(_v("cum_deaths", "upper"), 0),
                })
        fc_df = pd.DataFrame(rows)
        end_row = _append_table(ws, fc_df, row)

        # Interpretation
        interp = []
        for name, data in scenarios_out.items():
            med = _m("cum_confirmed") if False else (
                data["cum_confirmed"]["median"]
                if isinstance(data["cum_confirmed"], dict)
                else data["cum_confirmed"])
            interp.append(
                f"{scen_defaults[name]['label']}: cumulative confirmed reaches "
                f"{float(med[-1]):,.0f} by horizon end."
            )
        interp.append(
            "Shaded bands on the chart show 90% posterior predictive intervals "
            "derived by sampling R_t starting values from the Cori Gamma posterior."
        )
        _interpretation_block(ws, end_row + 1, " ".join(interp))

        for col_letter, w in {"A": 14, "B": 26, "C": 22, "D": 22, "E": 18,
                                "F": 16, "G": 18}.items():
            ws.column_dimensions[col_letter].width = w

    # =====================================================================
    # SHEET 6 — Step 4: End-of-outbreak
    # =====================================================================
    if valid_plc and eoo_probs is not None and eoo_days is not None:
        ws = wb.create_sheet("Step 4 — End of outbreak")
        _write_title_block(ws, "Step 4 — End-of-outbreak probability",
                            "Nishiura / Lloyd-Smith descendant-tree simulation")

        # Embed EOO chart (from session_state)
        thr = ss.get("eoo_threshold", 0.95)
        _fig_eoo = ss.get("chart_eoo")
        if _fig_eoo is not None:
            _embed_png(ws, _try_fig_to_png(_fig_eoo, width=1200, height=520),
                        "A4")

        # Per-scenario declaration dates
        row = 32
        ws.cell(row=row, column=1,
                 value="Per-scenario declaration dates").font = H2
        row += 1
        for col_i, h in enumerate(
            ["Scenario", "Projected last case", "WHO 42-day",
             "Djaafara 63-day", "Djaafara final (+90 d)",
             f"P(extinct) ≥ {int(thr * 100)}%"], start=1):
            cell = ws.cell(row=row, column=col_i, value=h)
            cell.fill = fill_brand
            cell.font = Font(name="Calibri", size=10, bold=True,
                              color="FFFFFF")
            cell.alignment = LEFT
            cell.border = BORDER
        row += 1

        cross = np.where(eoo_probs >= thr)[0]
        cross_day = int(eoo_days[cross[0]]) if len(cross) > 0 else None
        scen_lbl = {"S1": "Delayed response", "S2": "Moderate response",
                     "S3": "Strong response"}
        for s_name, plc in valid_plc.items():
            cross_date = ((plc + pd.Timedelta(days=cross_day)).strftime("%d %b %Y")
                           if cross_day is not None else "not reached")
            cells = [
                scen_lbl.get(s_name, s_name),
                plc.strftime("%d %b %Y"),
                (plc + pd.Timedelta(days=42)).strftime("%d %b %Y"),
                (plc + pd.Timedelta(days=63)).strftime("%d %b %Y"),
                (plc + pd.Timedelta(days=153)).strftime("%d %b %Y"),
                cross_date,
            ]
            for col_i, v in enumerate(cells, start=1):
                cell = ws.cell(row=row, column=col_i, value=v)
                cell.border = BORDER
                cell.alignment = LEFT
                cell.fill = fill_grey if row % 2 == 0 else fill_brand_light
            ws.row_dimensions[row].height = 20
            row += 1

        # EOO probability table
        row += 2
        ws.cell(row=row, column=1,
                 value="P(extinct) by days after projected last case").font = H2
        row += 1
        prob_rows = []
        for s_name, plc in valid_plc.items():
            for i, d in enumerate(eoo_days):
                prob_rows.append({
                    "scenario": scen_lbl.get(s_name, s_name),
                    "days_after_last_case": int(d),
                    "projected_date": (plc + pd.Timedelta(days=int(d))).strftime("%Y-%m-%d"),
                    "P(extinct)": round(float(eoo_probs[i]), 4),
                    "WHO 42-day met": int(d) >= 42,
                    "Djaafara 63-day met": int(d) >= 63,
                })
        eoo_table = pd.DataFrame(prob_rows)
        end_row = _append_table(ws, eoo_table, row)

        # Interpretation
        if cross_day is not None:
            interp = (
                f"P(extinct) crosses the {thr:.0%} declaration threshold at day "
                f"{cross_day} after the projected last case. Under the Delayed "
                "response scenario this corresponds to the latest end date; the "
                "Strong response scenario yields the earliest. WHO 42-day and "
                "Djaafara 63-day are heuristic operational rules; the "
                f"{thr:.0%} threshold is the standard statistical convention."
            )
        else:
            interp = (
                f"P(extinct) does not reach the {thr:.0%} threshold within "
                f"{int(eoo_days[-1])} days. Extend the days range or revisit Step 3 "
                "scenarios to find an end date."
            )
        _interpretation_block(ws, end_row + 1, interp)

        for col_letter, w in {"A": 26, "B": 18, "C": 18, "D": 18, "E": 22,
                                "F": 18}.items():
            ws.column_dimensions[col_letter].width = w

    # =====================================================================
    # SHEET 7 — Interpretation (overall narrative)
    # =====================================================================
    ws = wb.create_sheet("Interpretation")
    _write_title_block(ws, "Overall interpretation",
                        "Auto-generated narrative based on this run")
    row = 4
    narrative = []
    narrative.append(f"Scenario: {scenario_name}")
    narrative.append(f"Report generated: {now}")
    narrative.append("")
    if series is not None and len(series):
        _line = (
            f"DATA. Daily incidence series covers {len(series)} days "
            f"({pd.to_datetime(series['date']).min():%d %b %Y} → "
            f"{pd.to_datetime(series['date']).max():%d %b %Y}). "
            f"Cumulative observed: {float(series['cumulative_confirmed'].iloc[-1]):,.0f} "
            f"confirmed, "
            f"{float(series['cumulative_suspected'].iloc[-1]):,.0f} suspected."
        )
        if (ss.get("cfr_active")
                and "cumulative_cfr_estimated" in series.columns):
            _line += (
                f" CFR backcalculation active "
                f"(CFR = {ss.get('cfr_pct_active', '?')}%, "
                f"role = {ss.get('cfr_role_active', 'Total cases')}): "
                f"estimated cumulative "
                f"= {float(series['cumulative_cfr_estimated'].iloc[-1]):,.0f}."
            )
        narrative.append(_line)
    if rt_df is not None and not rt_df.empty:
        prim = rt_df[rt_df["si_mean_used"] == si_used].dropna(subset=["rt_mean"])
        if not prim.empty:
            latest = prim.iloc[-1]
            mean_rt = float(latest["rt_mean"])
            phrase = "growing" if mean_rt > 1 else "declining"
            narrative.append(
                f"TRANSMISSION. Latest R_t = {mean_rt:.2f} "
                f"(95% CrI {float(latest['rt_lower']):.2f}–{float(latest['rt_upper']):.2f}); "
                f"the outbreak is in a {phrase} phase as of the most recent "
                f"reporting window."
            )
    if scenarios_out:
        scen_lbl = {"S1": "Delayed", "S2": "Moderate", "S3": "Strong"}
        lines = []
        for name, data in scenarios_out.items():
            v = data["cum_confirmed"]
            med = v["median"] if isinstance(v, dict) else v
            lines.append(
                f"{scen_lbl.get(name, name)}: ~{float(med[-1]):,.0f} cumulative confirmed"
            )
        narrative.append(
            "FORECAST. Under the three response scenarios at the horizon end: "
            + "; ".join(lines) + "."
        )
    if valid_plc and eoo_probs is not None:
        thr = ss.get("eoo_threshold", 0.95)
        cross = np.where(eoo_probs >= thr)[0]
        if len(cross) > 0:
            cross_day = int(eoo_days[cross[0]])
            mid_scen = list(valid_plc.keys())[len(valid_plc) // 2]
            d = (valid_plc[mid_scen] + pd.Timedelta(days=cross_day)).strftime("%d %b %Y")
            narrative.append(
                f"END OF OUTBREAK. P(extinct) crosses {thr:.0%} on approximately "
                f"day {cross_day} after the projected last case. Under the "
                "moderate response scenario this corresponds to declaration "
                f"around {d}."
            )
        else:
            narrative.append(
                f"END OF OUTBREAK. P(extinct) does not reach {thr:.0%} within "
                "the evaluated horizon."
            )
    narrative.append("")
    narrative.append(
        "KEY CAVEATS. Daily counts between sparse cumulative snapshots are "
        "linearly interpolated. R_t scenarios are user-defined assumptions, not "
        "predictions of response speed. Forecast PI bands reflect R_t posterior "
        "sampling only — they do not include observation noise, CFR uncertainty, "
        "or SI uncertainty. EOO is computed via Nishiura/Lloyd-Smith offspring "
        "trees with NB dispersion k from Lloyd-Smith 2005."
    )

    for line in narrative:
        ws.cell(row=row, column=1, value=line).alignment = WRAP
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=11)
        ws.row_dimensions[row].height = max(20, min(80, len(line) // 10 + 18))
        ws.merge_cells(start_row=row, start_column=1, end_row=row,
                       end_column=10)
        row += 1
    ws.column_dimensions["A"].width = 120

    # ---- finalise ----
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


st.set_page_config(page_title="EVD Forecaster", layout="wide")

st.markdown(
    """
    <style>
      :root {
        --brand: #1f4e79;
        --brand-soft: #f1f5fa;
        --line: #e3e7ed;
        --muted: #5b6573;
      }
      #MainMenu, header[data-testid="stHeader"], footer {visibility: hidden;}
      .block-container {
        padding-top: 1.4rem;
        padding-bottom: 1rem;
        max-width: 1400px;
      }
      html, body, [class*="css"] {
        font-family: 'Inter', 'Segoe UI', system-ui, sans-serif;
      }
      h1.app-title {
        font-size: 1.65rem;
        font-weight: 600;
        color: var(--brand);
        margin: 0 0 0.15rem 0;
        letter-spacing: -0.01em;
      }
      .app-sub {
        color: var(--muted);
        font-size: 0.93rem;
        margin-bottom: 1.1rem;
      }
      .panel-title {
        font-size: 1.05rem;
        font-weight: 600;
        color: var(--brand);
        text-transform: uppercase;
        letter-spacing: 0.06em;
        margin: 0 0 0.6rem 0;
        padding-bottom: 0.4rem;
        border-bottom: 2px solid var(--brand);
      }
      .section-label {
        font-size: 0.82rem;
        font-weight: 600;
        color: var(--muted);
        text-transform: uppercase;
        letter-spacing: 0.05em;
        margin: 0.9rem 0 0.35rem 0;
      }
      div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:first-child {
        border-right: 1px solid var(--line);
        padding-right: 1.4rem;
      }
      div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:last-child {
        padding-left: 1.4rem;
      }
      /* Freeze the right output column so it stays visible while the left
         input column scrolls. The empty .freeze-right anchor (placed at the
         top of each right column) lets us target *only* those columns via
         :has(), leaving the header / Excel-row columns untouched. */
      [data-testid="stElementContainer"]:has(.freeze-right) { display: none; }
      div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:has(.freeze-right),
      div[data-testid="stHorizontalBlock"] > div[data-testid="stColumn"]:has(.freeze-right) {
        position: sticky;
        top: 0.75rem;
        align-self: flex-start;
        max-height: calc(100vh - 1.5rem);
        overflow-y: auto;
        overflow-x: hidden;
      }
      /* Compact checkbox legend below the Step-1 daily chart. Streamlit adds a
         .st-key-<key> class to each keyed widget's container, so these rules
         shrink ONLY the daily-legend checkboxes (keys start d1_show_). */
      .mini-legend-label {
        font-size: 0.66rem; color: var(--muted); text-transform: uppercase;
        letter-spacing: 0.05em; margin: 0.3rem 0 0.25rem 0;
      }
      .mini-legend-group {
        font-weight: 600; color: var(--brand); font-size: 0.7rem;
        text-transform: uppercase; letter-spacing: 0.04em;
        margin: 0 0 0.5rem 0; padding-bottom: 0.22rem;
        border-bottom: 1px solid var(--line);
      }
      /* Compact labels; no negative margins (those made a wrapped 2-line label
         collide with the next row, and a flush title collide with row one). */
      [class*="st-key-d1_show_"] p,
      [class*="st-key-c1_show_"] p { font-size: 0.72rem; line-height: 1.25; }
      [class*="st-key-d1_show_"] label,
      [class*="st-key-c1_show_"] label { align-items: flex-start; }
      /* trim Plotly's bottom whitespace so the legend sits right under it */
      div[data-testid="stPlotlyChart"] { margin-bottom: -0.4rem; }
      /* Tighten vertical spacing inside the frozen output panel so the chart
         and the legend below it fit without internal scrolling. Scoped to the
         right (output) column via the .freeze-right anchor. */
      div[data-testid="stColumn"]:has(.freeze-right) [data-testid="stVerticalBlock"],
      div[data-testid="stColumn"]:has(.freeze-right) [data-testid="stVerticalBlockBorderWrapper"] {
        gap: 0.55rem;
      }
      div[data-testid="stRadio"] label p { font-size: 0.88rem; }
      div[data-testid="stCaptionContainer"] { color: var(--muted); }
      .stButton > button[kind="primary"] {
        background: var(--brand);
        border: 1px solid var(--brand);
        font-weight: 600;
        letter-spacing: 0.02em;
      }
      .stButton > button[kind="primary"]:hover { background: #16385a; border-color: #16385a; }
      .stDownloadButton > button { font-weight: 500; }
      .stTabs [data-baseweb="tab-list"] { gap: 0.4rem; }
      .stTabs [data-baseweb="tab"] {
        padding: 0.4rem 0.9rem;
        font-weight: 500;
      }
      .stTabs [aria-selected="true"] { color: var(--brand); }
      .placeholder-card {
        background: var(--brand-soft);
        border: 1px dashed #c7d2e0;
        border-radius: 8px;
        padding: 1.6rem;
        color: var(--muted);
        text-align: center;
      }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown('<h1 class="app-title">EVD Forecaster</h1>', unsafe_allow_html=True)
st.markdown(
    '<div class="app-sub">Build a daily incidence series from cumulative DON '
    'snapshots or incidence data. Output feeds the renewal-equation forecast model.</div>',
    unsafe_allow_html=True,
)

# Scenario name (flows into chart titles and downloaded filenames)
hdr_l, hdr_m, hdr_r = st.columns([2.4, 0.9, 0.9])
with hdr_l:
    scenario_name = st.text_input(
        "Scenario / outbreak label",
        value=st.session_state.get("scenario_name",
                                   "DRC EVD — May 2026"),
        help="Stamped onto chart titles and CSV filenames so you can keep runs apart.",
        key="scenario_name",
    )
with hdr_m:
    st.markdown('<div style="margin-top:1.7rem;"></div>',
                unsafe_allow_html=True)
    if st.button("Input glossary", use_container_width=True,
                 key="open_glossary"):
        st.session_state["prev_step"] = st.session_state.get("step", "data")
        st.session_state["step"] = "help"
        st.rerun()
with hdr_r:
    st.markdown('<div style="margin-top:1.7rem;"></div>',
                unsafe_allow_html=True)
    if st.button("Reset all", use_container_width=True, key="reset_all",
                 help="Clear every input and result. The page reloads with defaults."):
        for k in list(st.session_state.keys()):
            del st.session_state[k]
        st.rerun()

# Persistent Excel-export button row (just below the header).
# IMPORTANT: build the report only when the user clicks — building it on every
# rerun re-ran the (slow, kaleido-based) chart image export on every widget
# interaction, which froze the app once charts existed in session_state.
ex_l, ex_r = st.columns([3.4, 1])
with ex_r:
    if st.button("Generate Excel report", use_container_width=True,
                 key="excel_build",
                 help="Builds a single .xlsx with Dashboard, Inputs, Daily "
                      "incidence, R_t estimates, Forecast, and EOO probability "
                      "sheets — based on whatever is currently in the app."):
        try:
            with st.spinner("Building Excel report…"):
                st.session_state["_excel_bytes"] = build_excel_report()
        except Exception as e:
            st.session_state.pop("_excel_bytes", None)
            st.caption(f"Excel export unavailable: {e}")
    if st.session_state.get("_excel_bytes"):
        slug = _slug(st.session_state.get("scenario_name", ""))
        fname = (f"{slug}__evd_forecaster_report__"
                 f"{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
        st.download_button(
            "Download Excel report",
            data=st.session_state["_excel_bytes"],
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="excel_export",
        )


# Project colour palette (matched to outputs/*.png)
COLOURS = {
    "new_confirmed": "#4682B4",
    "new_suspected": "#FF8C00",
    "new_deaths": "#B22222",
    "cumulative_confirmed": "#4682B4",
    "cumulative_suspected": "#FF8C00",
    "cumulative_deaths": "#B22222",
}

TABLE_COLS = ["date", "new_confirmed", "new_suspected", "new_deaths"]


# Keys whose persistence is tracked across step navigation. These keys hold
# inputs/widget values that survive switching between Step 1/2/3/4/help; the
# Reset all button clears them.
INPUT_STATE_KEYS = [
    "scenario_name", "snapshots_editor",
    "si_mean_primary_val", "si_mean_primary_src",
    "si_sd_primary_val", "si_sd_primary_src",
    "si_mean_sens_val", "si_mean_sens_src",
    "si_sd_sens_val", "si_sd_sens_src",
    "window_val", "window_src",
    "shape_prior_val", "shape_prior_src",
    "rate_prior_val", "rate_prior_src",
    "rt_preset", "selected_rt", "selected_rt_basis",
    "rt_case_type", "rt_source_kind", "rt_use_smoothed",
    "cfr_val", "cfr_src", "lag_val", "lag_src",
    "forecast_horizon", "forecast_start_date",
    "obs_conf_input", "obs_susp_input", "obs_death_input",
    "S1_target", "S1_days", "S2_target", "S2_days", "S3_target", "S3_days",
    "eoo_n_sim", "eoo_max_days", "eoo_threshold", "eoo_k_disp",
    "fc_y_log", "fc_independent_y", "fc_cfr", "result_series",
    "result_chart_snaps", "result_interpolated", "rt_df",
    "rt_si_mean", "rt_si_sd", "fc_scenarios", "fc_dates", "fc_baselines",
    "fc_scen_inputs", "fc_rt_start", "fc_n_samples",
    "fc_preview_traj", "fc_preview_dates",
    "eoo_days", "eoo_probs", "eoo_valid_plc",
    # CFR backcalculation (Step 1)
    "cfr_enabled", "cfr_choice", "cfr_pct_custom",
    "cfr_role", "cfr_t2", "cfr_alpha", "cfr_beta",
    "cfr_active", "cfr_role_active", "cfr_pct_active", "cfr_sma_active",
    # Standalone smoothing (Step 1)
    "smooth_label", "smooth_active", "smooth_window_active",
    # Estimate suspected from confirmed via ratio (Step 1)
    "suspest_enabled", "suspest_tpr", "suspest_basis",
    "suspest_active", "suspest_tpr_active", "suspest_basis_active",
]


def attach_source(df: pd.DataFrame, mode: str, whole: str) -> pd.DataFrame:
    if mode == "Per-row sources":
        if "source" not in df.columns:
            df["source"] = ""
        df["source"] = df["source"].fillna("").astype(str)
    else:
        df["source"] = whole
    return df


def _parse_date_column_strict(df: pd.DataFrame, col: str = "date",
                               where: str = "your file") -> pd.DataFrame:
    """Parse `col` to datetime using ONLY unambiguous formats.

    Order of attempts:
      1. Pass-through if values are already datetime / date objects
         (manual-entry path via st.column_config.DateColumn).
      2. Strict "DD Mmm YYYY"  — e.g. 01 May 2026, 29 May 2026.
      3. Strict "DD-Mmm-YYYY"  — backward-compat for older files.
      4. Strict ISO YYYY-MM-DD — e.g. 2026-05-29.

    Ambiguous numeric formats like 01/02/2026 are deliberately REJECTED
    because pandas would silently pick month/day order based on locale,
    which is the bug the user reported on the Step-1 chart x-axis.
    Unparseable rows raise an st.error that names the bad values."""
    raw = df[col]
    # 1. Already datetime64 → just clone
    if pd.api.types.is_datetime64_any_dtype(raw):
        return df.copy()
    # 2. Object column of Python date/datetime objects (e.g. from
    #    st.column_config.DateColumn) → convert directly.
    first = raw.dropna().iloc[0] if raw.notna().any() else None
    if isinstance(first, (pd.Timestamp, datetime, date)):
        out = df.copy()
        out[col] = pd.to_datetime(raw)
        return out
    # 3. String parsing — DD Mmm YYYY preferred, hyphenated and ISO as
    # backward-compatible fallbacks.
    raw_str = raw.astype(str).str.strip()
    parsed = pd.to_datetime(raw_str, format="%d %b %Y", errors="coerce")
    hy_mask = parsed.isna() & raw.notna()
    if hy_mask.any():
        hy_try = pd.to_datetime(raw_str[hy_mask], format="%d-%b-%Y",
                                 errors="coerce")
        parsed.loc[hy_mask] = hy_try
    iso_mask = parsed.isna() & raw.notna()
    if iso_mask.any():
        iso_try = pd.to_datetime(raw_str[iso_mask], format="%Y-%m-%d",
                                  errors="coerce")
        parsed.loc[iso_mask] = iso_try
    bad_mask = parsed.isna() & raw.notna() & (raw_str != "")
    if bad_mask.any():
        bad_rows = raw_str[bad_mask].head(10).tolist()
        st.error(
            f"❌ Could not read {int(bad_mask.sum())} date(s) in {where}. "
            f"First bad value(s): {bad_rows}. "
            f"Please use **DD Mmm YYYY** (e.g. 01 May 2026, 29 May 2026) "
            f"or YYYY-MM-DD. Numeric formats like 01/02/2026 are not "
            f"accepted (ambiguous month/day)."
        )
        st.stop()
    out = df.copy()
    out[col] = parsed
    return out


def interpolate_from_cumulative(snaps: pd.DataFrame) -> pd.DataFrame:
    snaps = snaps.dropna(subset=["date", "cumulative_confirmed",
                                 "cumulative_suspected", "cumulative_deaths"]).copy()
    if len(snaps) < 2:
        return pd.DataFrame()

    snaps = _parse_date_column_strict(snaps, "date",
                                       where="the cumulative snapshots")
    snaps = snaps.sort_values("date").reset_index(drop=True)

    # ------------------------------------------------------------------
    # BACK-EXTRAPOLATION of the first snapshot's cumulative count.
    # Without this, day 0's daily new = 0 (because there is no prior
    # cumulative to diff against), and the sum of new_X across the
    # displayed series equals (final cumulative - first cumulative)
    # rather than the final cumulative itself.
    #
    # Algorithm: use the per-day rate from the FIRST inter-snapshot
    # window to extend the series backward. For each cumulative
    # column the days needed to accumulate v1 at that rate are
    # ceil(v1 / rate); we pick the max across columns so a single
    # virtual day-0 (cumulative = 0 for every column) anchors the
    # back-extrapolation. Linear interpolation between that virtual
    # day and the first real snapshot then naturally distributes the
    # implied early cases. Result: sum(new_X) = final cumulative_X.
    # ------------------------------------------------------------------
    d1 = snaps.loc[0, "date"]
    d2 = snaps.loc[1, "date"]
    days_between = (d2 - d1).days
    if days_between > 0:
        days_back_per_col = []
        for col in ["cumulative_confirmed", "cumulative_suspected",
                    "cumulative_deaths"]:
            v1 = float(snaps.loc[0, col])
            v2 = float(snaps.loc[1, col])
            delta = v2 - v1
            if delta > 0 and v1 > 0:
                rate = delta / days_between
                days_back_per_col.append(int(np.ceil(v1 / rate)))
        if days_back_per_col:
            days_back = max(days_back_per_col)
            virtual_date = d1 - pd.Timedelta(days=days_back)
            virtual_row = {
                "date": virtual_date,
                "cumulative_confirmed": 0.0,
                "cumulative_suspected": 0.0,
                "cumulative_deaths": 0.0,
            }
            if "source" in snaps.columns:
                virtual_row["source"] = "(back-extrapolated baseline)"
            snaps = pd.concat([pd.DataFrame([virtual_row]), snaps],
                              ignore_index=True)

    full_dates = pd.date_range(snaps["date"].min(), snaps["date"].max(), freq="D")

    indexed = snaps.set_index("date")
    daily = pd.DataFrame(index=full_dates)
    daily.index.name = "date"

    for col in ["cumulative_confirmed", "cumulative_suspected", "cumulative_deaths"]:
        daily[col] = indexed[col].reindex(daily.index).interpolate(method="linear")

    if "source" in snaps.columns:
        s = indexed["source"].reindex(daily.index).ffill().bfill().fillna("")
        daily["source"] = s.values

    daily["new_confirmed"] = daily["cumulative_confirmed"].diff().clip(lower=0).fillna(0)
    daily["new_suspected"] = daily["cumulative_suspected"].diff().clip(lower=0).fillna(0)
    daily["new_deaths"] = daily["cumulative_deaths"].diff().clip(lower=0).fillna(0)

    out_cols = ["new_confirmed", "new_suspected", "new_deaths",
                "cumulative_confirmed", "cumulative_suspected", "cumulative_deaths"]
    if "source" in daily.columns:
        out_cols.append("source")
    return daily[out_cols].reset_index()


def expand_incidence(inc: pd.DataFrame) -> pd.DataFrame:
    inc = inc.dropna(subset=["date", "new_confirmed",
                             "new_suspected", "new_deaths"]).copy()
    if len(inc) < 1:
        return pd.DataFrame()

    inc = _parse_date_column_strict(inc, "date",
                                     where="the incidence file")
    inc = inc.sort_values("date").reset_index(drop=True)

    if len(inc) == 1:
        daily = inc.copy()
    else:
        gaps = inc["date"].diff().dt.days.dropna()
        is_daily = (gaps == 1).all()
        if is_daily:
            daily = inc.copy()
        else:
            rows = []
            prev_date = None
            prev_source = ""
            for _, r in inc.iterrows():
                d = r["date"]
                if prev_date is None:
                    rows.append({
                        "date": d,
                        "new_confirmed": float(r["new_confirmed"]),
                        "new_suspected": float(r["new_suspected"]),
                        "new_deaths": float(r["new_deaths"]),
                        "source": r.get("source", prev_source) or prev_source,
                    })
                else:
                    n = (d - prev_date).days
                    if n <= 0:
                        continue
                    src = r.get("source", "") or prev_source
                    share = {
                        "new_confirmed": float(r["new_confirmed"]) / n,
                        "new_suspected": float(r["new_suspected"]) / n,
                        "new_deaths": float(r["new_deaths"]) / n,
                    }
                    for k in range(1, n + 1):
                        rows.append({
                            "date": prev_date + pd.Timedelta(days=k),
                            **share, "source": src,
                        })
                    prev_source = src
                prev_date = d
            daily = pd.DataFrame(rows)

    daily["cumulative_confirmed"] = daily["new_confirmed"].cumsum()
    daily["cumulative_suspected"] = daily["new_suspected"].cumsum()
    daily["cumulative_deaths"] = daily["new_deaths"].cumsum()

    out_cols = ["date", "new_confirmed", "new_suspected", "new_deaths",
                "cumulative_confirmed", "cumulative_suspected", "cumulative_deaths"]
    if "source" in daily.columns:
        out_cols.append("source")
    return daily[out_cols]


# =========================================================================
# CFR backcalculation (Imperial College / Garske et al. style)
# =========================================================================
# Phase-bias multiplier (1 + r/β)^α corrects naive deaths/CFR for the fact
# that during an exponentially growing epidemic many recent cases have not
# yet died — so deaths reflect older, smaller incidence cohorts. r is the
# epidemic growth rate (log(2)/doubling-time); (α, β) parametrise the
# Gamma onset-to-death distribution.
# -------------------------------------------------------------------------
def cfr_backcalculate(series: pd.DataFrame, cfr_pct: float,
                      doubling_time: float, alpha: float, beta: float,
                      sma_window: int = 0) -> pd.DataFrame:
    out = series.copy()
    if "cumulative_deaths" not in out.columns:
        out["cumulative_deaths"] = out["new_deaths"].cumsum()

    r = np.log(2.0) / float(doubling_time)
    phase_bias = (1.0 + r / float(beta)) ** float(alpha)

    cum_est = (out["cumulative_deaths"].astype(float) * phase_bias) \
              / (float(cfr_pct) / 100.0)
    out["cumulative_cfr_estimated"] = cum_est.round().astype(int)

    new_est = out["cumulative_cfr_estimated"].diff()
    new_est.iloc[0] = out["cumulative_cfr_estimated"].iloc[0]
    out["new_cfr_estimated"] = new_est.astype(int)

    if sma_window and int(sma_window) >= 2:
        w = int(sma_window)
        out["new_deaths_sma"] = (
            out["new_deaths"].rolling(window=w, min_periods=1).mean()
        )
        out["new_cfr_estimated_sma"] = (
            out["new_cfr_estimated"].rolling(window=w, min_periods=1).mean()
        )

    out.attrs["cfr_phase_bias"] = float(phase_bias)
    out.attrs["cfr_growth_rate"] = float(r)
    out.attrs["cfr_pct"] = float(cfr_pct)
    out.attrs["cfr_sma_window"] = int(sma_window) if sma_window else 0
    return out


def smooth_incidence(series: pd.DataFrame, window: int) -> pd.DataFrame:
    """Standalone trailing simple moving average over the daily incidence
    columns, applied independently of the CFR backcalculation.

    The raw observed values are preserved in `<col>_raw`; the `<col>` values
    are replaced by the trailing SMA so every downstream step (R_t, forecast,
    EOO) consumes the smoothed series. A trailing (not centred) window is used
    so no future information leaks into the causal R_t estimate."""
    w = int(window)
    if w < 2:
        return series
    out = series.copy()
    for col in ["new_confirmed", "new_suspected", "new_deaths"]:
        if col in out.columns:
            out[f"{col}_raw"] = out[col].astype(float)
            out[col] = (out[col].astype(float)
                        .rolling(window=w, min_periods=1).mean())
    # Standalone-smoothed copy of the CFR-estimated incidence. Kept in a
    # separate column (not overwritten) so the raw CFR estimate in
    # new_cfr_estimated is preserved for the "Estimated Confirmed" option and
    # the Step 1 table. Uses the SAME window as the actual-case smoothing.
    if "new_cfr_estimated" in out.columns:
        out["new_cfr_estimated_smooth"] = (
            out["new_cfr_estimated"].astype(float)
            .rolling(window=w, min_periods=1).mean())
    # Smoothed copy of the ratio-derived estimated-suspected series (same window).
    if "new_suspected_estimated" in out.columns:
        out["new_suspected_estimated_smooth"] = (
            out["new_suspected_estimated"].astype(float)
            .rolling(window=w, min_periods=1).mean())
    out.attrs["smooth_window"] = w
    return out


def chart_with_line_filter(fig: go.Figure, key: str,
                           hint: bool = True) -> None:
    """Render a Plotly figure and let its legend act as the line filter.

    Plotly legends are already click-to-toggle: a single click on a legend
    entry hides that line (``visible="legendonly"``) and clicking again brings
    it back, while a double-click isolates a single line. Companion traces such
    as confidence-interval bands and baseline markers share each line's
    ``legendgroup``, so they hide and show together with it. A short caption
    points users at this behaviour.
    """
    # Make the toggle behaviour explicit (these are the Plotly defaults, but
    # we set them so the legend stays a reliable show/hide control).
    fig.update_layout(legend=dict(itemclick="toggle",
                                  itemdoubleclick="toggleothers"))
    st.plotly_chart(fig, use_container_width=True, key=key)
    if hint:
        st.caption("Tip: click a name in the legend to hide that line; "
                   "click again to show it. Double-click a name to show only "
                   "that line.")


# The Step-1 "Daily new" chart exposes exactly these 7 logical series, grouped
# into three legend containers. Each entry is
#   (label, [candidate columns in priority order], colour, dash, width).
# The first candidate column present in the built series is the one plotted;
# entries with no available column render as a disabled checkbox.
DAILY_LEGEND_GROUPS = [
    ("Suspected", [
        ("Actual Suspected", ["new_suspected_raw", "new_suspected"],
         "#FF8C00", "solid", 2.2),
        ("Estimated Suspected", ["new_suspected_estimated"],
         "#b8860b", "dot", 1.8),
        ("Estimated Suspected Smoothened",
         ["new_suspected_estimated_smooth"], "#b8860b", "dash", 2.2),
    ]),
    ("Confirmed", [
        ("Actual Confirmed", ["new_confirmed_raw", "new_confirmed"],
         "#4682B4", "solid", 2.2),
        ("Estimated Confirmed", ["new_cfr_estimated"],
         "#6a1b9a", "dot", 1.8),
        ("Estimated Confirmed Smoothened",
         ["new_cfr_estimated_smooth", "new_cfr_estimated_sma"],
         "#6a1b9a", "dash", 2.2),
    ]),
    ("Death", [
        ("Actual Death", ["new_deaths_raw", "new_deaths"],
         "#B22222", "solid", 2.2),
    ]),
]

# Checked by default: the three observed "actual" lines. Estimates are opt-in.
DAILY_LEGEND_DEFAULT_ON = {"Actual Suspected", "Actual Confirmed",
                           "Actual Death"}

# Coloured dot prefixed to each checkbox label so it reads like a legend.
_DAILY_SWATCH = {"#FF8C00": "🟠", "#b8860b": "🟡", "#4682B4": "🔵",
                 "#6a1b9a": "🟣", "#B22222": "🔴"}


def _resolve_daily_col(series: pd.DataFrame, candidates):
    """First candidate column actually present in the built series, else None."""
    for c in candidates:
        if c in series.columns:
            return c
    return None


def daily_selected_from_state(series: pd.DataFrame) -> list:
    """Resolve which daily-chart traces to draw from the checkbox state already
    in session_state. The checkboxes themselves are rendered *below* the chart
    by render_daily_legend, so here we just read their persisted values (with
    the same defaults) to decide what to plot above them."""
    selected = []
    for _group, items in DAILY_LEGEND_GROUPS:
        for label, candidates, colour, dash, width in items:
            resolved = _resolve_daily_col(series, candidates)
            if resolved is None:
                continue
            key = "d1_show_" + _slug(label)
            default = label in DAILY_LEGEND_DEFAULT_ON
            if st.session_state.get(key, default):
                selected.append((label, resolved, colour, dash, width))
    return selected


def render_daily_legend(series: pd.DataFrame) -> None:
    """Compact 3-container checkbox legend, rendered *below* the daily chart.
    Each of the 7 logical series has a colour-dot checkbox; ticking it adds the
    line. Series the current inputs can't produce show as disabled checkboxes."""
    st.markdown('<div class="mini-legend-label">Legend — tick to show</div>',
                unsafe_allow_html=True)
    cols = st.columns(len(DAILY_LEGEND_GROUPS))
    for (group_name, items), col in zip(DAILY_LEGEND_GROUPS, cols):
        with col:
            st.markdown(f'<div class="mini-legend-group">{group_name}</div>',
                        unsafe_allow_html=True)
            for label, candidates, colour, dash, width in items:
                available = _resolve_daily_col(series, candidates) is not None
                swatch = _DAILY_SWATCH.get(colour, "")
                st.checkbox(
                    f"{swatch} {label}",
                    value=available and label in DAILY_LEGEND_DEFAULT_ON,
                    key="d1_show_" + _slug(label),
                    disabled=not available,
                    help=None if available else
                    "Not available for the current inputs — enable the matching "
                    "estimate / smoothing option on the left, then regenerate.")


def daily_chart(series: pd.DataFrame, selected: list,
                interpolated: bool = True) -> go.Figure:
    """Daily-incidence chart driven by the checkbox legend. `selected` is the
    list returned by render_daily_legend — only those traces are drawn, and the
    built-in Plotly legend is suppressed in favour of the checkboxes."""
    fig = go.Figure()
    for label, col, colour, dash, width in selected:
        fig.add_trace(go.Scatter(
            x=series["date"], y=series[col],
            mode="lines", name=label,
            line=dict(color=colour, width=width, dash=dash),
            hovertemplate="%{x|%d %b %Y}<br>" + label
            + ": %{y:.1f}<extra></extra>",
        ))
    fig.update_layout(
        title=dict(text="Daily new cases" + (" (interpolated)"
                                             if interpolated else ""),
                   font=dict(size=15, color="#1f4e79"), x=0.01),
        xaxis_title="Date", yaxis_title="New cases per day",
        template="simple_white", hovermode="x unified",
        showlegend=False,
        margin=dict(l=60, r=20, t=44, b=24), height=340,
        font=dict(family="Inter, Segoe UI, sans-serif", size=12, color="#333"),
        plot_bgcolor="white",
    )
    fig.update_xaxes(showgrid=True, gridcolor="#eef1f5",
                      linecolor="#cfd6df", tickformat="%d %b %Y")
    fig.update_yaxes(showgrid=True, gridcolor="#eef1f5", linecolor="#cfd6df")
    if not selected:
        fig.add_annotation(text="Tick a series above to plot it",
                           xref="paper", yref="paper", x=0.5, y=0.5,
                           showarrow=False,
                           font=dict(size=13, color="#9aa3af"))
    return fig


# The Step-1 "Cumulative" chart mirrors the daily legend: three containers,
# each with a colour-dot checkbox. "snapshot" entries come from the raw DON
# snapshot table (snaps) and draw as dotted lines; "series" entries are the
# interpolated / estimated columns from the built daily series.
CUM_LEGEND_GROUPS = [
    ("Suspected", [
        ("Actual Suspected", "series", ["cumulative_suspected"],
         "#FF8C00", "solid", 2.2),
        ("Estimated Suspected", "series", ["cumulative_suspected_estimated"],
         "#b8860b", "dash", 2.4),
        ("Suspected (snapshots)", "snapshot", ["cumulative_suspected"],
         "#FF8C00", "dot", 1.6),
    ]),
    ("Confirmed", [
        ("Actual Confirmed", "series", ["cumulative_confirmed"],
         "#4682B4", "solid", 2.2),
        ("Estimated Confirmed", "series", ["cumulative_cfr_estimated"],
         "#6a1b9a", "solid", 2.4),
        ("Confirmed (snapshots)", "snapshot", ["cumulative_confirmed"],
         "#4682B4", "dot", 1.6),
    ]),
    ("Death", [
        ("Actual Death", "series", ["cumulative_deaths"],
         "#B22222", "solid", 2.2),
        ("Death (snapshots)", "snapshot", ["cumulative_deaths"],
         "#B22222", "dot", 1.6),
    ]),
]
CUM_LEGEND_DEFAULT_ON = {"Actual Suspected", "Actual Confirmed", "Actual Death"}


def _cum_resolve(series: pd.DataFrame, snaps, kind: str, candidates):
    """First available column for a cumulative legend item, or None.
    snapshot items resolve against the snaps table; series items against the
    built daily series."""
    if kind == "snapshot":
        if snaps is None or len(snaps) == 0:
            return None
        cols = snaps.columns
    else:
        cols = series.columns
    for c in candidates:
        if c in cols:
            return c
    return None


def cum_selected_from_state(series: pd.DataFrame, snaps) -> list:
    """Resolve which cumulative traces to draw from the checkbox state in
    session_state. Returns [(label, kind, column, colour, dash, width), ...]."""
    selected = []
    for _group, items in CUM_LEGEND_GROUPS:
        for label, kind, candidates, colour, dash, width in items:
            resolved = _cum_resolve(series, snaps, kind, candidates)
            if resolved is None:
                continue
            key = "c1_show_" + _slug(label)
            default = label in CUM_LEGEND_DEFAULT_ON
            if st.session_state.get(key, default):
                selected.append((label, kind, resolved, colour, dash, width))
    return selected


def render_cum_legend(series: pd.DataFrame, snaps) -> None:
    """Compact 3-container checkbox legend for the cumulative chart, rendered
    below it. Unavailable series show as disabled checkboxes."""
    st.markdown('<div class="mini-legend-label">Legend — tick to show</div>',
                unsafe_allow_html=True)
    cols = st.columns(len(CUM_LEGEND_GROUPS))
    for (group_name, items), col in zip(CUM_LEGEND_GROUPS, cols):
        with col:
            st.markdown(f'<div class="mini-legend-group">{group_name}</div>',
                        unsafe_allow_html=True)
            for label, kind, candidates, colour, dash, width in items:
                available = _cum_resolve(series, snaps, kind,
                                         candidates) is not None
                swatch = _DAILY_SWATCH.get(colour, "")
                st.checkbox(
                    f"{swatch} {label}",
                    value=available and label in CUM_LEGEND_DEFAULT_ON,
                    key="c1_show_" + _slug(label),
                    disabled=not available,
                    help=None if available else
                    "Not available for the current inputs — enable the matching "
                    "estimate option on the left, or add snapshot rows.")


def cumulative_chart(series: pd.DataFrame, snaps: pd.DataFrame | None,
                     selected: list, interpolated: bool = True) -> go.Figure:
    """Cumulative chart driven by the checkbox legend. Only `selected` traces
    are drawn; the built-in Plotly legend is suppressed."""
    fig = go.Figure()
    # Prepare the sorted snapshot table once (dates converted BEFORE sorting so
    # string dates don't sort lexicographically into a wrong x-axis order).
    snap_sorted = None
    if snaps is not None and len(snaps) > 0:
        snap_sorted = snaps.dropna(
            subset=["cumulative_confirmed", "cumulative_suspected",
                    "cumulative_deaths"]).copy()
        snap_sorted["date"] = pd.to_datetime(snap_sorted["date"])
        snap_sorted = snap_sorted.sort_values("date")
    has_source = snap_sorted is not None and "source" in snap_sorted.columns

    for label, kind, col, colour, dash, width in selected:
        if kind == "snapshot":
            if snap_sorted is None or col not in snap_sorted.columns:
                continue
            customdata = (snap_sorted["source"] if has_source
                          else [""] * len(snap_sorted))
            hover = ("%{x|%d %b %Y}<br>" + label + ": %{y:.0f}"
                     + ("<br>Source: %{customdata}" if has_source else "")
                     + "<extra></extra>")
            fig.add_trace(go.Scatter(
                x=snap_sorted["date"], y=snap_sorted[col],
                mode="markers+lines", name=label,
                line=dict(color=colour, width=width, dash=dash),
                marker=dict(size=6, color=colour),
                customdata=customdata, hovertemplate=hover,
            ))
        else:
            fig.add_trace(go.Scatter(
                x=series["date"], y=series[col],
                mode="lines", name=label,
                line=dict(color=colour, width=width, dash=dash),
                hovertemplate="%{x|%d %b %Y}<br>" + label
                + ": %{y:.0f}<extra></extra>",
            ))

    fig.update_layout(
        title=dict(text="Cumulative cases" + (" (interpolated)"
                                              if interpolated else ""),
                   font=dict(size=15, color="#1f4e79"), x=0.01),
        xaxis_title="Date", yaxis_title="Cumulative count",
        template="simple_white", hovermode="x unified",
        showlegend=False,
        margin=dict(l=60, r=20, t=44, b=24), height=340,
        font=dict(family="Inter, Segoe UI, sans-serif", size=12, color="#333"),
        plot_bgcolor="white",
    )
    fig.update_xaxes(showgrid=True, gridcolor="#eef1f5",
                      linecolor="#cfd6df", tickformat="%d %b %Y")
    fig.update_yaxes(showgrid=True, gridcolor="#eef1f5", linecolor="#cfd6df")
    if not selected:
        fig.add_annotation(text="Tick a series above to plot it",
                           xref="paper", yref="paper", x=0.5, y=0.5,
                           showarrow=False,
                           font=dict(size=13, color="#9aa3af"))
    return fig


# =========================================================================
# R_t estimation (Cori et al. 2013) — ported from scripts/rt_estimation.py
# =========================================================================
def si_weights(n_days: int, si_mean: float, si_sd: float) -> np.ndarray:
    shape = (si_mean / si_sd) ** 2
    scale = si_sd ** 2 / si_mean
    s = np.arange(1, n_days + 1, dtype=float)
    w = gamma_dist.cdf(s + 0.5, a=shape, scale=scale) - \
        gamma_dist.cdf(s - 0.5, a=shape, scale=scale)
    if w.sum() > 0:
        w = w / w.sum()
    return w


def estimate_rt(incidence: np.ndarray, si_mean: float, si_sd: float,
                window: int = 7, shape_prior: float = 1.0,
                rate_prior: float = 0.2) -> list:
    """Cori et al. (2013) sliding-window Bayesian R_t estimator.

    For a window of length τ ending at time t:
        I_window  = Σ_{s=t-τ+1}^{t} I_s        (total cases in window)
        Λ_window  = Σ_{s=t-τ+1}^{t} Λ_s        (SI-weighted past incidence,
                                                summed over the SAME window)
        Λ_s       = Σ_{k=1}^{s} w_k · I_{s-k}
    Posterior R_t ~ Gamma(α₀ + I_window, β₀ + Λ_window).

    Earlier versions used Λ_t (only at the right-edge day) which under-
    counts the denominator τ-fold and inflates R_t. This matches the
    EpiEstim R package and Cori 2013 eq. (4).
    """
    n = len(incidence)
    weights = si_weights(n, si_mean, si_sd)

    # Precompute Λ_s for every day s in the series, once.
    lambda_s = np.zeros(n, dtype=float)
    for s_day in range(n):
        acc = 0.0
        s_max = min(s_day, len(weights))
        for k in range(1, s_max + 1):
            acc += weights[k - 1] * incidence[s_day - k]
        lambda_s[s_day] = acc

    records = []
    for t in range(window - 1, n):
        I_window = float(incidence[t - window + 1: t + 1].sum())
        Lambda_window = float(lambda_s[t - window + 1: t + 1].sum())
        shape_post = shape_prior + I_window
        rate_post = rate_prior + Lambda_window
        if rate_post > 0:
            rt_mean = shape_post / rate_post
            rt_lower = gamma_dist.ppf(0.025, a=shape_post, scale=1 / rate_post)
            rt_upper = gamma_dist.ppf(0.975, a=shape_post, scale=1 / rate_post)
        else:
            rt_mean = rt_lower = rt_upper = np.nan
        records.append({
            "rt_mean": rt_mean, "rt_lower": rt_lower, "rt_upper": rt_upper,
            "shape_post": shape_post, "rate_post": rate_post,
        })
    return records


def _resolve_rt_col(daily: pd.DataFrame, source: str) -> str:
    """Map a Step 2 incidence-source key to the actual DataFrame column.

    Six keys, three families × {actual, smoothed}:
      - "confirmed" / "suspected"  → Actual observed counts. When Step 1
        smoothing is on it overwrote new_confirmed/new_suspected in place and
        kept the originals in <col>_raw, so read the raw copy if it exists.
      - "cfr"                       → Estimated Confirmed: the raw CFR
        back-calculated incidence (new_cfr_estimated), never overwritten.
      - "confirmed_smooth" / "suspected_smooth" → the smoothed values, which
        live in the base column after Step 1 smoothing.
      - "cfr_smooth"                → Smoothed Estimated Confirmed: the
        standalone-smoothed CFR copy (new_cfr_estimated_smooth)."""
    if source == "confirmed":
        return ("new_confirmed_raw" if "new_confirmed_raw" in daily.columns
                else "new_confirmed")
    if source == "suspected":
        return ("new_suspected_raw" if "new_suspected_raw" in daily.columns
                else "new_suspected")
    if source == "cfr":
        return "new_cfr_estimated"
    if source == "confirmed_smooth":
        return "new_confirmed"
    if source == "suspected_smooth":
        return "new_suspected"
    if source == "cfr_smooth":
        return "new_cfr_estimated_smooth"
    if source == "suspected_est":
        return "new_suspected_estimated"
    if source == "suspected_est_smooth":
        return "new_suspected_estimated_smooth"
    return "new_confirmed"


def compute_rt_table(daily: pd.DataFrame, si_mean: float, si_sd: float,
                     window: int, shape_prior: float, rate_prior: float,
                     run_sensitivity: bool, sens_mean: float,
                     sens_sd: float,
                     incidence_source: str = "confirmed") -> pd.DataFrame:
    """Build per-day R_t estimates. `incidence_source` selects which series
    feeds the Cori estimator — see _resolve_rt_col for the six keys (Actual /
    Estimated Confirmed/Suspected, raw or smoothed)."""
    col = _resolve_rt_col(daily, incidence_source)
    if col not in daily.columns:
        col = "new_confirmed"
    incidence = daily[col].astype(float).values
    dates = pd.to_datetime(daily["date"]).values
    if len(incidence) < window:
        return pd.DataFrame()
    result_dates = dates[window - 1:]

    rows = []
    for rec, d in zip(
        estimate_rt(incidence, si_mean, si_sd, window, shape_prior, rate_prior),
        result_dates,
    ):
        rows.append({"date": d, **rec, "si_mean_used": si_mean,
                     "window_size": window,
                     "incidence_source": incidence_source})

    if run_sensitivity:
        for rec, d in zip(
            estimate_rt(incidence, sens_mean, sens_sd, window, shape_prior, rate_prior),
            result_dates,
        ):
            rows.append({"date": d, **rec, "si_mean_used": sens_mean,
                         "window_size": window,
                         "incidence_source": incidence_source})

    return pd.DataFrame(rows)


def rt_combined_chart(daily: pd.DataFrame, rt_df: pd.DataFrame,
                      si_mean: float) -> go.Figure:
    """R_t trajectory with 95% CrI band, optional sensitivity overlay, R_t=1 line."""
    fig = go.Figure()
    primary = rt_df[rt_df["si_mean_used"] == si_mean].copy()
    primary["date"] = pd.to_datetime(primary["date"])

    fig.add_trace(go.Scatter(
        x=pd.concat([primary["date"], primary["date"][::-1]]),
        y=pd.concat([primary["rt_upper"], primary["rt_lower"][::-1]]),
        fill="toself", fillcolor="rgba(178,34,34,0.18)",
        line=dict(color="rgba(0,0,0,0)"), name="95% CrI",
        hoverinfo="skip", showlegend=True,
    ))
    fig.add_trace(go.Scatter(
        x=primary["date"], y=primary["rt_mean"],
        mode="lines+markers", name="R_t mean",
        line=dict(color="#8B0000", width=2.4),
        marker=dict(size=6, color="#8B0000"),
        hovertemplate="%{x|%d %b %Y}<br>R_t: %{y:.2f}<extra></extra>",
    ))

    sens_subset = rt_df[rt_df["si_mean_used"] != si_mean]
    if len(sens_subset) > 0:
        sens_mean = sens_subset["si_mean_used"].iloc[0]
        sens_subset = sens_subset.copy()
        sens_subset["date"] = pd.to_datetime(sens_subset["date"])
        fig.add_trace(go.Scatter(
            x=sens_subset["date"], y=sens_subset["rt_mean"],
            mode="lines+markers",
            name=f"Sensitivity (SI = {sens_mean:.1f} d)",
            line=dict(color="#2E8B8B", width=1.8, dash="dash"),
            marker=dict(size=5, color="#2E8B8B"),
            hovertemplate="%{x|%d %b %Y}<br>R_t: %{y:.2f}<extra></extra>",
        ))

    if len(primary) > 0:
        fig.add_trace(go.Scatter(
            x=[primary["date"].min(), primary["date"].max()],
            y=[1, 1], mode="lines",
            line=dict(color="black", dash="dot", width=1.4),
            name="R_t = 1", hoverinfo="skip",
        ))

    fig.update_layout(
        title=dict(text=f"Instantaneous R_t (SI mean = {si_mean:.1f} d)",
                   font=dict(size=15, color="#1f4e79"), x=0.01),
        xaxis_title="Date", yaxis_title="Instantaneous R_t",
        template="simple_white", hovermode="x unified",
        legend=dict(orientation="h", yanchor="top", y=-0.20,
                    xanchor="center", x=0.5, font=dict(size=11)),
        margin=dict(l=60, r=20, t=50, b=95),
        height=480,
        font=dict(family="Inter, Segoe UI, sans-serif", size=12, color="#333"),
        plot_bgcolor="white",
    )
    fig.update_xaxes(showgrid=True, gridcolor="#eef1f5", linecolor="#cfd6df")
    fig.update_yaxes(showgrid=True, gridcolor="#eef1f5", linecolor="#cfd6df")
    return fig


# =========================================================================
# Forecast (renewal equation, Nouvellet et al. 2018) — port of who_forecast.py
# =========================================================================
def _project_renewal_batch(seed: np.ndarray, rt_traj_2d: np.ndarray,
                           horizon: int, weights: np.ndarray) -> np.ndarray:
    """Vectorised renewal projection across many R_t trajectories at once.

    seed       : 1-D observed incidence, shared by every trajectory.
    rt_traj_2d : (n_sim, horizon) — one R_t trajectory per sample.
    weights    : serial-interval weights.
    Returns (n_sim, horizon) projected new incidence. Row i is identical (up to
    floating-point rounding) to the scalar recurrence
        I_t = R_t · Σ_{s=1}^{min(t, len(w))} w_{s-1} · I_{t-s}
    run for trajectory i. The time loop stays (each day depends on earlier
    projected days) but all samples advance together, and the per-day
    convolution is a single matrix–vector product instead of a Python sum.
    """
    seed = np.asarray(seed, dtype=float)
    n_seed = len(seed)
    n_sim = rt_traj_2d.shape[0]
    n_w = len(weights)
    full = np.zeros((n_sim, n_seed + horizon))
    full[:, :n_seed] = seed                       # broadcast seed to every row
    for k in range(horizon):
        t = n_seed + k
        s_max = min(t, n_w)
        # window columns are [t-s_max .. t-1]; reversed they line up with
        # weights[0..s_max-1] so the dot = Σ w_{s-1}·full[t-s].
        window = full[:, t - s_max:t]
        contrib = window[:, ::-1] @ weights[:s_max]
        full[:, t] = rt_traj_2d[:, k] * contrib
    return full[:, n_seed:]


def run_scenario_uncertain(seed_conf, seed_susp, rt_start_samples,
                            target, days_to_target, horizon,
                            obs_conf, obs_susp, obs_deaths,
                            cfr, death_lag, si_mean, si_sd):
    """Run the renewal forecast for N R_t-start samples; return median + 90% PI bands.

    rt_start_samples : 1-D array of starting R_t values sampled from the Cori
    Gamma posterior. Each becomes a linear trajectory to `target`
    over `days_to_target` days, then holds.

    Vectorised across the R_t samples (and the per-day serial-interval
    convolution) for speed; numerically identical to projecting each sample
    one at a time.
    """
    rt_start_samples = np.asarray(rt_start_samples, dtype=float)
    n_sim = len(rt_start_samples)
    seed_conf = np.asarray(seed_conf, dtype=float)
    seed_susp = np.asarray(seed_susp, dtype=float)
    n_seed = len(seed_conf)

    # Serial-interval weights (discretised Gamma), normalised to sum 1.
    shape = (si_mean / si_sd) ** 2
    scale = si_sd ** 2 / si_mean
    n_w = horizon + n_seed
    s = np.arange(1, n_w + 1, dtype=float)
    weights = gamma_dist.cdf(s + 0.5, a=shape, scale=scale) - \
              gamma_dist.cdf(s - 0.5, a=shape, scale=scale)
    if weights.sum() > 0:
        weights = weights / weights.sum()

    # R_t trajectory per sample: linear ramp from each sampled start to `target`
    # over days_to_target, then a plateau — build_rt_trajectory for all at once.
    d2t = max(1, min(int(days_to_target), int(horizon)))
    frac = np.linspace(0.0, 1.0, d2t)
    ramp = (rt_start_samples[:, None] * (1.0 - frac)[None, :]
            + target * frac[None, :])
    plateau = np.full((n_sim, horizon - d2t), float(target))
    rt_traj_2d = np.concatenate([ramp, plateau], axis=1)

    new_conf = _project_renewal_batch(seed_conf, rt_traj_2d, horizon, weights)
    new_susp = _project_renewal_batch(seed_susp, rt_traj_2d, horizon, weights)

    # Deaths on day t = cfr · confirmed on day (t - death_lag), drawing on the
    # observed seed history for the first death_lag days. Same rule as the
    # scalar version, expressed as a single shifted slice.
    combined = np.concatenate(
        [np.broadcast_to(seed_conf, (n_sim, n_seed)), new_conf], axis=1)
    new_death = np.zeros((n_sim, horizon))
    src_start = n_seed - death_lag        # combined-array index feeding day t=0
    t0 = max(0, -src_start)               # earlier days have no valid source → 0
    if t0 < horizon:
        i0 = src_start + t0
        count = horizon - t0
        new_death[:, t0:t0 + count] = cfr * combined[:, i0:i0 + count]

    cum_conf = obs_conf + np.cumsum(new_conf, axis=1)
    cum_susp = obs_susp + np.cumsum(new_susp, axis=1)
    cum_death = obs_deaths + np.cumsum(new_death, axis=1)

    def _stats(arr):
        return {
            "median": np.median(arr, axis=0),
            "lower":  np.percentile(arr, 5, axis=0),
            "upper":  np.percentile(arr, 95, axis=0),
        }
    return {
        "cum_confirmed": _stats(cum_conf),
        "cum_suspected": _stats(cum_susp),
        "cum_deaths":    _stats(cum_death),
        "new_confirmed": _stats(new_conf),
        "new_suspected": _stats(new_susp),
        "new_deaths":    _stats(new_death),
    }


def build_rt_trajectory(rt_start: float, target: float, days_to_target: int,
                        horizon: int) -> np.ndarray:
    days_to_target = max(1, min(int(days_to_target), int(horizon)))
    return np.concatenate([
        np.linspace(rt_start, target, days_to_target),
        np.full(horizon - days_to_target, target),
    ])


SCENARIO_COLOURS = {
    "S1": "#C0392B",
    "S2": "#E07B39",
    "S3": "#1E8449",
}


def rt_preview_chart(traj_dict: dict, dates) -> go.Figure:
    fig = go.Figure()

    SCENARIO_LABELS = {
        "S1": "S1 — Delayed",
        "S2": "S2 — Moderate",
        "S3": "S3 — Strong",
    }

    # Determine y-axis range so shaded zones fit cleanly
    y_max = max(float(np.max(t)) for t in traj_dict.values())
    y_max_plot = max(1.6, y_max * 1.12)

    # Background zones: red tint above R_t = 1 (growing), green below (declining)
    fig.add_shape(type="rect", xref="paper", yref="y",
                  x0=0, x1=1, y0=1, y1=y_max_plot,
                  fillcolor="rgba(178,34,34,0.05)",
                  line=dict(width=0), layer="below")
    fig.add_shape(type="rect", xref="paper", yref="y",
                  x0=0, x1=1, y0=0, y1=1,
                  fillcolor="rgba(30,132,73,0.05)",
                  line=dict(width=0), layer="below")

    # Zone labels (top-left and bottom-left)
    fig.add_annotation(
        xref="paper", yref="y",
        x=0.012, y=y_max_plot * 0.94, showarrow=False,
        text="<b>Growing</b> (R<sub>t</sub> > 1)",
        font=dict(size=10, color="#B22222"), xanchor="left",
    )
    fig.add_annotation(
        xref="paper", yref="y",
        x=0.012, y=0.18, showarrow=False,
        text="<b>Declining</b> (R<sub>t</sub> < 1)",
        font=dict(size=10, color="#1E8449"), xanchor="left",
    )

    # Scenario lines + marker at target-reached point
    for name, traj in traj_dict.items():
        traj_arr = np.asarray(traj, dtype=float)
        target = float(traj_arr[-1])
        target_idx = next(
            (i for i, v in enumerate(traj_arr) if abs(v - target) < 1e-6),
            len(traj_arr) - 1,
        )
        days_to_target = int(target_idx)
        legend_label = (f"{SCENARIO_LABELS.get(name, name)} "
                        f"(→ {target:.1f} over {days_to_target} d)")
        fig.add_trace(go.Scatter(
            x=dates, y=traj, mode="lines",
            name=legend_label,
            line=dict(color=SCENARIO_COLOURS[name], width=2.8),
            hovertemplate=f"<b>{name}</b><br>%{{x|%d %b %Y}}<br>"
                          "R<sub>t</sub>: %{y:.2f}<extra></extra>",
        ))
        if 0 < target_idx < len(dates):
            fig.add_trace(go.Scatter(
                x=[dates[target_idx]], y=[target],
                mode="markers",
                marker=dict(size=9, color=SCENARIO_COLOURS[name],
                            line=dict(color="white", width=1.6)),
                showlegend=False,
                hovertemplate=(
                    f"<b>{name}</b> target reached<br>"
                    "%{x|%d %b %Y}<br>"
                    f"R<sub>t</sub> = {target:.2f}<extra></extra>"
                ),
            ))

    # R_t = 1 threshold (line + annotation, not in legend)
    fig.add_shape(type="line", xref="paper", yref="y",
                  x0=0, x1=1, y0=1, y1=1,
                  line=dict(color="#444", dash="dash", width=1.4))
    fig.add_annotation(
        xref="paper", yref="y",
        x=0.995, y=1, showarrow=False,
        text="<b>R<sub>t</sub> = 1</b>  epidemic threshold",
        font=dict(size=10, color="#444"),
        bgcolor="rgba(255,255,255,0.85)",
        xanchor="right", yanchor="bottom",
    )

    fig.update_layout(
        title=dict(
            text="<b>Assumed R<sub>t</sub> trajectories</b> "
                 "— how R<sub>t</sub> evolves under each response scenario",
            font=dict(size=14, color="#1f4e79"), x=0.01, xanchor="left",
        ),
        xaxis_title="Date",
        yaxis_title="R<sub>t</sub> (reproduction number)",
        template="simple_white",
        height=340,
        margin=dict(l=70, r=30, t=65, b=85),
        legend=dict(orientation="h", yanchor="top", y=-0.22,
                    xanchor="center", x=0.5, font=dict(size=11),
                    bgcolor="rgba(255,255,255,0.7)"),
        font=dict(family="Inter, Segoe UI, sans-serif", size=12, color="#333"),
        plot_bgcolor="white",
        hovermode="x unified",
        yaxis=dict(range=[0, y_max_plot], zeroline=False),
    )
    fig.update_xaxes(showgrid=True, gridcolor="#eef1f5", linecolor="#cfd6df")
    fig.update_yaxes(showgrid=True, gridcolor="#eef1f5", linecolor="#cfd6df")
    return fig


def forecast_chart(scenarios: dict, horizon_dates, baselines: dict,
                   labels: dict, y_log: bool = True,
                   mode: str = "cumulative",
                   independent_y: bool = False) -> go.Figure:
    """One panel per response scenario, each showing 3 series (confirmed /
    suspected / deaths). mode='cumulative' (default) plots cumulative counts
    with the observed baseline as a dotted reference; mode='daily' plots
    daily new cases per day, no baseline line. Always draws 90% PI bands
    when median/lower/upper dicts are available.

    independent_y=False (default) shares one y-axis across the three panels so
    scenarios are directly comparable; True gives each panel its own y-axis
    scale (and its own tick labels), useful when one scenario dwarfs another."""
    panel_order = [s for s in ["S1", "S2", "S3"] if s in scenarios]
    titles = [labels.get(s, s) for s in panel_order]

    fig = make_subplots(rows=1, cols=len(panel_order),
                        subplot_titles=titles,
                        horizontal_spacing=(0.11 if independent_y else 0.09),
                        shared_yaxes=not independent_y)

    is_daily = mode == "daily"
    if is_daily:
        metric_map = [
            ("new_confirmed", "New confirmed",  "#4682B4",
             "rgba(70,130,180,0.18)", "confirmed"),
            ("new_suspected", "New suspected",  "#FF8C00",
             "rgba(255,140,0,0.18)", "suspected"),
            ("new_deaths",    "New deaths",     "#B22222",
             "rgba(178,34,34,0.18)", "deaths"),
        ]
    else:
        metric_map = [
            ("cum_confirmed", "Cumulative confirmed", "#4682B4",
             "rgba(70,130,180,0.18)", "confirmed"),
            ("cum_suspected", "Cumulative suspected", "#FF8C00",
             "rgba(255,140,0,0.18)", "suspected"),
            ("cum_deaths",    "Cumulative deaths",    "#B22222",
             "rgba(178,34,34,0.18)", "deaths"),
        ]

    def _split(metric_value):
        if isinstance(metric_value, dict):
            return (metric_value.get("median"),
                    metric_value.get("lower"),
                    metric_value.get("upper"))
        return metric_value, None, None

    for col_idx, s_name in enumerate(panel_order, start=1):
        s_data = scenarios[s_name]
        for col_key, label, colour, band_colour, baseline_key in metric_map:
            median, lower, upper = _split(s_data[col_key])
            # 90% PI band if available
            if lower is not None and upper is not None:
                fig.add_trace(go.Scatter(
                    x=list(horizon_dates) + list(horizon_dates[::-1]),
                    y=list(upper) + list(lower[::-1]),
                    fill="toself", fillcolor=band_colour,
                    line=dict(width=0),
                    legendgroup=label, showlegend=False,
                    hoverinfo="skip",
                ), row=1, col=col_idx)
            fig.add_trace(go.Scatter(
                x=horizon_dates, y=median,
                mode="lines",
                name=label,
                line=dict(color=colour, width=2.2),
                legendgroup=label,
                showlegend=(col_idx == 1),
                hovertemplate=(f"<b>{label}</b><br>"
                               "%{x|%d %b %Y}<br>%{y:,.0f}<extra></extra>"),
            ), row=1, col=col_idx)
            # Observed baseline marker (dotted horizontal) — only for cumulative
            if not is_daily:
                fig.add_trace(go.Scatter(
                    x=[horizon_dates[0], horizon_dates[-1]],
                    y=[baselines[baseline_key], baselines[baseline_key]],
                    mode="lines",
                    line=dict(color=colour, dash="dot", width=1),
                    opacity=0.55,
                    legendgroup=label,
                    showlegend=False,
                    hoverinfo="skip",
                ), row=1, col=col_idx)
        fig.update_yaxes(type=("log" if y_log else "linear"),
                         row=1, col=col_idx,
                         showgrid=True, gridcolor="#eef1f5",
                         linecolor="#cfd6df")
        fig.update_xaxes(row=1, col=col_idx,
                         showgrid=True, gridcolor="#eef1f5",
                         linecolor="#cfd6df")
        if col_idx == 1:
            if is_daily:
                title = ("New cases per day (log)" if y_log
                         else "New cases per day")
            else:
                title = ("Cumulative count (log)" if y_log
                         else "Cumulative count")
            fig.update_yaxes(title_text=title, row=1, col=col_idx)

    fig.update_layout(
        template="simple_white",
        height=560,
        margin=dict(l=60, r=20, t=140, b=95),
        legend=dict(orientation="h", yanchor="top", y=-0.18,
                    xanchor="center", x=0.5, font=dict(size=11),
                    bgcolor="rgba(255,255,255,0.7)"),
        font=dict(family="Inter, Segoe UI, sans-serif", size=12, color="#333"),
        plot_bgcolor="white",
        hovermode="x unified",
    )
    for ann in fig.layout.annotations:
        ann.font = dict(size=12, color="#1f4e79")
        ann.align = "center"
        # Lift the (now two-line) subplot title above the plot area.
        ann.y = (ann.y if ann.y is not None else 1) + 0.07
        ann.yanchor = "bottom"
    return fig


# =========================================================================
# End-of-Outbreak — Nishiura/Lloyd-Smith offspring-tree simulation
#
# Proper implementation:
#   - Each case generates NB(R, k) offspring (Lloyd-Smith 2005 superspreading,
#     k=0.18 for EBOV).
#   - Each offspring is offset by a Gamma serial-interval draw.
#   - Tree extinction time = max time across all descendants.
#   - P(extinct by T) = fraction of simulations with max time <= T.
#
# Initial unobserved chains seeded from a reporting fraction (under-ascertainment).
# R_t is sampled per realisation from the Cori Gamma posterior — NO arbitrary
# linear R_t decay over time.
# =========================================================================
def compute_eoo_curve(rt_shape_post: float, rt_rate_post: float,
                      si_mean: float, si_sd: float,
                      k_disp: float, max_days: int, n_sim: int = 1000,
                      p_report_lo: float = 0.10, p_report_hi: float = 0.30,
                      seed: int = 42) -> tuple:
    """Simulate descendant trees from the projected last case.

    rt_shape_post, rt_rate_post : Gamma posterior parameters of R_t (Cori).
    si_mean, si_sd              : Gamma SI distribution (days).
    k_disp                      : NB dispersion (Lloyd-Smith 2005, EBOV ~0.18).
    max_days                    : Length of EOO curve to compute (days).
    n_sim                       : Number of stochastic tree simulations.
    p_report_lo/hi              : Uniform range for case-reporting fraction.

    Returns: (days_range, eoo_probs) where probs[i] = P(extinct by day i).
    """
    rng = np.random.default_rng(seed)
    si_shape = (si_mean / si_sd) ** 2
    si_scale = si_sd ** 2 / si_mean

    # Bounded horizon to prevent runaway trees if R > 1 sampled.
    horizon_cap = max_days * 3
    safety_max_active = 50_000

    extinction_times = np.zeros(n_sim)
    for i in range(n_sim):
        # 1. Sample R_t from the Cori Gamma posterior (NO decay).
        R = float(rng.gamma(shape=rt_shape_post,
                             scale=1.0 / max(rt_rate_post, 1e-6)))
        # 2. Sample reporting fraction → initial unobserved chains.
        p_rep = rng.uniform(p_report_lo, p_report_hi)
        n0 = max(1, int(round(1.0 / p_rep)))
        # 3. NB success probability for given (R, k).
        nb_p = k_disp / (k_disp + R) if R > 0 else 1.0

        # Active = list of infection times (initial cases all at t=0).
        active_times = [0.0] * n0
        max_time_seen = 0.0
        runaway = False
        for _ in range(60):  # generation cap
            if not active_times:
                break
            n_active = len(active_times)
            if n_active > safety_max_active:
                runaway = True
                break
            # Vectorised draws for this generation.
            n_off_per_case = rng.negative_binomial(k_disp, nb_p, size=n_active)
            total_off = int(n_off_per_case.sum())
            if total_off == 0:
                break
            si_offsets = rng.gamma(si_shape, si_scale, size=total_off)
            # Parent infection times broadcast to each offspring.
            parent_times = np.repeat(np.asarray(active_times), n_off_per_case)
            offspring_times = parent_times + si_offsets
            # Drop offspring beyond horizon_cap (cannot affect EOO curve).
            mask = offspring_times <= horizon_cap
            active_times = offspring_times[mask].tolist()
            if active_times:
                max_time_seen = max(max_time_seen, float(np.max(active_times)))
        extinction_times[i] = np.inf if runaway else max_time_seen

    days_range = np.arange(0, max_days + 1)
    finite_mask = np.isfinite(extinction_times)
    probs = np.zeros(len(days_range), dtype=float)
    if finite_mask.any():
        finite_times = extinction_times[finite_mask]
        for j, d in enumerate(days_range):
            probs[j] = (finite_times <= d).sum() / n_sim
    return days_range, probs


RESPONSE_LABELS = {
    "S1": "Delayed response",
    "S2": "Moderate response",
    "S3": "Strong response",
}


def eoo_chart_multi(days_range, eoo_probs, scenarios_plc: dict,
                    threshold: float = 0.95) -> go.Figure:
    """Plot EOO probability for all valid scenarios on a calendar-date x-axis.

    Each scenario's curve is the same shape (function of days after its
    projected last case) but shifted to start from that scenario's last-case date.
    """
    fig = go.Figure()

    if not scenarios_plc:
        return fig

    all_dates = []
    for name, plc in scenarios_plc.items():
        dates = [plc + pd.Timedelta(days=int(d)) for d in days_range]
        all_dates.extend(dates)
        label = RESPONSE_LABELS.get(name, name)
        fig.add_trace(go.Scatter(
            x=dates, y=eoo_probs, mode="lines",
            name=f"{label} — last case {plc.strftime('%d %b %Y')}",
            line=dict(color=SCENARIO_COLOURS[name], width=2.6),
            hovertemplate=(
                f"<b>{label}</b><br>%{{x|%d %b %Y}}<br>"
                "P(extinct): %{y:.3f}<extra></extra>"
            ),
        ))

    # Horizontal declaration-threshold line
    x_min, x_max = min(all_dates), max(all_dates)
    fig.add_shape(type="line", xref="x", yref="y",
                  x0=x_min, x1=x_max, y0=threshold, y1=threshold,
                  line=dict(color="#7a7a7a", dash="dot", width=1.4))
    fig.add_annotation(
        xref="paper", yref="y",
        x=0.995, y=threshold, showarrow=False,
        text=f"<b>{int(threshold * 100)}% declaration threshold</b>",
        font=dict(size=10, color="#444"),
        bgcolor="rgba(255,255,255,0.85)",
        xanchor="right", yanchor="bottom",
    )

    fig.update_layout(
        title=dict(text="<b>End-of-outbreak probability</b> — per response scenario",
                   font=dict(size=14, color="#1f4e79"), x=0.01, xanchor="left"),
        xaxis_title="Calendar date",
        yaxis_title="P(outbreak extinct)",
        template="simple_white", hovermode="x unified",
        legend=dict(orientation="h", yanchor="top", y=-0.20,
                    xanchor="center", x=0.5, font=dict(size=11),
                    bgcolor="rgba(255,255,255,0.7)"),
        margin=dict(l=60, r=30, t=60, b=95),
        height=460, plot_bgcolor="white",
        font=dict(family="Inter, Segoe UI, sans-serif", size=12, color="#333"),
        yaxis=dict(range=[0, 1.05]),
    )
    fig.update_xaxes(showgrid=True, gridcolor="#eef1f5", linecolor="#cfd6df")
    fig.update_yaxes(showgrid=True, gridcolor="#eef1f5", linecolor="#cfd6df")
    return fig


# =========================================================================
# Step navigation
# =========================================================================
if "step" not in st.session_state:
    st.session_state["step"] = "data"

# Guards: fall back if pre-reqs not met
if st.session_state["step"] == "rt" and st.session_state.get("result_series") is None:
    st.session_state["step"] = "data"
if st.session_state["step"] == "forecast" and (
    st.session_state.get("result_series") is None
    or st.session_state.get("rt_df") is None
):
    st.session_state["step"] = "data"
if st.session_state["step"] == "eoo" and (
    st.session_state.get("rt_df") is None
    or st.session_state.get("fc_scenarios") is None
):
    st.session_state["step"] = "data"

# Step indicator (4 steps)
step_now = st.session_state["step"]
def _pill(label, active):
    return (
        f'<span style="padding:0.25rem 0.75rem; border-radius:14px; '
        f'background:{"#1f4e79" if active else "#e3e7ed"}; '
        f'color:{"white" if active else "#5b6573"}; font-weight:500;">{label}</span>'
    )
step_html = (
    '<div style="display:flex; gap:0.5rem; margin-bottom:1rem; font-size:0.85rem; '
    'flex-wrap:wrap;">'
    + _pill("Step 1 · Daily incidence", step_now == "data")
    + _pill("Step 2 · R_t estimation", step_now == "rt")
    + _pill("Step 3 · Forecast", step_now == "forecast")
    + _pill("Step 4 · End of outbreak", step_now == "eoo")
    + '</div>'
)
st.markdown(step_html, unsafe_allow_html=True)


# =========================================================================
# Two-column layout: Input (left) | Output (right)
# =========================================================================
left, right = st.columns([1, 1.25], gap="large")

# -------------------------------------------------------------------------
# HELP / GLOSSARY — rendered if step == "help", then st.stop()
# -------------------------------------------------------------------------
if st.session_state["step"] == "help":
    top_l, top_r = st.columns([3, 1])
    with top_l:
        st.markdown('<div class="panel-title">Input glossary</div>',
                    unsafe_allow_html=True)
        st.caption(
            "Every input in this app, explained twice — once for a "
            "modeller (scientific) and once for an MPH student (plain language)."
        )
    with top_r:
        if st.button("← Back to app", use_container_width=True,
                     key="glossary_back"):
            st.session_state["step"] = st.session_state.get("prev_step", "data")
            st.rerun()

    def gentry(name: str, default: str, scientific: str, simple: str):
        st.markdown(
            f'<div style="margin-top:1.1rem; font-weight:600; color:#1f4e79; '
            f'font-size:1.02rem;">{name}</div>',
            unsafe_allow_html=True,
        )
        if default:
            st.markdown(
                f'<div style="font-size:0.78rem; color:#5b6573; '
                f'margin-bottom:0.35rem;">Default: <code>{default}</code></div>',
                unsafe_allow_html=True,
            )
        c1, c2 = st.columns(2)
        with c1:
            st.markdown(
                f'<div style="background:#f6f8fb; border-left:3px solid #1f4e79; '
                f'padding:0.6rem 0.8rem; border-radius:4px; font-size:0.88rem; '
                f'line-height:1.55;">'
                f'<div style="font-size:0.72rem; color:#1f4e79; font-weight:600; '
                f'text-transform:uppercase; letter-spacing:0.05em; '
                f'margin-bottom:0.3rem;">Scientific</div>'
                f'{scientific}'
                f'</div>',
                unsafe_allow_html=True,
            )
        with c2:
            st.markdown(
                f'<div style="background:#fffaf0; border-left:3px solid #c97a1f; '
                f'padding:0.6rem 0.8rem; border-radius:4px; font-size:0.88rem; '
                f'line-height:1.55;">'
                f'<div style="font-size:0.72rem; color:#c97a1f; font-weight:600; '
                f'text-transform:uppercase; letter-spacing:0.05em; '
                f'margin-bottom:0.3rem;">Plain language (MPH)</div>'
                f'{simple}'
                f'</div>',
                unsafe_allow_html=True,
            )

    # ---------- Key indicators & values ----------
    st.markdown(
        '<h3 style="color:#1f4e79; margin-top:1.6rem; '
        'border-bottom:2px solid #1f4e79; padding-bottom:0.3rem;">'
        'Key indicators &amp; values</h3>',
        unsafe_allow_html=True,
    )
    st.caption(
        "Conceptual reference for every quantity used in this tool. "
        "Scientific definitions on the left, plain-language explanations on the right."
    )

    gentry(
        "R<sub>t</sub> — Effective (instantaneous) reproduction number",
        "Time-varying",
        "Expected number of secondary cases generated by a typical infectious "
        "individual at time <i>t</i>, given the current state of the epidemic "
        "(susceptible depletion, interventions, behaviour). Decision-relevant "
        "indicator of whether transmission is growing (R<sub>t</sub> > 1), "
        "stable (R<sub>t</sub> = 1) or declining (R<sub>t</sub> < 1).",
        "How many new people, on average, each infected person passes the "
        "disease to <b>right now</b>. If R<sub>t</sub> = 2, each sick person "
        "infects 2 others on average. Above 1 = outbreak growing; below 1 = "
        "outbreak shrinking.",
    )
    gentry(
        "R<sub>0</sub> — Basic reproduction number",
        "Pathogen-intrinsic",
        "Expected number of secondary cases per primary case in a fully "
        "susceptible population with no interventions. A property of the "
        "pathogen and its setting, not directly used in this tool — but the "
        "starting R<sub>t</sub> in early outbreak phases approximates R<sub>0</sub>.",
        "How contagious the disease is when nobody is immune and nothing is "
        "being done about it. EBOV ≈ 1.5–2.5; measles ≈ 12–18.",
    )
    gentry(
        "Serial interval (SI)",
        "Gamma(15.3 d, 9.3 d) for EBOV",
        "Time between symptom onset in a primary case and onset in a "
        "secondary case directly infected by them. Modelled as a Gamma "
        "distribution. Discretised into per-day weights w<sub>s</sub> for "
        "the renewal equation and R<sub>t</sub> estimator.",
        "On average, how many days between one person getting sick and the "
        "next person they infected getting sick. For Ebola: about 15 days.",
    )
    gentry(
        "Generation interval",
        "Same Gamma as SI here",
        "Time between infection events (primary → secondary). Theoretically "
        "distinct from SI; in this tool we use the SI as a proxy, as is "
        "standard practice in short-term forecasting (Cori 2013, Nouvellet 2018).",
        "Like the serial interval, but counting from the moment of infection "
        "rather than from when symptoms appear. Usually very close in value.",
    )
    gentry(
        "Incidence",
        "Per day or per window",
        "Number of new cases occurring per unit time. The native input for "
        "the renewal equation. May be reported daily or aggregated over a "
        "reporting window (e.g., weekly).",
        "How many <b>new</b> cases happen each day (or week).",
    )
    gentry(
        "Cumulative cases",
        "Running total",
        "Σ of all incident cases from outbreak start through date t. "
        "Convenient for surveillance reporting but must be first-differenced "
        "to recover incidence for modelling.",
        "Total cases counted up to a date, including all earlier days "
        "summed together.",
    )
    gentry(
        "Confirmed case",
        "Lab-positive",
        "WHO case classification: person with laboratory-confirmed EBOV "
        "infection by PCR or rapid diagnostic test.",
        "Person whose lab test came back positive for the disease.",
    )
    gentry(
        "Suspected / probable case",
        "Clinical definition",
        "Person meeting the clinical case definition (e.g., fever + epidemiological "
        "link + WHO symptom criteria) but without laboratory confirmation. Usually "
        "≥ confirmed in count.",
        "Person whose symptoms match the disease but who has not yet been "
        "lab-tested or whose result is pending.",
    )
    gentry(
        "CFR — Case fatality ratio",
        "0.34 for Bundibugyo EBOV",
        "Proportion of confirmed cases who die. Pathogen-and-context specific: "
        "Zaire EBOV 0.40–0.90, Sudan 0.50, Bundibugyo 0.30, Reston ≈ 0. "
        "Applied here as a deterministic multiplier on lagged confirmed cases.",
        "Of every 100 confirmed cases, how many die. For Bundibugyo: about 34.",
    )
    gentry(
        "Onset-to-death lag",
        "10 days median",
        "Median time from symptom onset to death among fatal cases. The "
        "true distribution is right-skewed Gamma (Wamala 2010); applied "
        "here as a deterministic delay between confirmed cases and projected "
        "deaths.",
        "After someone gets sick, how many days until they die (if the "
        "disease is fatal). ~10 days for Ebola.",
    )
    gentry(
        "Maximum incubation period",
        "21 days for EBOV",
        "99th-percentile time from exposure to symptom onset. WHO uses this "
        "to set quarantine windows and the basis of the 42-day EOO rule "
        "(2 × max incubation).",
        "The longest realistic time between getting infected and showing "
        "symptoms. 21 days for Ebola.",
    )
    gentry(
        "95% Credible Interval (CrI)",
        "Bayesian uncertainty",
        "Posterior interval [2.5%, 97.5%]: with the model assumptions, "
        "there is a 95% probability the true parameter lies inside. Distinct "
        "from a frequentist confidence interval.",
        "A range where, given what the model knows, there's a 95% chance "
        "the real value is inside.",
    )
    gentry(
        "Posterior Predictive Interval (PI)",
        "Forecast band",
        "90% interval covering the middle 90% of forecast trajectories drawn "
        "from the posterior of model parameters (here: R<sub>t</sub> starting "
        "value). Wider than a CrI because it also reflects projection-step "
        "variability.",
        "The shaded band on the forecast chart — 90% of likely future "
        "trajectories fall inside it.",
    )
    gentry(
        "Bayesian prior",
        "Gamma(1, 0.2) for R<sub>t</sub>",
        "Probability distribution representing belief about a parameter "
        "<b>before</b> observing data. The conjugate Gamma prior for the "
        "Cori R<sub>t</sub> estimator combines analytically with the "
        "Poisson likelihood.",
        "The model's starting guess about a value before it sees any actual "
        "data. A weak prior lets the data dominate.",
    )
    gentry(
        "Posterior",
        "Updated belief",
        "Probability distribution combining prior and observed-data "
        "likelihood via Bayes' theorem: P(θ | data) ∝ P(data | θ) × P(θ).",
        "The model's updated belief after looking at the data.",
    )
    gentry(
        "Sliding window (τ)",
        "7 days",
        "Time interval over which incidence is aggregated for parameter "
        "estimation. Longer windows give smoother, more stable R<sub>t</sub> "
        "estimates at the cost of timeliness; shorter windows react faster "
        "but are noisier.",
        "How many recent days the model looks at to make each calculation. "
        "Slides forward day by day. Short = jumpy; long = smooth.",
    )
    gentry(
        "Gamma distribution",
        "Shape α, rate β",
        "Continuous probability distribution on positive real numbers with "
        "mean α/β and variance α/β². Standard choice for SI, onset-to-death, "
        "incubation, and Bayesian R<sub>t</sub> posterior.",
        "A flexible bell-shaped mathematical curve that's always positive — "
        "used to describe how long things take.",
    )
    gentry(
        "Negative-Binomial offspring",
        "Mean R, dispersion k",
        "Discrete distribution for offspring count per case. Mean = R "
        "(reproduction number); dispersion k controls right-tail heaviness. "
        "Captures superspreading observed in EBOV (Lloyd-Smith 2005).",
        "A way to count how many new infections each person causes when "
        "transmission is unequal — most people infect few, a handful "
        "(\"superspreaders\") infect many.",
    )
    gentry(
        "Offspring dispersion (k)",
        "0.18 for EBOV",
        "NB dispersion parameter. <b>Small k → more superspreading</b>; "
        "large k → roughly equal transmission across cases. EBOV k ≈ 0.18 "
        "indicates ~20% of cases cause ~80% of transmission.",
        "A number describing how unequal transmission is. Small k (like "
        "0.18 for Ebola) = a few people drive most of the spread.",
    )
    gentry(
        "Renewal equation",
        "I_t = R_t · Σ w_s · I_{t-s}",
        "Today's expected incidence = R<sub>t</sub> multiplied by a "
        "serial-interval-weighted sum of recent incidence. Used both for "
        "R<sub>t</sub> estimation (Cori 2013) and forward projection "
        "(Nouvellet 2018).",
        "A formula that uses recent case counts and R<sub>t</sub> to predict "
        "how many new cases will appear today (and forward).",
    )
    gentry(
        "Branching process",
        "Generation-by-generation simulation",
        "Stochastic process where each case independently produces a random "
        "number of offspring drawn from an offspring distribution; offspring "
        "infect further offspring, and so on. Used here for end-of-outbreak.",
        "A computer simulation where each infected person produces a random "
        "number of next-generation cases, and we follow the chain until it "
        "dies out (or runs away).",
    )
    gentry(
        "End-of-outbreak (EOO) probability",
        "Time-dependent",
        "P(no active infectious chains exist at time T after the last "
        "reported case), conditional on observed data and the offspring + "
        "SI distributions. Declared when this probability crosses a "
        "threshold (commonly 95%).",
        "The chance that the outbreak has truly stopped at a given day "
        "after the last reported case.",
    )
    gentry(
        "WHO 42-day rule",
        "Heuristic",
        "Operational declaration: outbreak considered over 42 days "
        "(= 2 × maximum 21-day incubation) after the last confirmed case "
        "with two consecutive negative PCR tests. Heuristic — not a "
        "probabilistic statement.",
        "WHO declares an Ebola outbreak over once 42 days have passed with "
        "no new confirmed cases.",
    )
    gentry(
        "Djaafara framework (63 + 90 days)",
        "Reporting-adjusted",
        "Statistical EOO framework (Djaafara et al. 2021, <i>BMC Med</i>) "
        "with explicit reporting-rate adjustment. <b>Preliminary</b> "
        "declaration at 63 days; <b>final</b> declaration after an additional "
        "90 days of clear surveillance.",
        "A more conservative set of EOO rules: declare preliminary after "
        "63 days, then wait another 90 days for the final all-clear.",
    )

    # ---------- Step 1 ----------
    st.markdown(
        '<h3 style="color:#1f4e79; margin-top:1.6rem; '
        'border-bottom:2px solid #1f4e79; padding-bottom:0.3rem;">'
        'Step 1 — Daily incidence builder</h3>',
        unsafe_allow_html=True,
    )

    gentry(
        "Source attribution mode",
        "Single source for whole table",
        "Provenance metadata is a reproducibility requirement (EPIFORGE item 11). "
        "Per-row attribution supports mixed-source datasets (e.g., merging WHO DONs "
        "with national surveillance bulletins, where each row may have a different citation).",
        "A way to record where each row of data came from. Pick <b>single source</b> if "
        "every number is from one WHO bulletin; pick <b>per-row sources</b> if you mixed "
        "different reports.",
    )
    gentry(
        "Values type — Cumulative vs Incidence",
        "Cumulative",
        "Cumulative refers to running totals (Σ cases up to date t). Incidence is the "
        "count of new cases at time t (or over a reporting interval). Renewal-equation "
        "estimators consume incidence; cumulative is first-differenced. Linear interpolation "
        "is applied between sparse cumulative snapshots.",
        "<b>Cumulative</b> = total cases so far (\"134 cases as of May 29\"). "
        "<b>Incidence</b> = new cases per day (or per reporting window). The model needs "
        "daily incidence — cumulative inputs are converted for you.",
    )
    gentry(
        "Confirmed / Suspected / Deaths columns",
        "",
        "Standard WHO case classification. <b>Confirmed</b> = laboratory-confirmed "
        "(PCR or RDT positive). <b>Suspected</b> = clinical case definition compatible "
        "with EVD without lab confirmation. <b>Deaths</b> = deaths attributed to EVD "
        "(confirmed cause).",
        "<b>Confirmed</b> = lab-tested positive. <b>Suspected</b> = symptoms look right "
        "but no lab confirmation yet. <b>Deaths</b> = people who died from the disease.",
    )
    gentry(
        "Source / DOI per row",
        "",
        "Digital Object Identifier or persistent URL for the originating data source. "
        "Required for citation, traceability, and downstream meta-analysis.",
        "The web link or paper reference for that row's data. Lets reviewers verify "
        "your numbers.",
    )
    gentry(
        "CFR Backcalculation (optional)",
        "CFR 24% / 30% / 40% / Custom; SMA None / 3 / 5 / 7-day",
        "Estimates incidence from observed deaths under the assumption that "
        "deaths are better counted than cases. Cumulative deaths divided by the CFR "
        "gives a naive estimate; the phase-bias multiplier "
        "(1 + r/β)<sup>α</sup> corrects it for the fact that during exponential "
        "growth most recent cases have not yet died. r = ln(2)/t<sub>2</sub> is the "
        "epidemic growth rate; (α, β) are the Gamma onset-to-death shape and rate. "
        "The user picks whether the resulting series substitutes for confirmed, "
        "suspected, or both (Total cases) in Steps 2–4. (Garske et al. 2009; "
        "Imperial College methodology.)",
        "A way to estimate the true number of cases by working backwards from deaths. "
        "If <b>X</b>% of cases die, then dividing deaths by <b>X</b>% gives an "
        "estimate of total cases — adjusted for the fact that recent cases haven't "
        "had time to die yet. Useful when you suspect cases are under-reported.",
    )

    # ---------- Step 2 ----------
    st.markdown(
        '<h3 style="color:#1f4e79; margin-top:1.6rem; '
        'border-bottom:2px solid #1f4e79; padding-bottom:0.3rem;">'
        'Step 2 — R<sub>t</sub> estimation</h3>',
        unsafe_allow_html=True,
    )

    gentry(
        "SI mean (primary)",
        "15.3 days",
        "Mean of the serial-interval distribution — the expected time between symptom "
        "onset in an infector and onset in their infectee. Modelled here as Gamma(α, β) "
        "with mean μ = α/β. Default value is the Zaire EBOV proxy from WHO Ebola "
        "Response Team 2014 (NEJM).",
        "On average, how many days pass between one person showing symptoms and the "
        "next person they infected showing symptoms. For Ebola, about <b>15 days</b>.",
    )
    gentry(
        "SI SD (primary)",
        "9.3 days",
        "Standard deviation of the Gamma serial-interval distribution. Quantifies "
        "generation-time heterogeneity. Affects both R<sub>t</sub> estimation and the "
        "discretised SI weights used in the renewal equation.",
        "How much the gap between infections varies. A bigger SD means some people "
        "pass it on much faster or much slower than the average 15 days.",
    )
    gentry(
        "Sensitivity SI (mean / SD)",
        "12 d / 5 d (Legrand 2007)",
        "Alternative SI parameterisation used for sensitivity analysis. Overlaying a "
        "second R<sub>t</sub> trajectory shows whether conclusions are robust to SI "
        "mis-specification — a recommended EPIFORGE-style robustness check.",
        "A backup guess at the serial interval, plotted alongside the main one so "
        "you can see if the answer changes much when the assumption changes.",
    )
    gentry(
        "Sliding window (τ)",
        "7 days",
        "Aggregation window for the Cori et al. (2013) instantaneous R<sub>t</sub> "
        "estimator. Trade-off: longer τ → smoother estimates, more lag and lower "
        "temporal resolution; shorter τ → reactive estimates with higher variance. "
        "Window is right-aligned (estimate corresponds to the rightmost day).",
        "How many days the model looks back to estimate R<sub>t</sub>. <b>Short</b> "
        "(3 d) = jumpy estimates that react fast; <b>long</b> (14 d) = smooth "
        "estimates that lag behind real changes.",
    )
    gentry(
        "Prior shape (α₀) & rate (β₀)",
        "shape = 1, rate = 0.2",
        "Hyperparameters of the conjugate Gamma prior on R<sub>t</sub>. The Cori "
        "EpiEstim default (1, 0.2) gives prior mean R = 5, prior SD = 5 — weakly "
        "informative. Posterior: Gamma(α₀ + I<sub>t</sub>, β₀ + Λ<sub>t</sub>), "
        "where Λ<sub>t</sub> = Σ w<sub>s</sub> I<sub>t&minus;s</sub>.",
        "The model's starting guess about R<sub>t</sub> before seeing any data. The "
        "default expects R<sub>t</sub> around 5 with lots of uncertainty — meaning the "
        "real data dominates the answer.",
    )
    gentry(
        "Sensitivity overlay (checkbox)",
        "On",
        "Toggles the second SI-based R<sub>t</sub> trajectory on the plot for robustness "
        "comparison.",
        "Turns on/off the second R<sub>t</sub> line drawn with the backup serial interval.",
    )
    gentry(
        "R<sub>t</sub> for forecast (Latest / Mean 7d / Median / Mean all / Custom)",
        "Latest (recommended)",
        "Summary statistic of the R<sub>t</sub> posterior trajectory used to seed the "
        "renewal-equation projection. The latest right-aligned estimate is preferred "
        "for forward forecasts; means/medians smooth recent noise at the cost of timeliness.",
        "Which R<sub>t</sub> value to plug into the forecast. <b>Latest</b> = most "
        "recent — best for predicting forward. The others smooth across time if the "
        "latest number looks too noisy.",
    )

    # ---------- Step 3 ----------
    st.markdown(
        '<h3 style="color:#1f4e79; margin-top:1.6rem; '
        'border-bottom:2px solid #1f4e79; padding-bottom:0.3rem;">'
        'Step 3 — Renewal-equation forecast</h3>',
        unsafe_allow_html=True,
    )

    gentry(
        "Start date",
        "Last date of input data",
        "Anchor date for the projection. The renewal equation "
        "I<sub>t</sub> = R<sub>t</sub> · Σ w<sub>s</sub> I<sub>t&minus;s</sub> projects "
        "forward from this date using the seed incidence history.",
        "The day the forecast starts counting forward — usually the date of your "
        "most recent data.",
    )
    gentry(
        "Horizon (days)",
        "365",
        "Number of days projected forward. Long horizons amplify R<sub>t</sub> "
        "mis-specification; in practice operational forecasts cover 30–365 days.",
        "How many days into the future to predict. Longer horizons are less reliable.",
    )
    gentry(
        "Observed baseline (confirmed / suspected / deaths)",
        "Auto-filled from Step 1",
        "Cumulative counts at the start date used to anchor cumulative projection "
        "trajectories. Editable to allow correction or override.",
        "How many cases you've already seen. The forecast adds new cases on top "
        "of these.",
    )
    gentry(
        "CFR (Case Fatality Ratio)",
        "0.34 (Wamala 2010, Bundibugyo)",
        "Proportion of confirmed cases that result in death. EBOV CFR varies by species: "
        "Zaire ~0.40–0.90, Sudan ~0.50, Bundibugyo ~0.30, Reston ~0. Default reflects "
        "Bundibugyo 2007 (Wamala 2010, CDC EID).",
        "Out of every 100 confirmed cases, how many die. For Bundibugyo Ebola, about "
        "<b>34 die per 100</b>.",
    )
    gentry(
        "Onset-to-death lag (days)",
        "10 days (Wamala 2010)",
        "Median time from symptom onset to death. Applied as a deterministic delay: "
        "D<sub>t</sub> = CFR · I<sub>t&minus;lag</sub><sup>confirmed</sup>. A delta "
        "lag underestimates the right-skewed onset-to-death distribution.",
        "After someone gets sick, how many days until they die (if they die). About "
        "<b>10 days</b> for Ebola.",
    )
    gentry(
        "R<sub>t</sub> scenario — target & days to target",
        "S1: 0.9/180d · S2: 0.9/90d · S3: 0.6/30d",
        "Each scenario defines a linear R<sub>t</sub> trajectory: starts at the user-"
        "selected R<sub>t</sub>, moves linearly to <code>target</code> over "
        "<code>days_to_target</code> days, then holds. Represents an assumption "
        "about response speed; not a prediction.",
        "Each scenario asks: \"What if the response brings R<sub>t</sub> down to "
        "<b>X</b> over <b>Y</b> days?\" Faster, deeper control = smaller outbreak.",
    )

    # ---------- Step 4 ----------
    st.markdown(
        '<h3 style="color:#1f4e79; margin-top:1.6rem; '
        'border-bottom:2px solid #1f4e79; padding-bottom:0.3rem;">'
        'Step 4 — End-of-outbreak predictor</h3>',
        unsafe_allow_html=True,
    )

    gentry(
        "Simulations",
        "1000",
        "Number of independent Monte-Carlo descendant-tree realisations. Higher N "
        "tightens the Monte-Carlo error on P(extinct). Compute scales O(N · "
        "generations).",
        "How many random \"what if\" runs the computer does. More runs = more precise "
        "probability, but slower.",
    )
    gentry(
        "Days range",
        "180 days",
        "Time horizon (days after projected last case) over which P(extinct by day t) "
        "is evaluated.",
        "How far ahead to track the probability that the outbreak has truly ended.",
    )
    gentry(
        "Declaration threshold (P)",
        "0.95",
        "Probability level at which end-of-outbreak is declared. WHO operational "
        "rule (42-day rule) is heuristic; the 95% threshold is the standard "
        "statistical convention.",
        "How sure you want to be before saying \"the outbreak is over\". Default "
        "<b>95%</b> = \"we're 95% confident\".",
    )
    gentry(
        "Offspring dispersion (k)",
        "0.18 (Lloyd-Smith 2005)",
        "Dispersion parameter of the Negative-Binomial offspring distribution. Lower "
        "k → more superspreading (heavy right tail in secondary cases). EBOV ≈ 0.18 "
        "(Lloyd-Smith et al. 2005, Nature).",
        "A number describing how unequal transmission is. Small k means most people "
        "infect nobody, but a few <b>superspreaders</b> infect many. EBOV has small k.",
    )

    st.stop()


# -------------------------------------------------------------------------
# STEP 4 — End-of-outbreak predictor (rendered if step == "eoo", then st.stop())
# -------------------------------------------------------------------------
if st.session_state["step"] == "eoo":
    scenarios_out = st.session_state["fc_scenarios"]
    horizon_dates = st.session_state["fc_dates"]
    rt_df = st.session_state["rt_df"]
    # Stable SI keys saved by Step 2 (see Step 2 "Estimate R_t"). Using the
    # widget key here previously made this filter miss every row when the user
    # changed the SI, producing a false "No R_t estimates" error.
    si_mean_used = float(st.session_state.get("rt_si_mean", 15.3))

    # Latest R_t — use the Gamma posterior parameters directly (no Normal SD approx).
    primary = rt_df[rt_df["si_mean_used"] == si_mean_used].dropna(
        subset=["rt_mean"])
    if primary.empty:
        st.warning("No R_t estimates available. Re-run Step 2 first.")
        st.stop()
    latest = primary.iloc[-1]
    rt_mean_latest = float(latest["rt_mean"])
    rt_shape_post = float(latest.get("shape_post", 1.0 + rt_mean_latest))
    rt_rate_post = float(latest.get("rate_post",
                                     max(1e-6, rt_shape_post / rt_mean_latest)))

    # Anchor on the R_t the user selected in Step 2 (Latest / Mean 7d /
    # Median / Mean all / Custom), not the time-series last row. We keep
    # the latest window's posterior *dispersion* and rescale so the mean
    # matches selected_rt — same trick used by Step 3 (Forecast) at the
    # rt_samples block. For Gamma(shape, rate), scaling X by c gives
    # Gamma(shape, rate / c); choosing c = selected_rt / rt_mean_latest
    # moves the mean to selected_rt while preserving shape.
    rt_selected = float(st.session_state.get("selected_rt", rt_mean_latest))
    rt_basis = st.session_state.get("selected_rt_basis", "Latest")
    if rt_mean_latest > 0 and rt_selected > 0:
        _scale = rt_selected / rt_mean_latest
        rt_rate_post = rt_rate_post / _scale
    si_sd_used = float(st.session_state.get("rt_si_sd", 9.3))

    def projected_last_case(scen_key: str):
        data = scenarios_out[scen_key]
        series = data["new_confirmed"]
        if isinstance(series, dict):
            series = series["median"]
        above = np.where(np.asarray(series) >= 0.5)[0]
        if len(above) == 0 or above[-1] >= len(horizon_dates) - 1:
            return None
        return horizon_dates[above[-1]]

    left4, right4 = st.columns([1, 1.25], gap="large")

    with left4:
        st.markdown('<div class="panel-title">EOO inputs</div>',
                    unsafe_allow_html=True)

        st.markdown('<div class="section-label">Locked from previous steps</div>',
                    unsafe_allow_html=True)
        # CrI bounds from the latest window, rescaled by the same factor
        # that moves the mean to rt_selected (Gamma quantiles scale linearly).
        _cri_scale = (rt_selected / rt_mean_latest
                      if rt_mean_latest > 0 else 1.0)
        _cri_lower = float(latest["rt_lower"]) * _cri_scale
        _cri_upper = float(latest["rt_upper"]) * _cri_scale
        f1, f2 = st.columns(2)
        f1.markdown(
            f'<div style="border:1px solid #d8dde4; background:#f6f8fb; '
            f'border-radius:6px; padding:0.5rem 0.7rem;">'
            f'<div style="font-size:0.7rem; color:#5b6573; '
            f'text-transform:uppercase; letter-spacing:0.05em;">'
            f'🔒 Selected R_t (Step 2)</div>'
            f'<div style="font-size:0.95rem; font-weight:600; color:#1f4e79; '
            f'margin-top:0.15rem;">{rt_selected:.2f}</div>'
            f'<div style="font-size:0.72rem; color:#5b6573;">'
            f'basis: {rt_basis} · 95% CrI (rescaled): '
            f'{_cri_lower:.2f}–{_cri_upper:.2f}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )
        f2.markdown(
            f'<div style="border:1px solid #d8dde4; background:#f6f8fb; '
            f'border-radius:6px; padding:0.5rem 0.7rem;">'
            f'<div style="font-size:0.7rem; color:#5b6573; '
            f'text-transform:uppercase; letter-spacing:0.05em;">'
            f'R_t Gamma posterior (rescaled)</div>'
            f'<div style="font-size:0.95rem; font-weight:600; color:#1f4e79; '
            f'margin-top:0.15rem;">'
            f'shape={rt_shape_post:.2f}, rate={rt_rate_post:.2f}</div>'
            f'<div style="font-size:0.72rem; color:#5b6573;">'
            f'mean anchored on selected R_t; latest-window dispersion '
            f'preserved · sampled per realisation</div>'
            f'</div>',
            unsafe_allow_html=True,
        )

        st.markdown('<div class="section-label">Projected last case per scenario</div>',
                    unsafe_allow_html=True)
        valid_plc = {}
        for s_name in scenarios_out:
            plc = projected_last_case(s_name)
            label = RESPONSE_LABELS.get(s_name, s_name)
            if plc is not None:
                valid_plc[s_name] = plc
                st.markdown(
                    f'<div style="font-size:0.82rem; margin:0.25rem 0;">'
                    f'<span style="color:{SCENARIO_COLOURS[s_name]};">●</span> '
                    f'<b>{label}</b> · projected last case: '
                    f'<b>{plc.strftime("%d %b %Y")}</b>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            else:
                st.markdown(
                    f'<div style="font-size:0.82rem; margin:0.25rem 0; '
                    f'color:#B22222;">'
                    f'<span>●</span> '
                    f'<b>{label}</b> · '
                    f'outbreak still active at horizon end (no end date).'
                    f'</div>',
                    unsafe_allow_html=True,
                )

        if not valid_plc:
            st.warning(
                "No scenario tails off below 0.5 cases/day within the forecast "
                "horizon. Go back to Step 3 and extend the horizon to find end dates."
            )

        st.markdown('<div class="section-label">Simulation settings</div>',
                    unsafe_allow_html=True)
        s1, s2 = st.columns(2)
        with s1:
            n_sim = st.number_input(
                "Simulations", min_value=100, max_value=5000,
                value=st.session_state.get("eoo_n_sim", 1000), step=100,
                key="eoo_n_sim",
                help="Independent stochastic descendant-tree simulations.",
            )
        with s2:
            max_days = st.number_input(
                "Days range", min_value=42, max_value=365,
                value=st.session_state.get("eoo_max_days", 180), step=10,
                key="eoo_max_days",
                help="Compute P(extinct) from day 0 to this many days after "
                     "the projected last case.",
            )
        s3, s4 = st.columns(2)
        with s3:
            threshold = st.slider(
                "Declaration threshold (P)", 0.50, 0.99,
                st.session_state.get("eoo_threshold_input", 0.95), 0.01,
                key="eoo_threshold_input",
                help="EOO is declared when P(extinct) exceeds this.",
            )
        with s4:
            k_disp = st.number_input(
                "Offspring dispersion k", min_value=0.05, max_value=10.0,
                value=st.session_state.get("eoo_k_disp", 0.18),
                step=0.01, format="%.2f", key="eoo_k_disp",
                help="NB dispersion. Small k = more superspreading. "
                     "EBOV ≈ 0.18 (Lloyd-Smith 2005, Nature).",
            )
        run_eoo = st.button("Run EOO simulation", type="primary",
                            use_container_width=True,
                            disabled=len(valid_plc) == 0)

        if run_eoo and len(valid_plc) > 0:
            with st.spinner("Running descendant-tree simulations…"):
                days_range, eoo_probs = compute_eoo_curve(
                    rt_shape_post=rt_shape_post,
                    rt_rate_post=rt_rate_post,
                    si_mean=si_mean_used,
                    si_sd=si_sd_used,
                    k_disp=float(k_disp),
                    max_days=int(max_days),
                    n_sim=int(n_sim),
                )
            st.session_state["eoo_days"] = days_range
            st.session_state["eoo_probs"] = eoo_probs
            st.session_state["eoo_valid_plc"] = valid_plc
            st.session_state["eoo_threshold"] = float(threshold)

    with right4:
        st.markdown('<div class="freeze-right"></div>', unsafe_allow_html=True)
        top_row = st.columns([1, 0.18])
        with top_row[0]:
            st.markdown('<div class="panel-title">EOO output</div>',
                        unsafe_allow_html=True)
        with top_row[1]:
            if st.button("← Back", use_container_width=True, key="back_to_fc"):
                st.session_state["step"] = "forecast"
                st.rerun()

        days_range = st.session_state.get("eoo_days")
        eoo_probs = st.session_state.get("eoo_probs")
        valid_plc_used = st.session_state.get("eoo_valid_plc", {})
        thr = st.session_state.get("eoo_threshold", 0.95)

        if days_range is None or eoo_probs is None:
            st.markdown(
                '<div class="placeholder-card">'
                'Set the simulation settings and click '
                '<b>Run EOO simulation</b>. You will see one EOO probability '
                'curve per response scenario (S1, S2, S3) on a calendar date '
                'axis, plus the WHO 42-day and Djaafara 63-day declaration dates.'
                '</div>',
                unsafe_allow_html=True,
            )
        else:
            _fig_eoo = eoo_chart_multi(days_range, eoo_probs, valid_plc_used, thr)
            st.session_state["chart_eoo"] = _fig_eoo
            chart_with_line_filter(_fig_eoo, key="eoo")

            # --- Per-scenario key dates summary ---
            cross_idx = np.where(eoo_probs >= thr)[0]
            cross_day = int(days_range[cross_idx[0]]) if len(cross_idx) > 0 else None

            for s_name, plc_used in valid_plc_used.items():
                who_date = plc_used + pd.Timedelta(days=42)
                djaafara_prelim = plc_used + pd.Timedelta(days=63)
                djaafara_final = djaafara_prelim + pd.Timedelta(days=90)

                if cross_day is not None:
                    cross_date = plc_used + pd.Timedelta(days=cross_day)
                    cross_line = (
                        f'P(extinct) ≥ {thr:.0%}: <b style="color:#1f4e79;">'
                        f'{cross_date.strftime("%d %b %Y")}</b> '
                        f'(day {cross_day} after last case)'
                    )
                else:
                    cross_line = (
                        f'<span style="color:#B22222;">P(extinct) does not reach '
                        f'{thr:.0%} within {int(days_range[-1])} days — '
                        f'extend the range.</span>'
                    )

                st.markdown(
                    f'<div style="padding:0.6rem 0.9rem; '
                    f'border-left:4px solid {SCENARIO_COLOURS[s_name]}; '
                    f'background:#fafbfc; margin:0.4rem 0; font-size:0.85rem; '
                    f'line-height:1.55;">'
                    f'<b style="color:{SCENARIO_COLOURS[s_name]};">'
                    f'{RESPONSE_LABELS.get(s_name, s_name)}</b> · '
                    f'projected last case <b>{plc_used.strftime("%d %b %Y")}</b><br>'
                    f'WHO 42-day: <b style="color:#E07B39;">'
                    f'{who_date.strftime("%d %b %Y")}</b>  '
                    f'· Djaafara prelim (63 d): <b style="color:#1E8449;">'
                    f'{djaafara_prelim.strftime("%d %b %Y")}</b>  '
                    f'· Djaafara final (+90 d): <b style="color:#1E8449;">'
                    f'{djaafara_final.strftime("%d %b %Y")}</b><br>'
                    f'{cross_line}'
                    f'</div>',
                    unsafe_allow_html=True,
                )

            with st.expander("Show table / data", expanded=False):
                # Long-format CSV: one row per (scenario, day)
                rows = []
                for s_name, plc_used in valid_plc_used.items():
                    for i, d in enumerate(days_range):
                        rows.append({
                            "scenario": s_name,
                            "days_after_last_case": int(d),
                            "projected_date": (plc_used + pd.Timedelta(
                                days=int(d))).strftime("%Y-%m-%d"),
                            "eoo_probability": round(float(eoo_probs[i]), 4),
                            "who_42day_met": int(d) >= 42,
                            "djaafara_63day_met": int(d) >= 63,
                        })
                eoo_df = pd.DataFrame(rows)
                preview = eoo_df.copy()
                preview["projected_date"] = pd.to_datetime(
                    preview["projected_date"]).dt.strftime("%d %b %Y")
                st.dataframe(preview, use_container_width=True, height=320,
                             hide_index=True)
                slug = _slug(st.session_state.get("scenario_name", ""))
                st.download_button(
                    "Download eoo_probability.csv",
                    eoo_df.to_csv(index=False).encode("utf-8"),
                    f"{slug}__eoo_probability.csv", "text/csv",
                    use_container_width=True,
                )

    # Method note + equations
    with st.expander("📐 Method & equations — Nishiura / Lloyd-Smith branching",
                      expanded=False):
        st.markdown("**Step 1 — R_t drawn from the Cori 2013 Gamma posterior:**")
        st.latex(r"R \;\sim\; \mathrm{Gamma}(\alpha + I_{[t-\tau+1,t]},\; "
                  r"\beta + \Lambda_{[t-\tau+1,t]})")
        st.caption(
            "Each tree gets its own R drawn from the posterior — propagates Step 2 "
            "uncertainty into the EOO calculation."
        )

        st.markdown("**Step 2 — Initial unobserved chains:**")
        st.latex(r"p_{\rm rep} \sim \mathrm{Uniform}(0.10,\,0.30) "
                  r"\qquad "
                  r"n_0 = \left\lceil \dfrac{1}{p_{\rm rep}} \right\rceil")
        st.caption(
            "If only 1 in p_rep cases is reported, we expect n_0 hidden infectious "
            "chains active at the last reported case date."
        )

        st.markdown("**Step 3 — Negative-Binomial offspring (per case):**")
        st.latex(r"X \;\sim\; \mathrm{NB}\!\left(k,\; \dfrac{k}{k + R}\right),"
                  r"\quad \mathbb{E}[X] = R")
        st.caption(
            "k is the dispersion (Lloyd-Smith 2005; EBOV k ≈ 0.18). Small k = "
            "heavy right tail = superspreading."
        )

        st.markdown("**Step 4 — Generation time per offspring:**")
        st.latex(r"\tau \;\sim\; \mathrm{Gamma}(\mu_{SI},\, \sigma_{SI}^2) "
                  r"\qquad t_{\rm offspring} = t_{\rm parent} + \tau")
        st.caption("Same Gamma SI as Step 2.")

        st.markdown("**Step 5 — End-of-outbreak probability:**")
        st.latex(r"P\bigl(\text{extinct by day } T\bigr) = "
                  r"\dfrac{1}{N_{\rm sim}}\sum_{i=1}^{N_{\rm sim}}"
                  r"\mathbf{1}\!\left[ \max_{j} t^{(i)}_{j} \le T \right]")
        st.caption(
            "Fraction of simulated trees whose last descendant occurred by day T "
            "after the projected last case."
        )

        st.markdown(
            "**Declaration thresholds**: WHO **42 days** (2 × max incubation 21 d). "
            "**Djaafara et al. 2021** uses a more conservative **63-day preliminary** + "
            "**+90-day final** with explicit reporting-rate adjustment. The **95% "
            "probability threshold** is the standard cut-off for high confidence."
        )

    st.stop()


# -------------------------------------------------------------------------
# STEP 3 — Forecast page (rendered if step == "forecast", then st.stop())
# -------------------------------------------------------------------------
if st.session_state["step"] == "forecast":
    series_in = st.session_state["result_series"]
    rt_starting = float(st.session_state.get("selected_rt", 4.0))
    rt_basis = st.session_state.get("selected_rt_basis", "Latest")
    # Read the serial interval from the stable keys saved by Step 2's
    # "Estimate R_t" (not the widget keys, which are dropped on navigation).
    si_mean_used = float(st.session_state.get("rt_si_mean", 15.3))
    si_sd_used = float(st.session_state.get("rt_si_sd", 9.3))
    si_src_used = st.session_state.get("si_mean_primary_src",
                                       "https://doi.org/10.1056/NEJMoa1411100")

    series_dates = pd.to_datetime(series_in["date"])
    last_input_date = series_dates.max()

    # Baseline defaults are computed *after* the start_date input below so
    # they track the user's chosen projection start. See "Observed baseline
    # at start" block.

    DEFAULT_FC_SRC = {
        "cfr": "https://doi.org/10.3201/eid1607.090536",
        "lag": "https://doi.org/10.3201/eid1607.090536",
    }
    DEFAULT_FC_LABEL = {
        "cfr": "Wamala 2010, CDC Emerging Infect Dis (Bundibugyo 2007)",
        "lag": "Wamala 2010, CDC Emerging Infect Dis (Bundibugyo 2007)",
    }

    def frozen_card(label: str, value: str, sub: str = "") -> str:
        sub_html = (f'<div style="font-size:0.72rem; color:#5b6573; '
                    f'margin-top:0.1rem;">{sub}</div>') if sub else ""
        return (
            f'<div style="border:1px solid #d8dde4; background:#f6f8fb; '
            f'border-radius:6px; padding:0.5rem 0.7rem;">'
            f'<div style="font-size:0.7rem; color:#5b6573; '
            f'text-transform:uppercase; letter-spacing:0.05em;">'
            f'🔒 {label}</div>'
            f'<div style="font-size:0.95rem; font-weight:600; '
            f'color:#1f4e79; margin-top:0.15rem;">{value}</div>'
            f'{sub_html}'
            f'</div>'
        )

    left3, right3 = st.columns([1, 1.25], gap="large")

    with left3:
        st.markdown('<div class="panel-title">Forecast inputs</div>',
                    unsafe_allow_html=True)

        # --- Frozen inputs ---
        st.markdown('<div class="section-label">Locked from previous steps</div>',
                    unsafe_allow_html=True)
        f1, f2 = st.columns(2)
        f1.markdown(frozen_card(
            "Daily series",
            f"{len(series_in)} days",
            f"{series_dates.min():%d %b %Y} → {last_input_date:%d %b %Y}",
        ), unsafe_allow_html=True)
        f2.markdown(frozen_card(
            "R_t starting value",
            f"{rt_starting:.2f}",
            f"basis: {rt_basis}",
        ), unsafe_allow_html=True)
        st.markdown(
            f'<div style="margin-top:0.4rem;">'
            f'{frozen_card("Serial interval", f"Gamma({si_mean_used:.1f}, {si_sd_used:.1f}) d", "")}'
            f'<div style="margin:-0.35rem 0 0.3rem 0; font-size:0.75rem;">'
            f'<a href="{si_src_used}" target="_blank" rel="noopener" '
            f'style="color:#1f4e79; text-decoration:none; '
            f'border-bottom:1px dotted #1f4e79;">{si_src_used}</a></div>'
            f'</div>',
            unsafe_allow_html=True,
        )

        # --- Forecast window ---
        st.markdown('<div class="section-label">Forecast window</div>',
                    unsafe_allow_html=True)
        w1, w2 = st.columns(2)
        with w1:
            start_date = st.date_input(
                "Start date",
                value=st.session_state.get("forecast_start_date",
                                            last_input_date.date()),
                help="Projection starts on the day AFTER this date.",
                key="forecast_start_date",
            )
        with w2:
            horizon = st.number_input(
                "Horizon (days)", min_value=7, max_value=720,
                value=st.session_state.get("forecast_horizon", 365), step=7,
                key="forecast_horizon",
            )

        # --- Baseline observed cumulative ---
        # Look up cumulative observed at the SELECTED start_date (not just
        # the last row of the series). This fixes two prior bugs:
        #   (a) Baseline ignored start_date — earlier start dates kept
        #       showing the end-of-series cumulative.
        #   (b) Stale session state — re-uploading data left old user
        #       inputs in place because Streamlit cached the prior value.
        # When the data-derived defaults change (data re-uploaded OR
        # start_date moved), we clear the cached user inputs so the fields
        # re-seed from the new data. User edits after that point persist
        # until the next data/start_date change.
        _start_ts = pd.Timestamp(start_date)
        _eligible = series_in[pd.to_datetime(series_in["date"]) <= _start_ts]
        _base_row = (_eligible.iloc[-1] if len(_eligible)
                      else series_in.iloc[0])
        # Decide the Confirmed and Suspected baselines INDEPENDENTLY so the
        # seed and the baseline always come from the same series (otherwise
        # their scales would differ). Precedence — kept identical to the
        # seeding block and the source labels below; all three must agree.
        #   Confirmed: CFR estimate if CFR routes there (Confirmed/Total),
        #              else observed confirmed.
        #   Suspected: CFR estimate if CFR routes to Suspected; else the
        #              ratio-based Estimated Suspected (Step 1) if active;
        #              else 0 when CFR took everything as Total; else observed.
        _cfr_role_fc = st.session_state.get("cfr_role_active", "Total cases")
        _cfr_on = (st.session_state.get("cfr_active")
                   and "cumulative_cfr_estimated" in series_in.columns)
        _susp_est_on = (st.session_state.get("suspest_active")
                        and "cumulative_suspected_estimated" in series_in.columns)
        _cfr_cum = float(_base_row.get("cumulative_cfr_estimated", 0)) if _cfr_on else 0.0

        if _cfr_on and _cfr_role_fc in ("Confirmed", "Total cases"):
            base_conf_default = int(round(_cfr_cum))
        else:
            base_conf_default = int(round(float(_base_row.get("cumulative_confirmed", 0))))

        if _cfr_on and _cfr_role_fc == "Suspected":
            base_susp_default = int(round(_cfr_cum))
        elif _susp_est_on:
            base_susp_default = int(round(float(
                _base_row.get("cumulative_suspected_estimated", 0))))
        elif _cfr_on and _cfr_role_fc == "Total cases":
            base_susp_default = 0  # CFR total folded into Confirmed
        else:
            base_susp_default = int(round(float(_base_row.get("cumulative_suspected", 0))))
        base_death_default = int(round(float(
            _base_row.get("cumulative_deaths", 0))))

        # Plain-language labels for which series feeds each slot — shown to the
        # user below so the forecast seeding is never a black box. These MUST
        # track the precedence used for the baseline and the seeds above.
        _conf_src_label = (
            "Estimated Confirmed (CFR back-calc)"
            if _cfr_on and _cfr_role_fc in ("Confirmed", "Total cases")
            else "Observed confirmed")
        if _cfr_on and _cfr_role_fc == "Suspected":
            _susp_src_label = "Estimated via CFR back-calc"
        elif _susp_est_on:
            _susp_src_label = "Estimated Suspected (ratio, from Step 1)"
        elif _cfr_on and _cfr_role_fc == "Total cases":
            _susp_src_label = "Not used (0 — CFR total folded into Confirmed)"
        else:
            _susp_src_label = "Observed suspected"
        _baseline_sig = (base_conf_default, base_susp_default,
                          base_death_default)
        if st.session_state.get("fc_baseline_sig") != _baseline_sig:
            for _k in ("obs_conf_input", "obs_susp_input", "obs_death_input"):
                st.session_state.pop(_k, None)
            st.session_state["fc_baseline_sig"] = _baseline_sig

        st.markdown('<div class="section-label">Observed baseline at start</div>',
                    unsafe_allow_html=True)
        st.caption(
            f"Auto-filled from your data on {_start_ts:%d %b %Y}. "
            "Edit if you need to override."
        )
        st.caption(
            f"**Forecast is seeded from** — Confirmed: {_conf_src_label} · "
            f"Suspected: {_susp_src_label} · "
            f"Deaths: computed from cases via the CFR below."
        )

        # --- Suspected-zeroed notice (problem #6) ---
        # With the DEFAULT CFR role "Total cases", the back-calculated total is
        # folded entirely into Confirmed and Suspected is set to 0 (to avoid
        # double-counting). That is intended, but easy to miss — make it loud.
        if _cfr_on and _cfr_role_fc == "Total cases" and not _susp_est_on:
            st.warning(
                "**Suspected = 0 in this forecast.** The CFR role is "
                "**\"Total cases\"** (the default), so the back-calculated total "
                "is counted entirely as Confirmed and Suspected is dropped to "
                "avoid double-counting. To project Suspected separately, set the "
                "CFR role to **Confirmed** in Step 1, or enable **Estimate "
                "Suspected** there."
            )

        # --- Step 2 ↔ Step 3 consistency check (problem #5) ---
        # R_t was estimated on ONE stream in Step 2 (rt_incidence_source), and
        # the forecast applies that R_t to the seeds chosen here. If, for the
        # SAME case type, Step 2 used Estimated but the forecast seeds Observed
        # (or vice versa), the growth rate and the projected series describe
        # different things. We do NOT silently remap (that would fight the
        # Step 1 CFR role — see Fix #7); instead we surface it so you can align
        # Step 1's CFR role or Step 2's source deliberately.
        _rt_src = st.session_state.get("rt_incidence_source", "confirmed")
        _rt_case = "Suspected" if _rt_src.startswith("suspected") else "Confirmed"
        _rt_estimated = _rt_src in ("cfr", "cfr_smooth",
                                    "suspected_est", "suspected_est_smooth")
        _conf_seed_estimated = bool(
            _cfr_on and _cfr_role_fc in ("Confirmed", "Total cases"))
        _susp_seed_estimated = bool(
            (_cfr_on and _cfr_role_fc == "Suspected") or _susp_est_on)
        _seed_estimated = (_conf_seed_estimated if _rt_case == "Confirmed"
                           else _susp_seed_estimated)
        _seed_lbl = (_conf_src_label if _rt_case == "Confirmed"
                     else _susp_src_label)
        if _rt_estimated != _seed_estimated:
            st.warning(
                f"**R_t / forecast mismatch.** R_t was estimated on "
                f"**{'Estimated' if _rt_estimated else 'Observed'} {_rt_case}** "
                f"in Step 2, but the forecast seeds {_rt_case} from "
                f"**{_seed_lbl}**. For a coherent forecast these should describe "
                f"the same series — adjust the CFR role in Step 1 or the R_t "
                f"source in Step 2 so they match."
            )
        b1, b2, b3 = st.columns(3)
        with b1:
            obs_conf = st.number_input(
                "Confirmed", min_value=0,
                value=st.session_state.get("obs_conf_input",
                                            base_conf_default),
                step=1, key="obs_conf_input",
            )
        with b2:
            obs_susp = st.number_input(
                "Suspected", min_value=0,
                value=st.session_state.get("obs_susp_input",
                                            base_susp_default),
                step=1, key="obs_susp_input",
            )
        with b3:
            obs_death = st.number_input(
                "Deaths", min_value=0,
                value=st.session_state.get("obs_death_input",
                                            base_death_default),
                step=1, key="obs_death_input",
            )

        # --- Disease parameters ---
        st.markdown('<div class="section-label">Disease parameters</div>',
                    unsafe_allow_html=True)

        def disease_param(label, default_val, key, min_val, max_val, step,
                          is_int=False):
            st.markdown(
                f'<div style="font-size:0.85rem; font-weight:500; '
                f'margin:0.45rem 0 0.15rem 0; color:#333;">{label}</div>',
                unsafe_allow_html=True,
            )
            c1, c2 = st.columns([1, 1.6])
            with c1:
                if is_int:
                    val = st.number_input("v", min_value=int(min_val),
                                           max_value=int(max_val),
                                           value=int(default_val),
                                           step=int(step), key=f"{key}_val",
                                           label_visibility="collapsed")
                else:
                    val = st.number_input("v", min_value=float(min_val),
                                           max_value=float(max_val),
                                           value=float(default_val),
                                           step=float(step), key=f"{key}_val",
                                           label_visibility="collapsed")
            with c2:
                src = st.text_input("s", value=DEFAULT_FC_SRC[key],
                                     key=f"{key}_src",
                                     label_visibility="collapsed")
            if src and src.strip():
                lbl = (DEFAULT_FC_LABEL.get(key, "")
                       if src == DEFAULT_FC_SRC[key] else src)
                st.markdown(
                    f'<div style="margin:-0.35rem 0 0.25rem 0; font-size:0.78rem;">'
                    f'<a href="{src}" target="_blank" rel="noopener" '
                    f'style="color:#1f4e79; text-decoration:none; '
                    f'border-bottom:1px dotted #1f4e79;">{lbl}</a></div>',
                    unsafe_allow_html=True,
                )
            return val

        # CFR is a single quantity. Step 1's back-calculation (cases =
        # deaths / CFR) and this death projection (deaths = cases x CFR) are
        # inverse uses of the SAME CFR. So when Step 1's CFR back-calculation
        # was used, inherit that value here (locked) instead of asking a second
        # time — two different CFRs would be self-contradictory. cfr_pct_active
        # is a percent; the forecast uses a fraction.
        if (st.session_state.get("cfr_active")
                and "cfr_pct_active" in st.session_state):
            cfr = float(st.session_state["cfr_pct_active"]) / 100.0
            st.markdown(
                '<div style="font-size:0.85rem; font-weight:500; '
                'margin:0.45rem 0 0.3rem 0; color:#333;">'
                'CFR (case fatality ratio)</div>',
                unsafe_allow_html=True)
            st.markdown(
                frozen_card("CFR — inherited from Step 1", f"{cfr * 100:.1f}%",
                            "Same CFR as the Step 1 back-calculation. "
                            "Change it in Step 1."),
                unsafe_allow_html=True)
        else:
            cfr = disease_param("CFR (case fatality ratio)", 0.34, "cfr",
                                0.01, 1.0, 0.01)
        # Stable copy of the CFR the forecast actually uses, for the export
        # sheet (the editable widget key is dropped once you leave Step 3).
        st.session_state["fc_cfr"] = cfr
        death_lag = disease_param("Onset-to-death lag (days)", 10, "lag",
                                   0, 60, 1, is_int=True)

        # --- R_t scenarios ---
        st.markdown(
            '<div class="section-label">R_t trajectories (3 scenarios)</div>',
            unsafe_allow_html=True,
        )
        st.caption(
            f"All scenarios start from R_t = {rt_starting:.2f} (your selected "
            f"value from Step 2) and move linearly to the target over the "
            f"given days, then hold."
        )

        scen_defaults = {
            "S1": {"label": "Delayed response", "target": 0.9, "days": 180},
            "S2": {"label": "Moderate response", "target": 0.9, "days": 90},
            "S3": {"label": "Strong response", "target": 0.6, "days": 30},
        }
        scen_inputs = {}
        for name, d in scen_defaults.items():
            st.markdown(
                f'<div style="font-size:0.82rem; font-weight:500; '
                f'margin:0.5rem 0 0.15rem 0;">'
                f'<span style="color:{SCENARIO_COLOURS[name]};">●</span> '
                f'{name} — {d["label"]}</div>',
                unsafe_allow_html=True,
            )
            sc1, sc2 = st.columns(2)
            with sc1:
                tgt = st.number_input(f"Target R_t ({name})",
                                       min_value=0.01, max_value=20.0,
                                       value=float(d["target"]),
                                       step=0.05, key=f"{name}_target",
                                       label_visibility="visible")
            with sc2:
                dys = st.number_input(f"Days to target ({name})",
                                       min_value=1, max_value=720,
                                       value=int(d["days"]), step=1,
                                       key=f"{name}_days",
                                       label_visibility="visible")
            scen_inputs[name] = {"target": float(tgt), "days": int(dys)}

        run_fc = st.button("Run forecast", type="primary",
                           use_container_width=True)

        # Stash preview data for the right panel
        try:
            horizon_int = int(horizon)
            st.session_state["fc_preview_traj"] = {
                name: build_rt_trajectory(rt_starting, v["target"],
                                          v["days"], horizon_int)
                for name, v in scen_inputs.items()
            }
            st.session_state["fc_preview_dates"] = pd.date_range(
                pd.Timestamp(start_date) + pd.Timedelta(days=1),
                periods=horizon_int, freq="D",
            )
        except Exception:
            pass
        if run_fc:
            horizon_int = int(horizon)
            seed_conf = series_in["new_confirmed"].astype(float).values
            seed_susp = series_in["new_suspected"].astype(float).values

            # Pick the Confirmed and Suspected seeds INDEPENDENTLY, using the
            # SAME precedence as the baseline block above (they must agree):
            #   Confirmed: CFR estimate if CFR routes there (Confirmed/Total),
            #              else observed confirmed.
            #   Suspected: CFR estimate if CFR routes to Suspected; else the
            #              ratio-based Estimated Suspected (Step 1) if active;
            #              else 0 when CFR took everything as Total; else observed.
            _cfr_on = (st.session_state.get("cfr_active")
                       and "new_cfr_estimated" in series_in.columns)
            _susp_est_on = (st.session_state.get("suspest_active")
                            and "new_suspected_estimated" in series_in.columns)
            _role = st.session_state.get("cfr_role_active", "Total cases")
            _cfr_est = (series_in["new_cfr_estimated"].astype(float).values
                        if _cfr_on else None)

            if _cfr_on and _role in ("Confirmed", "Total cases"):
                seed_conf = _cfr_est

            if _cfr_on and _role == "Suspected":
                seed_susp = _cfr_est
            elif _susp_est_on:
                seed_susp = series_in["new_suspected_estimated"].astype(float).values
            elif _cfr_on and _role == "Total cases":
                # CFR total folded into Confirmed → no separate suspected.
                seed_susp = np.zeros_like(seed_conf)
            # else: observed suspected (CFR=Confirmed or no CFR) — already set.

            # --- Sample R_t starting values from the Cori Gamma posterior ---
            # The user picked a preset (Latest / Mean 7d / etc.) in Step 2,
            # and we have selected_rt + the posterior shape/rate of the latest
            # window. We anchor the *mean* of the sampling distribution at
            # selected_rt by rescaling Gamma(shape_post, 1/rate_post).
            rt_df_local = st.session_state.get("rt_df")
            shape_p, rate_p = None, None
            if rt_df_local is not None:
                _primary = rt_df_local[
                    rt_df_local["si_mean_used"] == si_mean_used
                ].dropna(subset=["rt_mean"])
                if not _primary.empty:
                    _last = _primary.iloc[-1]
                    shape_p = float(_last.get("shape_post", np.nan))
                    rate_p = float(_last.get("rate_post", np.nan))

            N_SAMPLES = 200
            rng = np.random.default_rng(42)
            if (shape_p is not None and np.isfinite(shape_p)
                    and rate_p is not None and rate_p > 0):
                raw = rng.gamma(shape=shape_p, scale=1.0 / rate_p,
                                size=N_SAMPLES)
                posterior_mean = shape_p / rate_p
                # Rescale so the *sample mean* matches the user's selected R_t
                # (preserves posterior dispersion, anchors the central value).
                rt_samples = raw * (rt_starting / posterior_mean)
            else:
                # Fallback: tight CV=10% around the chosen point R_t.
                rt_samples = rng.normal(rt_starting, max(0.05, rt_starting * 0.10),
                                         size=N_SAMPLES)
            rt_samples = np.clip(rt_samples, 0.05, 20.0)

            scenarios_out = {}
            with st.spinner("Sampling R_t posterior & projecting…"):
                for name, v in scen_inputs.items():
                    scenarios_out[name] = run_scenario_uncertain(
                        seed_conf, seed_susp, rt_samples,
                        v["target"], v["days"], horizon_int,
                        obs_conf, obs_susp, obs_death,
                        cfr, int(death_lag), si_mean_used, si_sd_used,
                    )
            horizon_dates = pd.date_range(
                pd.Timestamp(start_date) + pd.Timedelta(days=1),
                periods=horizon_int, freq="D",
            )
            st.session_state["fc_scenarios"] = scenarios_out
            st.session_state["fc_dates"] = horizon_dates
            st.session_state["fc_baselines"] = {
                "confirmed": obs_conf,
                "suspected": obs_susp,
                "deaths": obs_death,
            }
            st.session_state["fc_scen_inputs"] = scen_inputs
            st.session_state["fc_rt_start"] = rt_starting
            st.session_state["fc_n_samples"] = N_SAMPLES

    with right3:
        st.markdown('<div class="freeze-right"></div>', unsafe_allow_html=True)
        top_row = st.columns([1, 0.18])
        with top_row[0]:
            st.markdown('<div class="panel-title">Forecast output</div>',
                        unsafe_allow_html=True)
        with top_row[1]:
            if st.button("← Back", use_container_width=True, key="back_to_rt"):
                st.session_state["step"] = "rt"
                st.rerun()

        # R_t trajectory preview always visible (updates as inputs change)
        preview_traj = st.session_state.get("fc_preview_traj")
        preview_dates = st.session_state.get("fc_preview_dates")
        # R_t trajectory preview chart removed per UI request — the scenarios
        # are visible in the legend and chart annotations of the forecast
        # chart itself.

        scenarios_out = st.session_state.get("fc_scenarios")
        horizon_dates = st.session_state.get("fc_dates")
        baselines = st.session_state.get("fc_baselines")

        if not scenarios_out:
            st.markdown(
                '<div class="placeholder-card">'
                'Adjust the editable inputs on the left and click '
                '<b>Run forecast</b>. You will see a 3-panel cumulative '
                'projection (confirmed / suspected / deaths) for the three '
                'response scenarios, plus a downloadable forecast table.'
                '</div>',
                unsafe_allow_html=True,
            )
        else:
            scen_inputs_used = st.session_state.get("fc_scen_inputs", {})
            scen_labels = {
                name: (
                    f"<b>{scen_defaults[name]['label']}</b><br>"
                    f"<span style='font-size:10px;color:#5b6573;'>"
                    f"R<sub>t</sub> → "
                    f"{scen_inputs_used.get(name, {}).get('target', 0):.1f} "
                    f"over {scen_inputs_used.get(name, {}).get('days', 0)} d"
                    f"</span>"
                )
                for name in scenarios_out
            }
            tog_l, tog_m, tog_r, tog_r2 = st.columns([1.7, 1, 1, 1])
            with tog_m:
                mode_view = st.radio(
                    "View", ["Daily", "Cumulative"], horizontal=True,
                    key="fc_view_mode", label_visibility="collapsed",
                )
            with tog_r:
                y_log = st.toggle("Log y-axis", value=True,
                                   help="Toggle between log and linear scale.",
                                   key="fc_y_log")
            with tog_r2:
                independent_y = st.toggle(
                    "Independent Y", value=False,
                    help="Off: all three panels share one y-axis scale "
                         "(directly comparable). On: each scenario panel gets "
                         "its own y-axis scale.",
                    key="fc_independent_y")
            with tog_l:
                st.caption(
                    f"Forecast for **{scenario_name}** · "
                    f"shaded band = 90% posterior predictive interval "
                    f"from {st.session_state.get('fc_n_samples', '?')} R_t draws."
                )
            _fig_forecast = forecast_chart(
                scenarios_out, horizon_dates, baselines, scen_labels,
                y_log=bool(y_log),
                mode=("daily" if mode_view == "Daily" else "cumulative"),
                independent_y=bool(independent_y),
            )
            st.session_state["chart_forecast"] = _fig_forecast
            chart_with_line_filter(_fig_forecast, key="forecast")

            # --- Summary cards: 180-day totals per scenario ---
            st.markdown(
                '<div style="font-size:0.78rem; font-weight:600; color:#5b6573; '
                'text-transform:uppercase; letter-spacing:0.05em; '
                'margin:0.7rem 0 0.35rem 0;">'
                f'End-of-horizon ({len(horizon_dates)}-day) totals'
                '</div>',
                unsafe_allow_html=True,
            )
            def _median(metric):
                return metric["median"] if isinstance(metric, dict) else metric

            def _lo(metric):
                return (metric["lower"] if isinstance(metric, dict)
                        else metric)

            def _hi(metric):
                return (metric["upper"] if isinstance(metric, dict)
                        else metric)

            for name, data in scenarios_out.items():
                cum_conf_med = _median(data["cum_confirmed"])
                cum_susp_med = _median(data["cum_suspected"])
                cum_death_med = _median(data["cum_deaths"])
                new_conf_med = _median(data["new_confirmed"])
                c_end = float(cum_conf_med[-1])
                c_lo = float(_lo(data["cum_confirmed"])[-1])
                c_hi = float(_hi(data["cum_confirmed"])[-1])
                s_end = float(cum_susp_med[-1])
                d_end = float(cum_death_med[-1])
                d_lo = float(_lo(data["cum_deaths"])[-1])
                d_hi = float(_hi(data["cum_deaths"])[-1])
                peak_day = horizon_dates[int(np.argmax(new_conf_med))]
                peak_val = float(np.max(new_conf_med))

                # Projected last case (median trajectory)
                above_thresh = np.where(new_conf_med >= 0.5)[0]
                if len(above_thresh) > 0 and above_thresh[-1] < len(horizon_dates) - 1:
                    last_case_date = horizon_dates[above_thresh[-1]]
                    who_eoo_date = last_case_date + pd.Timedelta(days=42)
                    eoo_html = (
                        f'<br><span style="color:#5b6573;">Projected last case: '
                        f'<b>{last_case_date.strftime("%d %b %Y")}</b> · '
                        f'WHO 42-day EOO declaration: '
                        f'<b style="color:#1f7a3a;">'
                        f'{who_eoo_date.strftime("%d %b %Y")}</b></span>'
                    )
                else:
                    eoo_html = (
                        '<br><span style="color:#B22222;">Outbreak still active at '
                        f'horizon end ({horizon_dates[-1].strftime("%d %b %Y")}) '
                        '— extend horizon to find an end date.</span>'
                    )

                st.markdown(
                    f'<div style="border-left:4px solid {SCENARIO_COLOURS[name]}; '
                    f'background:#fafbfc; padding:0.5rem 0.8rem; '
                    f'margin-bottom:0.4rem; font-size:0.85rem;">'
                    f'<b style="color:{SCENARIO_COLOURS[name]};">{name}</b> '
                    f'— {scen_defaults[name]["label"]}<br>'
                    f'Confirmed: <b>{c_end:,.0f}</b> '
                    f'<span style="color:#5b6573;">[{c_lo:,.0f}–{c_hi:,.0f}]</span>  '
                    f'· Suspected: <b>{s_end:,.0f}</b>  '
                    f'· Deaths: <b>{d_end:,.0f}</b> '
                    f'<span style="color:#5b6573;">[{d_lo:,.0f}–{d_hi:,.0f}]</span><br>'
                    f'<span style="color:#5b6573;">Peak new confirmed (median): '
                    f'{peak_val:.1f} on {peak_day.strftime("%d %b %Y")} · '
                    f'90% PI from {st.session_state.get("fc_n_samples", "?")} '
                    f'posterior draws</span>'
                    f'{eoo_html}'
                    f'</div>',
                    unsafe_allow_html=True,
                )

            # --- Table + download (collapsed) ---
            with st.expander("Show table / data", expanded=False):
                rows = []
                for name, data in scenarios_out.items():
                    for i, d in enumerate(horizon_dates):
                        def _get(metric, key, idx):
                            v = data[metric]
                            return float(v[key][idx]) if isinstance(v, dict) else float(v[idx])
                        rows.append({
                            "date": d.strftime("%Y-%m-%d"),
                            "scenario": name,
                            "new_confirmed_median":   _get("new_confirmed", "median", i),
                            "new_confirmed_lower":    _get("new_confirmed", "lower", i),
                            "new_confirmed_upper":    _get("new_confirmed", "upper", i),
                            "new_suspected_median":   _get("new_suspected", "median", i),
                            "new_deaths_median":      _get("new_deaths", "median", i),
                            "cumulative_confirmed_median": _get("cum_confirmed", "median", i),
                            "cumulative_confirmed_lower":  _get("cum_confirmed", "lower", i),
                            "cumulative_confirmed_upper":  _get("cum_confirmed", "upper", i),
                            "cumulative_suspected_median": _get("cum_suspected", "median", i),
                            "cumulative_deaths_median":    _get("cum_deaths", "median", i),
                            "cumulative_deaths_lower":     _get("cum_deaths", "lower", i),
                            "cumulative_deaths_upper":     _get("cum_deaths", "upper", i),
                        })
                fc_df = pd.DataFrame(rows)
                preview = fc_df.copy()
                preview["date"] = pd.to_datetime(preview["date"]).dt.strftime(
                    "%d %b %Y")
                for c in [col for col in preview.columns
                          if col not in ("date", "scenario")]:
                    preview[c] = preview[c].round(1)
                st.dataframe(preview, use_container_width=True, height=320,
                             hide_index=True)
                slug = _slug(st.session_state.get("scenario_name", ""))
                st.download_button(
                    "Download who_forecast.csv",
                    fc_df.to_csv(index=False).encode("utf-8"),
                    f"{slug}__who_forecast.csv", "text/csv",
                    use_container_width=True,
                )

            # Next: End-of-outbreak (now Step 4)
            if st.button("Next: Step 4 — End-of-outbreak  →",
                          type="primary", use_container_width=True,
                          key="goto_eoo"):
                st.session_state["step"] = "eoo"
                st.rerun()

    # Method note + equations
    with st.expander("📐 Method & equations — Renewal-equation forecast "
                       "(Nouvellet 2018)", expanded=False):
        st.markdown("**Step 1 — Renewal equation (projection one day forward):**")
        st.latex(r"I_t \;=\; R_t \cdot \sum_{s=1}^{t} w_s \cdot I_{t-s}")
        st.caption(
            "Today's new cases = R_t times a weighted sum of recent cases. "
            "Weights w_s are the same Gamma SI weights as Step 2 (recent days "
            "matter more)."
        )

        st.markdown("**Step 2 — Linear R_t trajectory per scenario:**")
        st.latex(
            r"R_t = "
            r"\begin{cases}"
            r"R_0 + (R_{\rm target} - R_0)\cdot \dfrac{t}{D},"
            r"& t \le D \\[4pt]"
            r"R_{\rm target},"
            r"& t > D"
            r"\end{cases}"
        )
        st.caption(
            "Each scenario (S1/S2/S3) starts from the R_t selected in Step 2 "
            "and moves linearly to its target over D days, then holds."
        )

        st.markdown("**Step 3 — Deaths from confirmed cases with lag:**")
        st.latex(r"D_t \;=\; \mathrm{CFR} \cdot I_{t - \ell}^{\,\mathrm{confirmed}}")
        st.caption(
            "Deaths at day t are CFR × confirmed cases ℓ days ago (default ℓ = 10, "
            "Wamala 2010)."
        )

        st.markdown("**Step 4 — Cumulative trajectories:**")
        st.latex(r"C_t \;=\; C_{\rm obs} + \sum_{k=0}^{t} I_k")
        st.caption(
            "Anchored at the observed baseline at the start date. Confirmed and "
            "suspected are projected independently from their own seed series. "
            "Shaded band on chart = 90% posterior predictive interval from N R_t "
            "samples drawn from the Cori Gamma posterior of Step 2."
        )

    st.stop()


# -------------------------------------------------------------------------
# STEP 2 — R_t estimation page (rendered if step == "rt", then st.stop())
# -------------------------------------------------------------------------
if st.session_state["step"] == "rt":
    series_in = st.session_state["result_series"]
    left2, right2 = st.columns([1, 1.25], gap="large")

    DEFAULT_SRC = {
        "si_mean_primary": "https://doi.org/10.1056/NEJMoa1411100",
        "si_sd_primary":   "https://doi.org/10.1056/NEJMoa1411100",
        "si_mean_sens":    "https://doi.org/10.1017/S0950268806007217",
        "si_sd_sens":      "https://doi.org/10.1017/S0950268806007217",
        "window":          "https://doi.org/10.1093/aje/kwt133",
        "shape_prior":     "https://doi.org/10.1093/aje/kwt133",
        "rate_prior":      "https://doi.org/10.1093/aje/kwt133",
    }
    DEFAULT_SRC_LABEL = {
        "si_mean_primary": "WHO Ebola Response Team 2014, NEJM",
        "si_sd_primary":   "WHO Ebola Response Team 2014, NEJM",
        "si_mean_sens":    "Legrand et al. 2007, Epidemiol Infect",
        "si_sd_sens":      "Legrand et al. 2007, Epidemiol Infect",
        "window":          "Cori et al. 2013, Am J Epidemiol",
        "shape_prior":     "Cori et al. 2013, Am J Epidemiol",
        "rate_prior":      "Cori et al. 2013, Am J Epidemiol",
    }

    def param_with_source(label: str, default_val: float, key: str,
                          min_val: float, max_val: float, step: float,
                          is_int: bool = False, disabled: bool = False):
        st.markdown(
            f'<div style="font-size:0.85rem; font-weight:500; '
            f'margin:0.55rem 0 0.15rem 0; color:#333;">{label}</div>',
            unsafe_allow_html=True,
        )
        c1, c2 = st.columns([1, 1.6])
        with c1:
            if is_int:
                val = st.number_input("v", min_value=int(min_val),
                                       max_value=int(max_val),
                                       value=int(default_val), step=int(step),
                                       key=f"{key}_val", disabled=disabled,
                                       label_visibility="collapsed")
            else:
                val = st.number_input("v", min_value=float(min_val),
                                       max_value=float(max_val),
                                       value=float(default_val),
                                       step=float(step), key=f"{key}_val",
                                       disabled=disabled,
                                       label_visibility="collapsed")
        with c2:
            src = st.text_input("s", value=DEFAULT_SRC[key],
                                 placeholder="Source URL or DOI",
                                 key=f"{key}_src", disabled=disabled,
                                 label_visibility="collapsed")
        if src and src.strip():
            label_text = DEFAULT_SRC_LABEL.get(key, "") if src == DEFAULT_SRC[key] else src
            st.markdown(
                f'<div style="margin:-0.35rem 0 0.25rem 0; font-size:0.78rem;">'
                f'<a href="{src}" target="_blank" rel="noopener" '
                f'style="color:#1f4e79; text-decoration:none; '
                f'border-bottom:1px dotted #1f4e79;">{label_text}</a></div>',
                unsafe_allow_html=True,
            )
        return val

    with left2:
        st.markdown('<div class="panel-title">R_t inputs</div>',
                    unsafe_allow_html=True)

        st.markdown('<div class="section-label">Incidence series for R_t</div>',
                    unsafe_allow_html=True)
        # What Step 1 produced — drives which choices are offered.
        has_cfr = (st.session_state.get("cfr_active", False)
                   and "new_cfr_estimated" in series_in.columns)
        has_susp_est = "new_suspected_estimated" in series_in.columns
        has_smooth = "new_confirmed_raw" in series_in.columns

        # Choice 1 — Case type.
        case_type = st.radio(
            "Case type", ["Confirmed", "Suspected"], horizontal=True,
            key="rt_case_type",
            help="Which case stream drives the R_t estimate.")

        # Choice 2 — Actual vs Estimated (Estimated only offered when Step 1
        # produced an estimate for this case type).
        est_available = has_cfr if case_type == "Confirmed" else has_susp_est
        source_opts = ["Actual", "Estimated"] if est_available else ["Actual"]
        if st.session_state.get("rt_source_kind") not in source_opts:
            st.session_state.pop("rt_source_kind", None)
        source_kind = st.radio(
            "Source", source_opts, horizontal=True, key="rt_source_kind",
            help=("Actual = observed counts. Estimated = "
                  + ("infections back-calculated from deaths (CFR, Step 1)."
                     if case_type == "Confirmed"
                     else "derived from confirmed via the confirmed-to-"
                          "suspected ratio (Step 1).")))

        # Choice 3 — Smoothed? (only when Step 1 smoothing was applied).
        smoothed = False
        if has_smooth:
            smoothed = st.checkbox(
                "Use smoothed values",
                value=st.session_state.get("rt_use_smoothed", False),
                key="rt_use_smoothed",
                help="Trailing moving average using the Step 1 window.")

        # Map the three choices → internal source key.
        base_key = {
            ("Confirmed", "Actual"):    "confirmed",
            ("Suspected", "Actual"):    "suspected",
            ("Confirmed", "Estimated"): "cfr",
            ("Suspected", "Estimated"): "suspected_est",
        }[(case_type, source_kind)]
        incidence_source = base_key + ("_smooth" if smoothed else "")
        # Fall back to the unsmoothed series if the smoothed column for this
        # combination was not produced in Step 1.
        if (smoothed and _resolve_rt_col(series_in, incidence_source)
                not in series_in.columns):
            incidence_source = base_key
        st.session_state["rt_incidence_source"] = incidence_source

        st.markdown('<div class="section-label">Serial interval (primary)</div>',
                    unsafe_allow_html=True)
        # Defaults seed from the last-estimated SI (rt_si_mean / rt_si_sd) so
        # returning to Step 2 re-shows the values you used, not the literals.
        si_mean = param_with_source(
            "SI mean (days)",
            float(st.session_state.get("rt_si_mean", 15.3)),
            "si_mean_primary", 1.0, 60.0, 0.1)
        si_sd = param_with_source(
            "SI SD (days)",
            float(st.session_state.get("rt_si_sd", 9.3)),
            "si_sd_primary", 0.1, 30.0, 0.1)

        st.markdown('<div class="section-label">Sensitivity SI</div>',
                    unsafe_allow_html=True)
        run_sens = st.checkbox("Overlay a sensitivity-analysis SI",
                                value=st.session_state.get("run_sens", True),
                                key="run_sens")
        sens_mean = param_with_source("Sensitivity SI mean (days)", 12.0,
                                       "si_mean_sens", 1.0, 60.0, 0.1,
                                       disabled=not run_sens)
        sens_sd = param_with_source("Sensitivity SI SD (days)", 5.0,
                                     "si_sd_sens", 0.1, 30.0, 0.1,
                                     disabled=not run_sens)

        st.markdown('<div class="section-label">Sliding window</div>',
                    unsafe_allow_html=True)
        window = param_with_source("Window (days)", 7, "window",
                                    2, 21, 1, is_int=True)

        st.markdown('<div class="section-label">Bayesian priors</div>',
                    unsafe_allow_html=True)
        st.caption(
            "EpiEstim default: shape = 1, rate = 0.2 → prior R_t ~ Exp(mean = 5). "
            "This is weakly informative (does not pull R_t down). "
            "Smaller `rate` widens the prior; larger `rate` (e.g. 5) is **strongly** "
            "suppressive and biases R_t toward zero — only use for stable, late-phase outbreaks."
        )
        shape_prior = param_with_source("Prior shape", 1.0, "shape_prior",
                                         0.01, 100.0, 0.1)
        rate_prior = param_with_source("Prior rate", 0.2, "rate_prior",
                                        0.01, 100.0, 0.05)

        run_rt = st.button("Estimate R_t", type="primary",
                           use_container_width=True)

        if run_rt:
            rt_df = compute_rt_table(series_in, si_mean, si_sd, int(window),
                                     shape_prior, rate_prior,
                                     run_sens, sens_mean, sens_sd,
                                     incidence_source=incidence_source)
            if rt_df.empty:
                st.warning(
                    f"Not enough data — need at least {int(window)} days of incidence."
                )
            else:
                st.session_state["rt_df"] = rt_df
                # Persist the serial interval used for this R_t run in dedicated
                # (non-widget) keys. Widget keys like si_mean_primary_val are
                # garbage-collected once Step 2 is navigated away from, so
                # Step 3/4 must read these stable copies instead — otherwise
                # they silently fall back to the default SI.
                st.session_state["rt_si_mean"] = si_mean
                st.session_state["rt_si_sd"] = si_sd

    with right2:
        st.markdown('<div class="freeze-right"></div>', unsafe_allow_html=True)
        top_row = st.columns([1, 0.18])
        with top_row[0]:
            st.markdown('<div class="panel-title">R_t output</div>',
                        unsafe_allow_html=True)
        with top_row[1]:
            if st.button("← Back", use_container_width=True, key="back_btn"):
                st.session_state["step"] = "data"
                st.rerun()

        rt_df = st.session_state.get("rt_df")
        rt_si = st.session_state.get("rt_si_mean")
        if rt_df is None or rt_df.empty:
            st.markdown(
                '<div class="placeholder-card">'
                'Set the serial interval and priors on the left, then click '
                '<b>Estimate R_t</b>. The epidemic curve will appear on top and '
                'the R_t trajectory below, with 95% credible interval band.'
                '</div>',
                unsafe_allow_html=True,
            )
        else:
            _fig_rt = rt_combined_chart(series_in, rt_df, rt_si)
            st.session_state["chart_rt"] = _fig_rt
            chart_with_line_filter(_fig_rt, key="rt")

            primary = rt_df[rt_df["si_mean_used"] == rt_si].dropna(
                subset=["rt_mean"]).copy()
            if not primary.empty:
                latest = primary.iloc[-1]
                direction = "above" if latest["rt_mean"] > 1 else "below"
                color = "#B22222" if latest["rt_mean"] > 1 else "#1f7a3a"
                phase = "growing" if direction == "above" else "declining"
                st.markdown(
                    f'<div style="padding:0.7rem 1rem; border-left:4px solid {color}; '
                    f'background:#fafbfc; margin:0.5rem 0;">'
                    f'<b>Latest R_t</b>: {latest["rt_mean"]:.2f} '
                    f'(95% CrI {latest["rt_lower"]:.2f}–{latest["rt_upper"]:.2f}) '
                    f'on {pd.to_datetime(latest["date"]).strftime("%d %b %Y")} — '
                    f'<b style="color:{color}">{direction} 1</b>, outbreak is '
                    f'<b>{phase}</b>.'
                    f'</div>',
                    unsafe_allow_html=True,
                )

            # ----- R_t selector for forecast (compact, right below the chart) -----
            st.markdown(
                '<div style="font-size:0.78rem; font-weight:600; color:#5b6573; '
                'text-transform:uppercase; letter-spacing:0.05em; '
                'margin:0.6rem 0 0.35rem 0;">R_t for forecast</div>',
                unsafe_allow_html=True,
            )
            valid = primary["rt_mean"].dropna()
            if len(valid) > 0:
                latest_val = float(valid.iloc[-1])
                last7_val = float(valid.tail(7).mean())
                median_val = float(valid.median())
                overall_val = float(valid.mean())

                def _mini_card(label, value, recommended=False):
                    border = "#1f4e79" if recommended else "#e3e7ed"
                    bg = "#f1f5fa" if recommended else "#fafbfc"
                    return (
                        f'<div style="border:1px solid {border}; background:{bg}; '
                        f'border-radius:6px; padding:0.35rem 0.5rem; text-align:center;">'
                        f'<div style="font-size:0.66rem; color:#5b6573; '
                        f'text-transform:uppercase; letter-spacing:0.04em;">'
                        f'{label}</div>'
                        f'<div style="font-size:1.05rem; font-weight:600; '
                        f'color:#1f4e79; line-height:1.3;">{value:.2f}</div>'
                        f'</div>'
                    )

                m1, m2, m3, m4 = st.columns(4)
                m1.markdown(_mini_card("Latest", latest_val, recommended=True),
                            unsafe_allow_html=True)
                m2.markdown(_mini_card("Mean (last 7 d)", last7_val),
                            unsafe_allow_html=True)
                m3.markdown(_mini_card("Median (all)", median_val),
                            unsafe_allow_html=True)
                m4.markdown(_mini_card("Mean (all)", overall_val),
                            unsafe_allow_html=True)

                preset = st.radio(
                    "Use which R_t for the forecast?",
                    ["Latest (recommended)", "Mean (last 7 d)",
                     "Median (all)", "Mean (all)", "Custom"],
                    horizontal=True, key="rt_preset",
                    label_visibility="collapsed",
                )
                preset_map = {
                    "Latest (recommended)": latest_val,
                    "Mean (last 7 d)": last7_val,
                    "Median (all)": median_val,
                    "Mean (all)": overall_val,
                }
                base = preset_map.get(preset, latest_val)
                selected_rt = st.number_input(
                    "R_t value to use", min_value=0.01, max_value=20.0,
                    value=round(base, 2), step=0.05,
                    disabled=preset != "Custom",
                    help="Pre-filled from the choice above. Switch to Custom to override.",
                )
                st.session_state["selected_rt"] = selected_rt
                st.session_state["selected_rt_basis"] = preset

                # Next: Forecast
                if st.button("Next: Step 3 — Forecast  →", type="primary",
                              use_container_width=True,
                              key="goto_forecast"):
                    st.session_state["step"] = "forecast"
                    st.rerun()

            # ----- Show table & downloads (collapsed by default) -----
            with st.expander("Show table / data", expanded=False):
                display = primary[["date", "rt_mean", "rt_lower", "rt_upper"]].copy()
                display["date"] = pd.to_datetime(display["date"]).dt.strftime("%d %b %Y")
                for c in ["rt_mean", "rt_lower", "rt_upper"]:
                    display[c] = display[c].round(2)
                display = display.rename(columns={
                    "date": "Date", "rt_mean": "R_t mean",
                    "rt_lower": "R_t lower (2.5%)", "rt_upper": "R_t upper (97.5%)",
                })
                st.dataframe(display, use_container_width=True, height=280,
                             hide_index=True)

                csv_bytes = rt_df.copy()
                csv_bytes["date"] = pd.to_datetime(
                    csv_bytes["date"]).dt.strftime("%Y-%m-%d")
                slug = _slug(st.session_state.get("scenario_name", ""))
                st.download_button(
                    "Download rt_estimates.csv",
                    csv_bytes.to_csv(index=False).encode("utf-8"),
                    f"{slug}__rt_estimates.csv", "text/csv",
                    use_container_width=True,
                )

    # Method note + equations
    with st.expander("📐 Method & equations — Cori 2013 R_t estimator",
                      expanded=False):
        st.markdown(
            "Instantaneous reproduction number using the **Cori et al. (2013)** "
            "sliding-window Bayesian method. For each window of length τ ending at time _t_:"
        )
        st.markdown("**Step 1 — Discretise the serial interval** (Gamma):")
        st.latex(r"w_s = F_\Gamma(s + 0.5) - F_\Gamma(s - 0.5),\quad s = 1, 2, \dots")
        st.caption("How much weight each past day carries for explaining today's cases.")

        st.markdown("**Step 2 — SI-weighted past incidence** at every day _s_:")
        st.latex(r"\Lambda_s \;=\; \sum_{k=1}^{s} w_k \cdot I_{s-k}")
        st.caption("How many secondary infections we would have expected today if R_t = 1.")

        st.markdown("**Step 3 — Window sums** over the last τ days:")
        st.latex(r"I_{[t-\tau+1,\,t]} = \sum_{s=t-\tau+1}^{t} I_s "
                  r"\qquad "
                  r"\Lambda_{[t-\tau+1,\,t]} = \sum_{s=t-\tau+1}^{t} \Lambda_s")
        st.caption("Both incidence and Λ are summed over the SAME window — "
                    "matches Cori 2013 eq. (4) and the EpiEstim R package.")

        st.markdown("**Step 4 — Conjugate Gamma posterior:**")
        st.latex(r"R_t \mid \mathrm{data} \;\sim\; "
                  r"\mathrm{Gamma}\!\left( "
                  r"\alpha_0 + I_{[t-\tau+1,t]}, \; "
                  r"\beta_0 + \Lambda_{[t-\tau+1,t]}"
                  r"\right)")
        st.latex(r"\widehat{R}_t = "
                  r"\frac{\alpha_0 + I_{[t-\tau+1,t]}}"
                  r"{\beta_0 + \Lambda_{[t-\tau+1,t]}}")
        st.caption("95% CrI from the Gamma 2.5% and 97.5% quantiles.")

        st.markdown(
            "**Defaults:** SI Gamma(15.3 d, 9.3 d) — Zaire EBOV proxy (WHO 2014 NEJM); "
            "τ = 7 days; prior Gamma(α₀ = 1, β₀ = 0.2) → prior mean R = 5 "
            "(EpiEstim default, weakly informative). "
            "**R_t = 1** separates a growing outbreak (above) from a declining one (below)."
        )

    st.stop()


# -------------------------------------------------------------------------
# LEFT — Input panel  (Step 1)
# -------------------------------------------------------------------------
with left:
    st.markdown('<div class="panel-title">Input</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-label">Source attribution</div>',
                unsafe_allow_html=True)
    source_mode = st.radio(
        "Source / DOI",
        ["Single source for whole table", "Per-row sources"],
        horizontal=True,
        label_visibility="collapsed",
        key="source_mode",
    )
    whole_table_source = ""
    if source_mode == "Single source for whole table":
        whole_table_source = st.text_input(
            "Source URL or DOI",
            value=st.session_state.get(
                "whole_table_source",
                "Africa CDC — Bundibugyo Virus Disease Outbreak Situational "
                "Report"),
            placeholder="https://www.who.int/.../2026-DON605",
            key="whole_table_source",
        )

    st.markdown('<div class="section-label">Data</div>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        value_type = st.radio(
            "Values are",
            ["Cumulative", "Incidence (daily new)"],
            index=1,
            key="value_type",
        )
    with c2:
        input_method = st.radio(
            "Input method",
            ["Manual entry", "CSV upload"],
            key="input_method",
        )

    per_row_source = source_mode == "Per-row sources"
    is_cumulative = value_type == "Cumulative"

    snapshots = None
    incidence_df = None

    VALUE_COLS = (
        ["cumulative_confirmed", "cumulative_suspected", "cumulative_deaths"]
        if is_cumulative
        else ["new_confirmed", "new_suspected", "new_deaths"]
    )
    VALUE_LABELS = {
        "cumulative_confirmed": "Cum. confirmed",
        "cumulative_suspected": "Cum. suspected",
        "cumulative_deaths": "Cum. deaths",
        "new_confirmed": "New confirmed",
        "new_suspected": "New suspected",
        "new_deaths": "New deaths",
    }

    if input_method == "Manual entry":
        _data_src = (
            "DRC DON602/603/605 cumulative"
            if is_cumulative
            else "African CDC Bundibugyo Virus Disease Outbreak Situational "
                 "Report — daily incidence, 01 May–05 Jun 2026"
        )
        st.markdown(
            f'<div style="margin:0.3rem 0 0.6rem 0; padding:0.5rem 0.8rem; '
            f'background:#fff8e1; border:1px solid #f0d678; border-radius:5px; '
            f'font-size:0.82rem; color:#6b5a14;">'
            f'<b>Data shown below</b> ({_data_src}). '
            f'Edit the rows, add your own, or delete to start fresh before running.'
            f'</div>',
            unsafe_allow_html=True,
        )
        if is_cumulative:
            default_cols = {
                "date": [date(2026, 5, 16), date(2026, 5, 21),
                         date(2026, 5, 29)],
                "cumulative_confirmed": [10, 85, 134],
                "cumulative_suspected": [246, 746, 906],
                "cumulative_deaths": [84, 186, 241],
            }
        else:
            # African CDC "Bundibugyo Virus Disease Outbreak Situational
            # Report" — daily incidence, 01 May 2026 → 05 Jun 2026.
            # Dates absent from the report (03–04 Jun) are filled with zeros.
            default_cols = {
                "date": [date(2026, 5, d) for d in range(1, 32)]
                        + [date(2026, 6, d) for d in range(1, 6)],
                "new_confirmed": [
                    1, 1, 2, 0, 1, 0, 2, 2, 0, 2, 2, 1, 5, 3, 7, 20, 9, 20,
                    11, 1, 0, 0, 5, 10, 4, 0, 16, 4, 78, 51, 19,
                    39, 23, 0, 0, 105],
                "new_suspected": [0] * 36,
                "new_deaths": [
                    0, 0, 1, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 2, 1, 2, 3, 1,
                    0, 0, 0, 0, 0, 0, 2, 5, 0, 0, 25, 0,
                    6, 12, 0, 0, 16],
            }
        if per_row_source:
            n_rows = len(default_cols["date"])
            default_cols["source"] = [
                "https://www.who.int/.../2026-DON602",
                "https://www.who.int/.../2026-DON603",
                "https://www.who.int/.../2026-DON605",
            ][:n_rows] + [""] * max(0, n_rows - 3)
        default_df = pd.DataFrame(default_cols)

        column_config = {"date": st.column_config.DateColumn(
            "Date", format="DD MMM YYYY", required=True,
            help="Format: DD Mmm YYYY (e.g. 01 May 2026)")}
        for col in VALUE_COLS:
            column_config[col] = st.column_config.NumberColumn(
                VALUE_LABELS[col], min_value=0, step=1, required=True
            )
        if per_row_source:
            column_config["source"] = st.column_config.TextColumn(
                "Source / DOI", help="Per-row source URL or DOI"
            )

        edited = st.data_editor(
            default_df,
            num_rows="dynamic",
            column_config=column_config,
            use_container_width=True,
            key=f"editor_{value_type}_{per_row_source}",
            height=260,
        )
        if is_cumulative:
            snapshots = edited
        else:
            incidence_df = edited

        st.caption(
            "Linear interpolation between snapshots. The first snapshot's "
            "cumulative count is back-extrapolated at the first inter-snapshot "
            "rate so the sum of daily new cases equals the final cumulative."
            if is_cumulative
            else "Daily rows used as-is; gappy rows distributed uniformly across the window."
        )
    else:
        required = ["date"] + VALUE_COLS
        optional_note = " Optional: `source`." if per_row_source else ""
        st.caption(
            f"Required columns: `{', '.join(required)}`." + optional_note
            + " **Date format: DD Mmm YYYY** (e.g. 01 May 2026, "
            "29 May 2026). YYYY-MM-DD also accepted. Numeric formats "
            "like 01/02/2026 are rejected (ambiguous month/day)."
        )
        uploaded = st.file_uploader("Upload CSV", type=["csv"])
        if uploaded is not None:
            try:
                # Read date column as raw strings — our strict parser
                # validates the format below.
                raw_df = pd.read_csv(uploaded, dtype={"date": str})
            except Exception as e:
                st.error(f"Could not read CSV: {e}")
                raw_df = None
            if raw_df is not None:
                missing = [c for c in required if c not in raw_df.columns]
                if missing:
                    st.error(f"CSV missing columns: {missing}")
                else:
                    raw_df = _parse_date_column_strict(
                        raw_df, "date", where="the uploaded CSV")
                    raw_df = raw_df.sort_values("date").reset_index(drop=True)
                    st.success(f"Loaded {len(raw_df)} rows.")
                    preview = raw_df.head(8).copy()
                    preview["date"] = preview["date"].dt.strftime("%d %b %Y")
                    st.dataframe(preview, use_container_width=True, height=220)
                    if is_cumulative:
                        snapshots = raw_df
                    else:
                        incidence_df = raw_df

    # ------------------------------------------------------------------
    # Standalone smoothing (independent of CFR) — a moving average over the
    # daily incidence before it feeds the model.
    # ------------------------------------------------------------------
    st.markdown('<div class="section-label">Smoothing</div>',
                unsafe_allow_html=True)
    smooth_label = st.selectbox(
        "Moving-average smoothing",
        ["None", "3-Day", "5-Day", "7-Day"],
        index=["None", "3-Day", "5-Day", "7-Day"].index(
            st.session_state.get("smooth_label", "None")),
        key="smooth_label",
        help="Trailing simple moving average applied to the daily new "
             "confirmed / suspected / deaths series before it feeds Steps "
             "2–4. Independent of the CFR backcalculation — works with it "
             "off. Damps day-to-day reporting noise; raw values are kept "
             "for reference and shown faintly on the chart.",
    )
    smooth_window = {"None": 0, "3-Day": 3,
                     "5-Day": 5, "7-Day": 7}[smooth_label]

    # ------------------------------------------------------------------
    # CFR backcalculation controls (Imperial / Garske et al.)
    # ------------------------------------------------------------------
    st.markdown('<div class="section-label">CFR backcalculation</div>',
                unsafe_allow_html=True)
    cfr_enabled = st.checkbox(
        "Enable CFR Backcalculation Method",
        value=st.session_state.get("cfr_enabled", False),
        key="cfr_enabled",
        help="Estimates the true number of cases from observed deaths, "
             "applying a phase-bias correction (1 + r/β)^α to compensate "
             "for the lag between case onset and death during a growing "
             "epidemic.",
    )

    cfr_pct = 30.0
    cfr_sma = 0
    cfr_role = "Total cases"
    cfr_t2 = 14.0
    cfr_alpha = 4.42
    cfr_beta = 0.388

    if cfr_enabled:
        cfr_choice = st.radio(
            "CFR scenario",
            ["24%", "30%", "40%", "Custom"],
            index=["24%", "30%", "40%", "Custom"].index(
                st.session_state.get("cfr_choice", "30%")),
            key="cfr_choice",
            horizontal=True,
        )
        if cfr_choice == "Custom":
            cfr_pct = st.number_input(
                "Custom CFR (%)",
                min_value=0.1, max_value=100.0,
                value=float(st.session_state.get("cfr_pct_custom", 30.0)),
                step=0.5, key="cfr_pct_custom",
            )
        else:
            cfr_pct = float(cfr_choice.rstrip("%"))
        # (CFR's own SMA removed — the standalone Smoothing control above also
        # smooths the CFR-estimated series, so a second SMA here was redundant.)

        cfr_role = st.radio(
            "Treat estimated cases as",
            ["Total cases", "Confirmed", "Suspected"],
            index=["Total cases", "Confirmed", "Suspected"].index(
                st.session_state.get("cfr_role", "Total cases")),
            horizontal=True,
            key="cfr_role",
            help="How the CFR-estimated cases flow into Steps 2–4. "
                 "Total cases: a single combined incidence series. "
                 "Confirmed/Suspected: replaces that column downstream.",
        )

        with st.expander("Advanced Epidemiological Parameters", expanded=False):
            cfr_t2 = st.number_input(
                "Doubling time t₂ (days)",
                min_value=0.5, max_value=120.0,
                value=float(st.session_state.get("cfr_t2", 14.0)),
                step=0.5, key="cfr_t2",
                help="Epidemic doubling time. Growth rate r = ln(2)/t₂.",
            )
            cfr_alpha = st.number_input(
                "Shape α",
                min_value=0.01, max_value=50.0,
                value=float(st.session_state.get("cfr_alpha", 4.42)),
                step=0.01, key="cfr_alpha",
                help="Shape parameter of the Gamma onset-to-death distribution.",
            )
            cfr_beta = st.number_input(
                "Rate β",
                min_value=0.001, max_value=10.0,
                value=float(st.session_state.get("cfr_beta", 0.388)),
                step=0.001, format="%.3f", key="cfr_beta",
                help="Rate parameter of the Gamma onset-to-death distribution.",
            )

    # ------------------------------------------------------------------
    # Estimate suspected from confirmed using the confirmed-to-suspected ratio.
    # Estimated Suspected = Confirmed ÷ ratio.
    # ------------------------------------------------------------------
    st.markdown('<div class="section-label">Estimate suspected (from confirmed)</div>',
                unsafe_allow_html=True)
    suspest_enabled = st.checkbox(
        "Estimate Suspected from Confirmed",
        value=st.session_state.get("suspest_enabled", False),
        key="suspest_enabled",
        help="Derives an estimated suspected-case series as Confirmed ÷ ratio, "
             "where ratio = confirmed ÷ suspected. Feeds Step 2 as the "
             "'Estimated' suspected source.",
    )
    suspest_tpr = 0.15
    suspest_basis = "Actual Confirmed"
    if suspest_enabled:
        se_a, se_b = st.columns(2)
        with se_a:
            suspest_tpr = st.number_input(
                "Confirmed-to-suspected ratio",
                min_value=0.01, max_value=1.0,
                value=float(st.session_state.get("suspest_tpr", 0.15)),
                step=0.01, format="%.3f", key="suspest_tpr",
                help="Fraction of suspected cases that are confirmed "
                     "(confirmed ÷ suspected). Estimated Suspected = "
                     "Confirmed ÷ this ratio.",
            )
        with se_b:
            basis_opts = (["Actual Confirmed", "Estimated Confirmed (CFR)"]
                          if cfr_enabled else ["Actual Confirmed"])
            if st.session_state.get("suspest_basis") not in basis_opts:
                st.session_state.pop("suspest_basis", None)
            suspest_basis = st.selectbox(
                "Calculate from",
                basis_opts,
                index=0,
                key="suspest_basis",
                help="Which confirmed series to divide by the ratio. 'Estimated "
                     "Confirmed (CFR)' needs the CFR backcalculation enabled.",
            )

    generate = st.button("Generate daily series", type="primary",
                         use_container_width=True)

    if generate:
        if snapshots is None and incidence_df is None:
            st.warning("Please provide input data first.")
        else:
            built_series = None
            if snapshots is not None:
                if len(snapshots) < 2:
                    st.warning("Need at least 2 cumulative snapshot rows.")
                else:
                    snaps = attach_source(snapshots.copy(), source_mode,
                                          whole_table_source)
                    built_series = interpolate_from_cumulative(snaps)
                    st.session_state["result_chart_snaps"] = snaps
                    # Snapshot mode always interpolates between cumulative points.
                    st.session_state["result_interpolated"] = True
            else:
                inc = attach_source(incidence_df.copy(), source_mode,
                                    whole_table_source)
                built_series = expand_incidence(inc)
                # Convert dates BEFORE sorting — sorting string dates is
                # lexicographic ("10-Jan-2026" before "2-Feb-2026") and
                # corrupts the x-axis on the Step 1 chart.
                inc_sorted = _parse_date_column_strict(
                    inc, "date", where="the incidence file")
                inc_sorted = inc_sorted.sort_values("date")
                # Daily-spaced incidence is used as-is; only gappy incidence is
                # linearly filled by expand_incidence. Flag the title accordingly.
                _gaps = inc_sorted["date"].diff().dt.days.dropna()
                st.session_state["result_interpolated"] = (
                    len(inc_sorted) > 1 and not bool((_gaps == 1).all()))
                chart_snaps = inc_sorted.assign(
                    cumulative_confirmed=inc_sorted["new_confirmed"].cumsum(),
                    cumulative_suspected=inc_sorted["new_suspected"].cumsum(),
                    cumulative_deaths=inc_sorted["new_deaths"].cumsum(),
                )
                st.session_state["result_chart_snaps"] = chart_snaps

            if built_series is not None and len(built_series) > 0:
                if cfr_enabled:
                    built_series = cfr_backcalculate(
                        built_series, cfr_pct=cfr_pct,
                        doubling_time=cfr_t2, alpha=cfr_alpha, beta=cfr_beta,
                        sma_window=cfr_sma,
                    )
                    st.session_state["cfr_active"] = True
                    st.session_state["cfr_role_active"] = cfr_role
                    st.session_state["cfr_pct_active"] = cfr_pct
                    st.session_state["cfr_sma_active"] = cfr_sma
                else:
                    st.session_state["cfr_active"] = False
                    st.session_state.pop("cfr_role_active", None)

                # Estimate Suspected from Confirmed via ratio. Runs after CFR so
                # the Estimated-Confirmed basis is available, and before
                # smoothing so a smoothed copy follows the usual pattern.
                if suspest_enabled:
                    tpr_val = max(float(suspest_tpr), 1e-9)
                    if (suspest_basis.startswith("Estimated")
                            and "new_cfr_estimated" in built_series.columns):
                        basis_col = "new_cfr_estimated"
                    else:
                        basis_col = "new_confirmed"
                    built_series["new_suspected_estimated"] = (
                        (built_series[basis_col].astype(float) / tpr_val)
                        .round().astype(int))
                    built_series["cumulative_suspected_estimated"] = (
                        built_series["new_suspected_estimated"].cumsum())
                    st.session_state["suspest_active"] = True
                    st.session_state["suspest_tpr_active"] = tpr_val
                    st.session_state["suspest_basis_active"] = basis_col
                else:
                    st.session_state["suspest_active"] = False
                    st.session_state.pop("suspest_tpr_active", None)
                    st.session_state.pop("suspest_basis_active", None)

                # Standalone smoothing — applied after CFR so it also damps
                # the CFR-estimated base columns; works fine with CFR off.
                if smooth_window >= 2:
                    built_series = smooth_incidence(built_series, smooth_window)
                    st.session_state["smooth_active"] = True
                    st.session_state["smooth_window_active"] = smooth_window
                else:
                    st.session_state["smooth_active"] = False
                    st.session_state.pop("smooth_window_active", None)

                st.session_state["result_series"] = built_series

# -------------------------------------------------------------------------
# RIGHT — Output panel
# -------------------------------------------------------------------------
with right:
    st.markdown('<div class="freeze-right"></div>', unsafe_allow_html=True)
    st.markdown('<div class="panel-title">Output</div>', unsafe_allow_html=True)
    series = st.session_state.get("result_series")
    chart_snaps = st.session_state.get("result_chart_snaps")

    if series is None or len(series) == 0:
        st.markdown(
            '<div class="placeholder-card">'
            'Configure inputs on the left and click <b>Generate daily series</b> '
            'to see the daily incidence series, cumulative trends, and the '
            'model-ready table.'
            '</div>',
            unsafe_allow_html=True,
        )
    else:
        # --- Compact summary dashboard above the charts ---
        def _raw_sum(col):
            rc = f"{col}_raw"
            src = series[rc] if rc in series.columns else series.get(col)
            return float(src.sum()) if src is not None else 0.0

        _dates = pd.to_datetime(series["date"])
        _conf_src = (series["new_confirmed_raw"]
                     if "new_confirmed_raw" in series.columns
                     else series["new_confirmed"])
        tot_conf = _raw_sum("new_confirmed")
        tot_susp = _raw_sum("new_suspected")
        tot_death = _raw_sum("new_deaths")
        cfr_naive = (tot_death / tot_conf * 100) if tot_conf > 0 else 0.0

        chips = [
            ("Days", f"{len(series)}", "#1f4e79"),
            ("Confirmed", f"{tot_conf:,.0f}", "#4682B4"),
            ("Suspected", f"{tot_susp:,.0f}", "#FF8C00"),
            ("Deaths", f"{tot_death:,.0f}", "#B22222"),
            ("Peak conf./day", f"{float(_conf_src.max()):,.0f}", "#4682B4"),
            ("Naive CFR", f"{cfr_naive:.1f}%", "#B22222"),
        ]
        if "cumulative_cfr_estimated" in series.columns:
            chips.append(("Est. cases (CFR)",
                          f"{float(series['cumulative_cfr_estimated'].iloc[-1]):,.0f}",
                          "#6a1b9a"))
        if "cumulative_suspected_estimated" in series.columns:
            chips.append(("Est. suspected",
                          f"{float(series['cumulative_suspected_estimated'].iloc[-1]):,.0f}",
                          "#b8860b"))

        _chip_html = ('<div style="display:flex; flex-wrap:wrap; gap:0.4rem; '
                      'margin:0.1rem 0 0.5rem 0;">')
        for lbl, val, colour in chips:
            _chip_html += (
                f'<div style="flex:1 1 90px; min-width:84px; background:#f6f8fb; '
                f'border:1px solid #e3e7ed; border-left:3px solid {colour}; '
                f'border-radius:7px; padding:0.35rem 0.55rem;">'
                f'<div style="font-size:0.62rem; color:#6b7280; '
                f'text-transform:uppercase; letter-spacing:0.04em;">{lbl}</div>'
                f'<div style="font-size:1.0rem; font-weight:600; color:{colour};">'
                f'{val}</div></div>')
        _chip_html += '</div>'
        st.markdown(_chip_html, unsafe_allow_html=True)
        st.caption(f"{_dates.min():%d %b %Y} → {_dates.max():%d %b %Y}")

        tab1, tab2, tab3 = st.tabs(["Daily new", "Cumulative", "Table"])
        with tab1:
            _daily_selected = daily_selected_from_state(series)
            _interp = st.session_state.get("result_interpolated", True)
            _fig_daily = daily_chart(series, _daily_selected, _interp)
            st.session_state["chart_daily"] = _fig_daily
            st.plotly_chart(_fig_daily, use_container_width=True, key="daily")
            render_daily_legend(series)
        with tab2:
            _cum_selected = cum_selected_from_state(series, chart_snaps)
            _interp = st.session_state.get("result_interpolated", True)
            _fig_cum = cumulative_chart(series, chart_snaps, _cum_selected,
                                        _interp)
            st.session_state["chart_cumulative"] = _fig_cum
            st.plotly_chart(_fig_cum, use_container_width=True, key="cum")
            render_cum_legend(series, chart_snaps)
        with tab3:
            smooth_raw_cols = [c for c in
                               ["new_confirmed_raw", "new_suspected_raw",
                                "new_deaths_raw"] if c in series.columns]
            smoothing_on = bool(smooth_raw_cols)
            extra_cfr_cols = [c for c in
                              ["new_cfr_estimated", "cumulative_cfr_estimated",
                               "new_cfr_estimated_sma", "new_deaths_sma",
                               "new_suspected_estimated"]
                              if c in series.columns]
            display = series[TABLE_COLS + smooth_raw_cols + extra_cfr_cols].copy()
            display["date"] = pd.to_datetime(display["date"]).dt.strftime("%d %b %Y")
            for col in ["new_confirmed", "new_suspected", "new_deaths"]:
                if smoothing_on:
                    display[col] = display[col].round(1)
                else:
                    display[col] = display[col].round(0).astype(int)
            for col in smooth_raw_cols:
                display[col] = display[col].round(0).astype(int)
            for col in extra_cfr_cols:
                display[col] = display[col].round(1)
            rename_map = {
                "date": "Date",
                "new_confirmed": "New confirmed (smoothed)" if smoothing_on
                                 else "New confirmed",
                "new_suspected": "New suspected (smoothed)" if smoothing_on
                                 else "New suspected",
                "new_deaths": "New deaths (smoothed)" if smoothing_on
                              else "New deaths",
                "new_confirmed_raw": "New confirmed (raw)",
                "new_suspected_raw": "New suspected (raw)",
                "new_deaths_raw": "New deaths (raw)",
                "new_cfr_estimated": "Est. new cases (CFR)",
                "cumulative_cfr_estimated": "Est. cumulative (CFR)",
                "new_cfr_estimated_sma": "Est. new cases (CFR, SMA)",
                "new_deaths_sma": "New deaths (SMA)",
                "new_suspected_estimated": "Est. suspected",
            }
            display = display.rename(columns=rename_map)
            st.dataframe(display, use_container_width=True, height=440,
                         hide_index=True)

        csv_out = series.copy()
        csv_out["date"] = pd.to_datetime(csv_out["date"]).dt.strftime("%Y-%m-%d")

        dl_col, next_col = st.columns([1, 1])
        with dl_col:
            slug = _slug(st.session_state.get("scenario_name", ""))
            cfr_tag = ""
            if st.session_state.get("cfr_active"):
                cfr_tag = f"__cfr{int(st.session_state.get('cfr_pct_active', 0))}"
            st.download_button(
                "Download daily_incidence.csv",
                csv_out.to_csv(index=False).encode("utf-8"),
                f"{slug}__daily_incidence{cfr_tag}.csv", "text/csv",
                use_container_width=True,
            )
        with next_col:
            if st.button("Next: Estimate R_t  →", type="primary",
                         use_container_width=True, key="goto_rt"):
                st.session_state["step"] = "rt"
                st.rerun()
